import os
import re
import logging
import sqlite3
import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
import pandas as pd
import subprocess
import traceback
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, date
from PIL import Image, ImageTk

# Set up logging
logging.basicConfig(
    filename=os.path.join(os.path.dirname(__file__), 'project_booking_app.log'),
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s %(message)s'
)

# Set customtkinter appearance - Saipem theme
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")  # We'll override with custom Saipem colors

class ProjectBookingApp:
    """Project Booking & Resource Allocation Application"""
    
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("Project Booking & Resource Allocation System")
        self.root.geometry("1600x1000")  # Increased window size for better table visibility
        
        # Database connection
        self.db_path = "workload.db"
        self.init_extended_database()
        
        # Variables for dropdowns
        self.selected_technical_unit = tk.StringVar()
        self.selected_project = tk.StringVar()
        self.selected_employee = tk.StringVar()
        
        # Store mapping of dropdown display names to IDs
        self.technical_unit_map = {}
        self.project_map = {}
        self.employee_map = {}
        
        # Data storage
        self.technical_units = []
        self.projects = []
        self.employees = []
        
        # Store selected rows for multi-select deletion
        self.selected_rows = set()
        
        # Filter and sort functionality - like Fabsi app
        self.active_column_filters = {}
        self.filter_vars = {}
        self.current_sort_column = None
        self.current_sort_ascending = True
        self.filter_popup = None
        self.filter_window = None
        self.filter_menus = {}
        self.current_filter_values = []
        self.current_column_filter_vars = {}
        self.search_var = None
        self.select_all_var = None
        
        # Auto-refresh timer (refresh every 30 seconds) - DISABLED BY DEFAULT
        self.auto_refresh_enabled = False  # FIXED: Disabled to prevent filter clearing
        # self.schedule_auto_refresh()  # COMMENTED OUT - user can enable manually
        self.current_bookings = []
        
        # Main data DataFrame for filtering - like Fabsi app
        self.df = pd.DataFrame()
        self.original_df = pd.DataFrame()
        
        self.setup_ui()
        self.load_data()
        
    def init_extended_database(self):
        """Initialize database schema for project booking - using existing tables only"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Check if tables exist - do not create new ones, use existing structure
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='service'")
            if not cursor.fetchone():
                print("Warning: service table not found")
            
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='employee'")
            if not cursor.fetchone():
                print("Warning: employee table not found")
            
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='project_bookings'")
            if not cursor.fetchone():
                # Only create project_bookings if it doesn't exist, using existing table references
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS project_bookings (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        employee_id INTEGER,
                        technical_unit_id INTEGER,
                        project_id INTEGER,
                        service_id INTEGER,
                        actual_hours DECIMAL(10,2),
                        hourly_rate DECIMAL(10,2),
                        total_cost DECIMAL(10,2),
                        booking_status VARCHAR(50) DEFAULT 'Pending',
                        booking_date DATE,
                        start_date DATE,
                        end_date DATE,
                        created_by VARCHAR(100),
                        approved_by VARCHAR(100),
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (employee_id) REFERENCES employee (id),
                        FOREIGN KEY (technical_unit_id) REFERENCES technical_unit (id),
                        FOREIGN KEY (project_id) REFERENCES project (id),
                        FOREIGN KEY (service_id) REFERENCES service (id)
                    )
                ''')
            
            conn.commit()
            conn.close()
            print("Database schema checked successfully - using existing tables")
            
        except Exception as e:
            print(f"Error checking database: {e}")
            logging.error(f"Database check error: {e}")
    
    def load_logo_image(self, image_path, width, height):
        """Load and resize logo image for display"""
        try:
            if os.path.exists(image_path):
                # Load and resize the image
                pil_image = Image.open(image_path)
                # Convert to RGBA if not already
                if pil_image.mode != 'RGBA':
                    pil_image = pil_image.convert('RGBA')
                # Resize maintaining aspect ratio
                pil_image.thumbnail((width, height), Image.Resampling.LANCZOS)
                # Convert to CTkImage
                return ctk.CTkImage(light_image=pil_image, dark_image=pil_image, size=(width, height))
            else:
                print(f"Logo image not found: {image_path}")
                return None
        except Exception as e:
            print(f"Error loading logo image {image_path}: {e}")
            return None
    
    def setup_ui(self):
        """Setup the main user interface - Dropdowns + Employee Data Table Only"""
        # Add Saipem header first
        self.setup_header()
        
        # Main container
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.pack(fill="both", expand=True, padx=3, pady=3)  # Minimal padding for maximum space
        
        # Keep Selection Panel for filtering
        self.setup_selection_panel()
        
        # Show only Employee Data Table with 29 columns
        self.setup_employee_data_table_only()
        
    def setup_header(self):
        """Setup application header with Saipem and FABSI branding"""
        # Create header frame with exact Saipem colors from Fabsi app
        header_frame = ctk.CTkFrame(self.root, height=40, corner_radius=10)
        header_frame.pack(fill='x', padx=10, pady=(2, 3))
        header_frame.pack_propagate(False)

        # Left logo (Saipem) - same size as Fabsi app
        left_logo_frame = ctk.CTkFrame(header_frame, width=80, height=35, corner_radius=5)
        left_logo_frame.pack(side='left', padx=5)
        left_logo_frame.pack_propagate(False)
        
        # Load Saipem logo
        saipem_logo_path = os.path.join(os.path.dirname(__file__), 'photos', 'saipem_logo.png')
        saipem_logo_image = self.load_logo_image(saipem_logo_path, 75, 30)
        
        if saipem_logo_image:
            left_logo_label = ctk.CTkLabel(left_logo_frame, image=saipem_logo_image, text="")
        else:
            left_logo_label = ctk.CTkLabel(left_logo_frame, text="SAIPEM", 
                                         font=("Arial", 9), text_color="#003d52")
        left_logo_label.pack(expand=True)

        # Title in center - same style as Fabsi app
        title_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_frame.pack(side='left', expand=True)
        title_label = ctk.CTkLabel(
            title_frame, 
            text="Project Booking & Resource Allocation System",
            font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
            text_color="#003d52"
        )
        title_label.pack(expand=True)

        # Right logo (FABSI) - same as Fabsi app
        right_logo_frame = ctk.CTkFrame(header_frame, width=60, height=35, corner_radius=5)
        right_logo_frame.pack(side='right', padx=5)
        right_logo_frame.pack_propagate(False)
        
        # Load FABSI logo
        fabsi_logo_path = os.path.join(os.path.dirname(__file__), 'photos', 'fabsi_logo.png')
        fabsi_logo_image = self.load_logo_image(fabsi_logo_path, 55, 30)
        
        if fabsi_logo_image:
            right_logo_label = ctk.CTkLabel(right_logo_frame, image=fabsi_logo_image, text="")
        else:
            right_logo_label = ctk.CTkLabel(right_logo_frame, text="FABSI", 
                                          font=("Arial", 7), text_color="#ef8827")
        right_logo_label.pack(expand=True)
        
        # Add separator - same as Fabsi app
        separator_frame = ctk.CTkFrame(self.root, height=1, fg_color="#22505f")
        separator_frame.pack(fill='x', padx=10, pady=(0, 3))
    
    def setup_selection_panel(self):
        """Setup the three-level selection panel - FOR ADDING DATA"""
        selection_frame = ctk.CTkFrame(self.main_frame)
        selection_frame.pack(fill="x", padx=3, pady=(3, 5))  # Minimal padding
        
        selection_label = ctk.CTkLabel(
            selection_frame, 
            text="Add Service Data",
            font=ctk.CTkFont(size=14, weight="bold")  # Clear purpose
        )
        selection_label.pack(pady=(3, 2))  # Minimal padding
        
        # Dropdown container
        dropdown_frame = ctk.CTkFrame(selection_frame)
        dropdown_frame.pack(fill="x", padx=5, pady=2)  # Minimal padding
        
        # Technical Unit dropdown
        tech_unit_frame = ctk.CTkFrame(dropdown_frame)
        tech_unit_frame.pack(side="left", fill="x", expand=True, padx=3)
        
        ctk.CTkLabel(tech_unit_frame, text="Technical Unit:", font=ctk.CTkFont(size=11)).pack(pady=(2, 0))
        self.tech_unit_dropdown = ctk.CTkComboBox(
            tech_unit_frame,
            variable=self.selected_technical_unit,
            command=self.on_technical_unit_change,
            width=200
        )
        self.tech_unit_dropdown.pack(pady=3, padx=3)
        
        # Project dropdown
        project_frame = ctk.CTkFrame(dropdown_frame)
        project_frame.pack(side="left", fill="x", expand=True, padx=3)
        
        ctk.CTkLabel(project_frame, text="Project:", font=ctk.CTkFont(size=11)).pack(pady=(2, 0))
        self.project_dropdown = ctk.CTkComboBox(
            project_frame,
            variable=self.selected_project,
            command=self.on_project_change,
            width=200
        )
        self.project_dropdown.pack(pady=3, padx=3)
        
        # Employee dropdown
        employee_frame = ctk.CTkFrame(dropdown_frame)
        employee_frame.pack(side="left", fill="x", expand=True, padx=3)
        
        ctk.CTkLabel(employee_frame, text="Employee:", font=ctk.CTkFont(size=11)).pack(pady=(2, 0))
        self.employee_dropdown = ctk.CTkComboBox(
            employee_frame,
            variable=self.selected_employee,
            command=self.on_employee_change,
            width=200
        )
        self.employee_dropdown.pack(pady=3, padx=3)
        
        # Action buttons under dropdowns
        button_frame = ctk.CTkFrame(selection_frame)
        button_frame.pack(fill="x", padx=5, pady=3)
        
        self.refresh_emp_data_btn = ctk.CTkButton(
            button_frame, 
            text="Smart Refresh", 
            command=self.smart_refresh,
            width=120,
            fg_color="#003d52",
            hover_color="#255c7b"
        )
        self.refresh_emp_data_btn.pack(side="left", padx=3)
        
        self.delete_selected_btn = ctk.CTkButton(
            button_frame, 
            text="🗑️ Delete Selected", 
            command=self.delete_selected_rows,
            width=140,
            fg_color="#003d52",
            hover_color="#255c7b"
        )
        self.delete_selected_btn.pack(side="left", padx=3)
        
        self.select_all_btn = ctk.CTkButton(
            button_frame, 
            text="✅ Select All", 
            command=self.select_all_rows,
            width=120,
            fg_color="#003d52",
            hover_color="#255c7b"
        )
        self.select_all_btn.pack(side="left", padx=3)
        
        self.deselect_all_btn = ctk.CTkButton(
            button_frame, 
            text="❌ Deselect All", 
            command=self.deselect_all_rows,
            width=120,
            fg_color="#003d52",
            hover_color="#255c7b"
        )
        self.deselect_all_btn.pack(side="left", padx=3)
        
        self.clear_filters_btn = ctk.CTkButton(
            button_frame, 
            text="🔄 Clear Filters", 
            command=self.clear_all_filters,
            width=120,
            fg_color="#003d52",
            hover_color="#255c7b"
        )
        self.clear_filters_btn.pack(side="left", padx=3)
        
        # Export buttons frame - same as Fabsi app
        export_frame = ctk.CTkFrame(button_frame, corner_radius=8)
        export_frame.pack(side='left', padx=10)
        
        ctk.CTkLabel(export_frame, text="Export Data:", 
                    font=ctk.CTkFont(family='Arial', size=11, weight='bold')).pack(side='left', padx=5)
        
        ctk.CTkButton(export_frame, text="📊 Excel", 
                     command=self.export_report,
                     fg_color="#003d52", hover_color="#255c7b",
                     width=80).pack(side='left', padx=5, pady=5)
        
        self.import_btn = ctk.CTkButton(
            button_frame, 
            text="📁 Import Excel", 
            command=self.import_excel_data,
            width=120,
            fg_color="#003d52",
            hover_color="#255c7b"
        )
        self.import_btn.pack(side="left", padx=3)
    
    def setup_employee_data_table_only(self):
        """Setup only the employee data table to maximize space"""
        # Updated title for unified table approach
        title_label = ctk.CTkLabel(
            self.main_frame, 
            text="Employee Booking Data (Unified Tables)",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        title_label.pack(pady=(3, 2))
        
        # Create header frame for filter buttons
        self.header_frame = ctk.CTkFrame(self.main_frame)
        self.header_frame.pack(fill="x", padx=2, pady=(0, 2))
        
        # Add filter control buttons to header frame
        ctk.CTkButton(
            self.header_frame, 
            text="🔄 Clear All Filters",
            command=self.reset_filters,
            width=140,
            fg_color="#003d52",
            hover_color="#255c7b"
        ).pack(side="left", padx=5, pady=5)
        
        ctk.CTkLabel(
            self.header_frame,
            text="📋 Click column headers with ▼ to filter data (Excel-like filtering)",
            font=ctk.CTkFont(size=11)
        ).pack(side="left", padx=10, pady=5)
        
        # Employee data grid with booking information (simplified view)
        self.setup_employee_data_grid_maximized(self.main_frame)
    
    def setup_employee_data_grid_maximized(self, parent):
        """Setup treeview for employee data with booking information - MAXIMIZED"""
        emp_tree_frame = ctk.CTkFrame(parent)
        emp_tree_frame.pack(fill="both", expand=True, padx=2, pady=2)  # Maximum space
        
        # Create treeview with scrollbars
        emp_tree_container = tk.Frame(emp_tree_frame)
        emp_tree_container.pack(fill="both", expand=True, padx=1, pady=1)  # Maximum space
        
        # Define simplified columns for unified view
        emp_columns = (
            "Select", "ID", "Employee Name", "Total Bookings", "Est. Hours", "Actual Hours", 
            "Total Cost", "Projects", "Technical Units", "Status", "Last Booking"
        )
        
        # Maximum height for the table - most of the screen
        self.employee_tree = ttk.Treeview(emp_tree_container, columns=emp_columns, show="headings", height=32)
        
        # Configure column headings and widths for simplified view with Excel-like filters
        column_widths = {
            "Select": 60, "ID": 50, "Employee Name": 200, "Total Bookings": 100, 
            "Est. Hours": 100, "Actual Hours": 100, "Total Cost": 120,
            "Projects": 150, "Technical Units": 150, "Status": 100, 
            "Last Booking": 120
        }
        
        # Configure headers with Excel-like filter arrows (similar to Fabsi app)
        from functools import partial
        for col in emp_columns:
            # Add filter arrow only for non-Select and non-ID columns
            if col not in ["Select", "ID"]:
                header_text = f"{col} ▼"
                # Use functools.partial to avoid lambda closure issues
                self.employee_tree.heading(col, text=header_text, command=partial(self.show_filter_menu, col))
            else:
                self.employee_tree.heading(col, text=col)
            
            width = column_widths.get(col, 100)
            anchor = 'w' if col in ["Employee Name", "Projects", "Technical Units", "Status"] else 'center'
            self.employee_tree.column(col, width=width, anchor=anchor, minwidth=80)
        
        # Scrollbars for employee data grid - PROPERLY POSITIONED
        emp_v_scrollbar = ttk.Scrollbar(emp_tree_container, orient="vertical", command=self.employee_tree.yview)
        emp_h_scrollbar = ttk.Scrollbar(emp_tree_container, orient="horizontal", command=self.employee_tree.xview)
        
        self.employee_tree.configure(yscrollcommand=emp_v_scrollbar.set, xscrollcommand=emp_h_scrollbar.set)
        
        # Grid layout for proper scrollbar positioning
        self.employee_tree.grid(row=0, column=0, sticky="nsew")
        emp_v_scrollbar.grid(row=0, column=1, sticky="ns")
        emp_h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights
        emp_tree_container.grid_rowconfigure(0, weight=1)
        emp_tree_container.grid_columnconfigure(0, weight=1)
        
        # Excel-like editing setup
        self.edit_mode = "double"  # "single" or "double" click
        
        # Bind double-click event for editing (Excel-like behavior)
        self.employee_tree.bind('<Double-1>', self.on_simplified_cell_edit)
        
        # Add keyboard shortcut for editing (F2 like Excel)
        self.employee_tree.bind('<F2>', self.on_keyboard_edit)
        
        # Add visual feedback for selected cells
        self.employee_tree.bind('<Button-1>', self.on_cell_select)
        
        # Add checkbox selection functionality
        self.employee_tree.bind('<Button-1>', self.toggle_row_selection, add='+')
        
        # Load employee data initially
        self.load_employee_data_grid()
    
    def setup_employee_details_panel(self):
        """Setup employee details display panel"""
        details_frame = ctk.CTkFrame(self.main_frame)
        details_frame.pack(fill="x", padx=5, pady=(0, 5))  # Minimal padding
        
        details_label = ctk.CTkLabel(
            details_frame, 
            text="Employee Details",
            font=ctk.CTkFont(size=14, weight="bold")  # Smaller font
        )
        details_label.pack(pady=(3, 2))  # Minimal padding
        
        # Details container
        self.details_container = ctk.CTkFrame(details_frame)
        self.details_container.pack(fill="x", padx=10, pady=3)  # Minimal padding
        
        # Employee info will be populated here
        self.employee_info_label = ctk.CTkLabel(
            self.details_container, 
            text="Select an employee to view details",
            font=ctk.CTkFont(size=10)  # Smaller font
        )
        self.employee_info_label.pack(pady=3)  # Minimal padding
    
    def setup_service_assignment_panel(self):
        """Setup service assignment panel"""
        service_frame = ctk.CTkFrame(self.main_frame)
        service_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))  # More space for table
        
        # Create tabview for different data views
        self.tabview = ctk.CTkTabview(service_frame)
        self.tabview.pack(fill="both", expand=True, padx=5, pady=2)  # Minimal padding, more space
        
        # Tab 1: Service Assignments
        self.tabview.add("Service Assignments")
        service_tab = self.tabview.tab("Service Assignments")
        
        service_label = ctk.CTkLabel(
            service_tab, 
            text="Service Assignment & Booking",
            font=ctk.CTkFont(size=14, weight="bold")  # Smaller font
        )
        service_label.pack(pady=(2, 1))  # Minimal padding
        
        # Service list with treeview
        self.setup_service_treeview(service_tab)
        
        # Service actions
        action_frame = ctk.CTkFrame(service_tab)
        action_frame.pack(fill="x", padx=5, pady=2)  # Minimal padding
        
        self.add_service_btn = ctk.CTkButton(
            action_frame, 
            text="Add Service Assignment", 
            command=self.add_service_assignment,
            width=200
        )
        self.add_service_btn.pack(side="left", padx=3)
        
        self.edit_service_btn = ctk.CTkButton(
            action_frame, 
            text="Edit Assignment", 
            command=self.edit_service_assignment,
            width=200
        )
        self.edit_service_btn.pack(side="left", padx=3)
        
        self.delete_service_btn = ctk.CTkButton(
            action_frame, 
            text="Remove Assignment", 
            command=self.delete_service_assignment,
            width=200
        )
        self.delete_service_btn.pack(side="left", padx=3)
        
        # Tab 2: Employee Data (29 Columns)
        self.tabview.add("Employee Data (29 Columns)")
        employee_tab = self.tabview.tab("Employee Data (29 Columns)")
        
        employee_data_label = ctk.CTkLabel(
            employee_tab, 
            text="Complete Employee Data - 29 Columns",
            font=ctk.CTkFont(size=14, weight="bold")  # Smaller font
        )
        employee_data_label.pack(pady=(2, 1))  # Minimal padding
        
        # Employee data grid with all 29 columns
        self.setup_employee_data_grid(employee_tab)
        
        # Employee data actions
        emp_action_frame = ctk.CTkFrame(employee_tab)
        emp_action_frame.pack(fill="x", padx=5, pady=2)  # Minimal padding
        
        self.delete_employee_btn = ctk.CTkButton(
            emp_action_frame, 
            text="Delete Employee", 
            command=self.delete_employee_record,
            width=150
        )
        self.delete_employee_btn.pack(side="left", padx=3)
        
        self.refresh_emp_data_btn = ctk.CTkButton(
            emp_action_frame, 
            text="Refresh Data", 
            command=self.load_employee_data_grid,
            width=150
        )
        self.refresh_emp_data_btn.pack(side="left", padx=3)
    
    def setup_service_treeview(self, parent):
        """Setup treeview for service assignments"""
        tree_frame = ctk.CTkFrame(parent)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=2)  # Minimal padding, more space
        
        # Create treeview with scrollbars
        tree_container = tk.Frame(tree_frame)
        tree_container.pack(fill="both", expand=True, padx=2, pady=2)  # Minimal padding
        
        # Define columns
        columns = ("Service", "Activity", "Est. Hours", "Actual Hours", "Rate", "Cost", "Status", "Period")
        
        self.service_tree = ttk.Treeview(tree_container, columns=columns, show="headings", height=22)  # Increased height more
        
        # Configure column headings and widths
        for col in columns:
            self.service_tree.heading(col, text=col)
            self.service_tree.column(col, width=120, anchor="center")
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.service_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_container, orient="horizontal", command=self.service_tree.xview)
        
        self.service_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack everything
        self.service_tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
    
    def setup_employee_data_grid(self, parent):
        """Setup treeview for complete employee data with all 29 columns"""
        # Initialize filter system
        self.active_column_filters = {}
        self.filter_menus = {}
        
        # Create header frame for filter buttons
        self.header_frame = ctk.CTkFrame(parent, height=30)
        self.header_frame.pack(fill="x", padx=5, pady=(2, 0))
        
        emp_tree_frame = ctk.CTkFrame(parent)
        emp_tree_frame.pack(fill="both", expand=True, padx=5, pady=2)  # Minimal padding, more space
        
        # Create treeview with scrollbars
        emp_tree_container = tk.Frame(emp_tree_frame)
        emp_tree_container.pack(fill="both", expand=True, padx=2, pady=2)  # Minimal padding
        
        # Define all 29 columns
        emp_columns = (
            "ID", "Cost Center", "GHRS ID", "Last Name", "First Name", "Dept. Description",
            "Work Location", "Business Unit", "Tipo", "Tipo Description", "SAP Tipo",
            "SAABU Rate (EUR)", "SAABU Rate (USD)", "Local Agency Rate (USD)", "Unit Rate (USD)",
            "Monthly Hours", "Annual Hours", "Workload 2025_Planned", "Workload 2025_Actual",
            "Remark", "Project", "Item", "Technical Unit", "Activities", "Booking Hours",
            "Booking Cost (Forecast)", "Booking Period", "Booking hours (Accepted by Project)",
            "Booking Period (Accepted by Project)", "Booking hours (Extra)"
        )
        
        # Increased height significantly to show more rows like FABSI app
        self.employee_tree = ttk.Treeview(emp_tree_container, columns=emp_columns, show="headings", height=30)  # Increased from 25 to 30
        
        # Configure column headings and widths with better optimization
        column_widths = {
            "ID": 50, "Cost Center": 100, "GHRS ID": 100, "Last Name": 120, "First Name": 120,
            "Dept. Description": 150, "Work Location": 120, "Business Unit": 120, "Tipo": 80,
            "Tipo Description": 150, "SAP Tipo": 80, "SAABU Rate (EUR)": 120, "SAABU Rate (USD)": 120,
            "Local Agency Rate (USD)": 150, "Unit Rate (USD)": 120, "Monthly Hours": 100,
            "Annual Hours": 100, "Workload 2025_Planned": 150, "Workload 2025_Actual": 150,
            "Remark": 200, "Project": 120, "Item": 100, "Technical Unit": 120, "Activities": 150,
            "Booking Hours": 100, "Booking Cost (Forecast)": 150, "Booking Period": 120,
            "Booking hours (Accepted by Project)": 200, "Booking Period (Accepted by Project)": 200,
            "Booking hours (Extra)": 150
        }
        
        for col in emp_columns:
            self.employee_tree.heading(col, text=col, command=lambda c=col: self.on_column_header_click(c))
            width = column_widths.get(col, 100)
            self.employee_tree.column(col, width=width, anchor="center", minwidth=80)
        
        # Initialize filter menus for each column
        for col in emp_columns:
            self.filter_menus[col] = None
        
        # Scrollbars for employee data grid
        emp_v_scrollbar = ttk.Scrollbar(emp_tree_container, orient="vertical", command=self.employee_tree.yview)
        emp_h_scrollbar = ttk.Scrollbar(emp_tree_container, orient="horizontal", command=self.employee_tree.xview)
        
        self.employee_tree.configure(yscrollcommand=emp_v_scrollbar.set, xscrollcommand=emp_h_scrollbar.set)
        
        # Pack employee data grid
        self.employee_tree.pack(side="left", fill="both", expand=True)
        emp_v_scrollbar.pack(side="right", fill="y")
        emp_h_scrollbar.pack(side="bottom", fill="x")
        
        # Position filter buttons after the tree is created
        self.root.after(100, self.position_filter_buttons)
        
    def on_column_header_click(self, column):
        """Handle column header click for sorting and filtering"""
        try:
            # Check if this is a filter button click area (right side of header)
            # For now, just show the filter - you can add more sophisticated click detection
            self.show_column_filter(column)
        except Exception as e:
            logging.error(f"Column header click error: {e}")
    
        
        # Load employee data initially
        self.load_employee_data_grid()
    
    def setup_booking_management_panel(self):
        """Setup booking management panel"""
        booking_frame = ctk.CTkFrame(self.main_frame)
        booking_frame.pack(fill="x", padx=5, pady=(0, 5))  # Minimal padding
        
        booking_label = ctk.CTkLabel(
            booking_frame, 
            text="Step 3: Booking Management",
            font=ctk.CTkFont(size=14, weight="bold")  # Smaller font
        )
        booking_label.pack(pady=(3, 2))  # Minimal padding
        
        # Booking summary
        summary_frame = ctk.CTkFrame(booking_frame)
        summary_frame.pack(fill="x", padx=10, pady=3)  # Minimal padding
        
        self.booking_summary_label = ctk.CTkLabel(
            summary_frame, 
            text="Total Booking Hours: 0 | Total Cost: $0.00 | Status: No bookings",
            font=ctk.CTkFont(size=12)  # Smaller font
        )
        self.booking_summary_label.pack(pady=3)  # Minimal padding
        
        # Booking actions
        booking_action_frame = ctk.CTkFrame(booking_frame)
        booking_action_frame.pack(fill="x", padx=10, pady=(0, 3))  # Minimal padding
        
        self.save_booking_btn = ctk.CTkButton(
            booking_action_frame, 
            text="Save Booking", 
            command=self.save_booking,
            width=150
        )
        self.save_booking_btn.pack(side="left", padx=3)
        
        self.approve_booking_btn = ctk.CTkButton(
            booking_action_frame, 
            text="Approve Booking", 
            command=self.approve_booking,
            width=150
        )
        self.approve_booking_btn.pack(side="left", padx=3)
        
        self.reject_booking_btn = ctk.CTkButton(
            booking_action_frame, 
            text="Reject Booking", 
            command=self.reject_booking,
            width=150
        )
        self.reject_booking_btn.pack(side="left", padx=3)
    
    def load_data(self):
        """Load data from database using unified tables"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Load technical units with mapping
            cursor.execute("SELECT id, name FROM technical_unit ORDER BY name")
            self.technical_units = cursor.fetchall()
            self.technical_unit_map = {tu[1]: tu[0] for tu in self.technical_units}
            tech_unit_names = [tu[1] for tu in self.technical_units]
            self.tech_unit_dropdown.configure(values=tech_unit_names)
            
            # Load projects with mapping
            cursor.execute("SELECT id, name FROM project ORDER BY name")
            self.projects = cursor.fetchall()
            self.project_map = {p[1]: p[0] for p in self.projects}
            project_names = [p[1] for p in self.projects]
            self.project_dropdown.configure(values=project_names)
            
            # Load employees from employee table only with mapping
            cursor.execute("SELECT id, name FROM employee ORDER BY name")
            all_employees = cursor.fetchall()
            
            # Create employee mapping - all employees are now unified
            self.employees = []
            self.employee_map = {}
            
            for emp in all_employees:
                display_name = emp[1]  # Just use the name without type indicator
                self.employees.append((emp[0], emp[1], "unified"))
                self.employee_map[display_name] = {"id": emp[0], "type": "unified"}
            
            employee_names = list(self.employee_map.keys())
            self.employee_dropdown.configure(values=employee_names)
            
            conn.close()
            
            # Load the employee data grid
            self.load_employee_data_grid()
            
            print("Data loaded successfully from unified tables")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")
            logging.error(f"Data loading error: {e}")
    
    def on_technical_unit_change(self, value):
        """Handle technical unit selection change"""
        if value:
            self.check_and_add_service_data()
    
    def on_project_change(self, value):
        """Handle project selection change"""
        if value:
            self.check_and_add_service_data()
    
    def on_employee_change(self, value):
        """Handle employee selection change"""
        if value:
            self.display_employee_details()  # Show employee details if panel exists
            self.load_employee_services()    # Load services for selected employee
            self.check_and_add_service_data() # Check if all dropdowns are selected
    
    def check_and_add_service_data(self):
        """Check if all dropdowns are selected and filter service table to add to project_bookings table"""
        try:
            tech_unit_name = self.selected_technical_unit.get()
            project_name = self.selected_project.get()
            employee_name = self.selected_employee.get()
            
            # Only proceed if all three have valid selections
            if (tech_unit_name and tech_unit_name in self.technical_unit_map and 
                project_name and project_name in self.project_map and 
                employee_name and employee_name in self.employee_map):
                
                # Get IDs from the mappings
                tech_unit_id = self.technical_unit_map[tech_unit_name]
                project_id = self.project_map[project_name]
                employee_info = self.employee_map[employee_name]
                employee_id = employee_info["id"]
                
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Filter service table by the three dropdown selections
                cursor.execute("""
                    SELECT id, estimated_internal_hours, estimated_external_hours, 
                           start_date, due_date, activities_id, title_id
                    FROM service 
                    WHERE technical_unit_id = ? AND project_id = ? AND employee_id = ?
                """, (tech_unit_id, project_id, employee_id))
                
                matching_services = cursor.fetchall()
                
                if matching_services:
                    # Add each matching service to project_bookings table
                    bookings_added = 0
                    for service in matching_services:
                        service_id = service[0]
                        estimated_internal = service[1] or 0
                        estimated_external = service[2] or 0
                        # Use total estimated hours (internal + external)
                        total_estimated = (estimated_internal or 0) + (estimated_external or 0)
                        start_date = service[3]
                        end_date = service[4]  # due_date
                        service_notes = service[5]
                        
                        # Check if this booking already exists
                        cursor.execute("""
                            SELECT id FROM project_bookings 
                            WHERE employee_id = ? AND technical_unit_id = ? 
                            AND project_id = ? AND service_id = ?
                        """, (employee_id, tech_unit_id, project_id, service_id))
                        
                        existing_booking = cursor.fetchone()
                        
                        if not existing_booking:
                            # Get employee data for the new fields
                            cursor.execute("""
                                SELECT cost_center, ghrs_id, COALESCE(first_name || ' ' || last_name, last_name, first_name, 'N/A') as employee_name, dept_description,
                                       work_location, business_unit, tipo, tipo_description, sap_tipo,
                                       saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                                       monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                                       remark
                                FROM employee_extended WHERE id = ?
                            """, (employee_id,))
                            emp_data = cursor.fetchone()
                            
                            # Fallback: get employee name from main employee table if extended doesn't have it
                            if not emp_data or not emp_data[2]:
                                cursor.execute("SELECT name FROM employee WHERE id = ?", (employee_id,))
                                emp_name_fallback = cursor.fetchone()
                                emp_name = emp_name_fallback[0] if emp_name_fallback else "N/A"
                            else:
                                emp_name = emp_data[2]
                            
                            # Get project name
                            cursor.execute("SELECT name FROM project WHERE id = ?", (project_id,))
                            project_data = cursor.fetchone()
                            project_name_val = project_data[0] if project_data else "N/A"
                            
                            # Get technical unit name
                            cursor.execute("SELECT name FROM technical_unit WHERE id = ?", (tech_unit_id,))
                            tu_data = cursor.fetchone()
                            tu_name_val = tu_data[0] if tu_data else "N/A"
                            
                            # Get activities name
                            cursor.execute("""
                                SELECT a.name FROM service s 
                                LEFT JOIN activities a ON s.activities_id = a.id 
                                WHERE s.id = ?
                            """, (service_id,))
                            activity_data = cursor.fetchone()
                            activity_name_val = activity_data[0] if activity_data else "N/A"
                            
                            # Prepare values with defaults
                            if emp_data:
                                (cost_center, ghrs_id, employee_name_from_query, dept_description,
                                 work_location, business_unit, tipo, tipo_description, sap_tipo,
                                 saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                                 monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                                 remark) = emp_data
                            else:
                                # Default values when employee_extended data not available
                                (cost_center, ghrs_id, dept_description,
                                 work_location, business_unit, tipo, tipo_description, sap_tipo,
                                 saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                                 monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                                 remark) = (None, None, None, None, None, None, None, None,
                                           0.00, 0.00, 0.00, 0.00, 0, 0, 0.00, 0.00, None)
                            
                            # Parse first_name and last_name from full name (FIXED VERSION 2025-08-08)
                            cursor.execute("SELECT name FROM employee WHERE id = ?", (employee_id,))
                            name_data = cursor.fetchone()
                            full_name = name_data[0] if name_data and name_data[0] else ""
                            
                            # Try to split the name into first and last names
                            if full_name and full_name.strip():
                                name_parts = full_name.strip().split()
                                if len(name_parts) >= 2:
                                    first_name = name_parts[0]
                                    last_name = " ".join(name_parts[1:])
                                else:
                                    first_name = ""
                                    last_name = full_name
                            else:
                                first_name = ""
                                last_name = ""
                            
                            # Insert comprehensive booking record with all 47 columns (excluding auto-increment id)
                            cursor.execute("""
                                INSERT INTO project_bookings 
                                (employee_id, technical_unit_id, project_id, service_id, actual_hours, hourly_rate, total_cost, 
                                 booking_status, booking_date, start_date, end_date,
                                 created_by, approved_by, created_at, updated_at,
                                 cost_center, ghrs_id, last_name, first_name, dept_description,
                                 work_location, business_unit, tipo, tipo_description, sap_tipo,
                                 saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                                 monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                                 remark, project_name, item, technical_unit_name, activities_name,
                                 booking_hours, booking_cost_forecast, booking_period,
                                 booking_hours_accepted, booking_period_accepted, booking_hours_extra,
                                 employee_name)
                                VALUES (?, ?, ?, ?, ?, NULL, NULL, NULL, 'Pending', 
                                        CURRENT_DATE, ?, ?, ?, NULL, NULL, 
                                        CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                                        ?, NULL, ?, ?, 0.00, 0.00, NULL, 0.00, NULL, 0.00, ?)
                            """, (employee_id, tech_unit_id, project_id, service_id,
                                  total_estimated, start_date, end_date, service_notes,
                                  cost_center, ghrs_id, last_name, first_name, dept_description,
                                  work_location, business_unit, tipo, tipo_description, sap_tipo,
                                  saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                                  monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                                  remark, project_name_val, tu_name_val, activity_name_val, emp_name))
                            bookings_added += 1
                    
                    conn.commit()
                    
                    # Immediate refresh of data and grids
                    self.refresh_employee_data()
                    self.load_data()  # Refresh dropdowns to include any new data
                    
                    # DON'T reset dropdowns - keep selections as requested
                    
                    if bookings_added > 0:
                        messagebox.showinfo("Success", 
                            f"Added {bookings_added} booking record(s) for {employee_name} to project {project_name}!\n" +
                            "Unknown fields are set to NULL and can be edited by double-clicking cells.")
                    else:
                        messagebox.showinfo("Info", 
                            f"All matching services for {employee_name} are already booked to project {project_name}.")
                else:
                    messagebox.showwarning("No Services Found", 
                        f"No services found matching:\n" +
                        f"Technical Unit: {tech_unit_name}\n" +
                        f"Project: {project_name}\n" +
                        f"Employee: {employee_name}")
                
                conn.close()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add service data: {e}")
            logging.error(f"Service data addition error: {e}")
            import traceback
            traceback.print_exc()
    
    def display_employee_details(self):
        """Display selected employee details from unified employee table"""
        selected = self.selected_employee.get()
        if not selected or selected not in self.employee_map:
            return
        
        try:
            employee_info = self.employee_map[selected]
            employee_id = employee_info["id"]
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Get employee information from unified employee table
            cursor.execute("SELECT name FROM employee WHERE id = ?", (employee_id,))
            emp_data = cursor.fetchone()
            
            if emp_data:
                details_text = f"Employee: {emp_data[0]}\nType: Unified Employee Table"
            else:
                details_text = "Employee details not found"
            
            if hasattr(self, 'employee_info_label'):
                self.employee_info_label.configure(text=details_text, justify="left")
            conn.close()
            
        except Exception as e:
            if hasattr(self, 'employee_info_label'):
                self.employee_info_label.configure(text=f"Error loading employee details: {e}")
            logging.error(f"Employee details error: {e}")
    
    def load_employee_services(self):
        """Load services assigned to selected employee from project_bookings table using unified tables"""
        selected_emp = self.selected_employee.get()
        if not selected_emp or selected_emp not in self.employee_map:
            # Clear the treeview if it exists
            if hasattr(self, 'service_tree'):
                for item in self.service_tree.get_children():
                    self.service_tree.delete(item)
            return
        
        try:
            # Clear existing items
            if hasattr(self, 'service_tree'):
                for item in self.service_tree.get_children():
                    self.service_tree.delete(item)
            
            # Get employee info from mapping
            employee_info = self.employee_map[selected_emp]
            employee_id = employee_info["id"]
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Load from project_bookings table with detailed service information
            cursor.execute("""
                SELECT pb.id, p.name as project_name, 
                       COALESCE(a.name, 'Unknown Activity') as activity_name,
                       pb.actual_hours,
                       pb.hourly_rate, pb.total_cost, pb.booking_status, 
                       COALESCE(pb.start_date, pb.booking_date) as period,
                       tu.name as technical_unit_name
                FROM project_bookings pb
                LEFT JOIN project p ON pb.project_id = p.id
                LEFT JOIN technical_unit tu ON pb.technical_unit_id = tu.id
                LEFT JOIN service s ON pb.service_id = s.id
                LEFT JOIN activities a ON s.activities_id = a.id
                WHERE pb.employee_id = ?
                ORDER BY pb.created_at DESC
            """, (employee_id,))
            
            bookings = cursor.fetchall()
            
            # Populate treeview if it exists
            total_hours = 0
            total_cost = 0
            
            if hasattr(self, 'service_tree'):
                for booking in bookings:
                    booking_id = booking[0]
                    project_name = booking[1] or "Unknown Project"
                    activity_name = booking[2] or "Unknown Activity"
                    actual_hours = booking[3] or 0
                    rate = booking[4] or 0
                    cost = booking[5] or 0
                    status = booking[6] or "Pending"
                    period = booking[7] or "N/A"
                    technical_unit = booking[8] or "N/A"
                    
                    # Format display values
                    display_service = f"{project_name} ({technical_unit})"
                    display_period = str(period)[:10] if period else "N/A"  # Show date only
                    
                    self.service_tree.insert("", "end", values=(
                        display_service, activity_name, actual_hours, 
                        rate, cost, status, display_period
                    ), tags=(str(booking_id),))  # Store booking ID in tags for future reference
                    
                    total_hours += float(actual_hours) if actual_hours else 0
                    total_cost += float(cost) if cost else 0
            
            # Update summary if it exists
            if hasattr(self, 'booking_summary_label'):
                self.booking_summary_label.configure(
                    text=f"Total Booking Hours: {total_hours:.1f} | Total Cost: ${total_cost:.2f} | Status: {len(bookings)} bookings"
                )
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employee services: {e}")
            logging.error(f"Service loading error: {e}")
            import traceback
            traceback.print_exc()
    
    def load_employee_data_grid(self):
        """Load complete project bookings data with foreign key lookups - similar to Fabsi service table"""
        try:
            # Clear existing items
            if hasattr(self, 'employee_tree'):
                for item in self.employee_tree.get_children():
                    self.employee_tree.delete(item)
                
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Get all project bookings with all columns and foreign key lookups
                cursor.execute("""
                    SELECT 
                        pb.id,
                        pb.cost_center,
                        COALESCE(e.ghrs_id, pb.ghrs_id, 'N/A') as ghrs_id,
                        COALESCE(e.name, pb.employee_name, 'N/A') as employee_name,
                        COALESCE(d.name, pb.dept_description, 'N/A') as department_name,
                        COALESCE(h.name, 'N/A') as hub_name,
                        pb.work_location,
                        pb.business_unit,
                        pb.tipo,
                        pb.tipo_description,
                        pb.sap_tipo,
                        pb.saabu_rate_eur,
                        pb.saabu_rate_usd,
                        pb.local_agency_rate_usd,
                        pb.unit_rate_usd,
                        pb.monthly_hours,
                        pb.annual_hours,
                        pb.workload_2025_planned,
                        pb.workload_2025_actual,
                        pb.remark,
                        COALESCE(pb.project_name, p.name, 'N/A') as project,
                        pb.item,
                        COALESCE(pb.technical_unit_name, tu.name, 'N/A') as technical_unit,
                        COALESCE(pb.activities_name, a.name, 'N/A') as activities,
                        pb.booking_period_from,
                        pb.booking_period_to,
                        pb.actual_hours,
                        pb.hourly_rate,
                        pb.total_cost,
                        pb.booking_status,
                        pb.booking_date,
                        pb.start_date,
                        pb.end_date 
                    FROM project_bookings pb
                    LEFT JOIN employee e ON pb.employee_id = e.id
                    LEFT JOIN technical_unit tu ON pb.technical_unit_id = tu.id
                    LEFT JOIN project p ON pb.project_id = p.id
                    LEFT JOIN service s ON pb.service_id = s.id
                    LEFT JOIN title t ON s.title_id = t.id
                    LEFT JOIN activities a ON s.activities_id = a.id
                    LEFT JOIN department d ON pb.department_id = d.id
                    LEFT JOIN hub h ON pb.hub_id = h.id
                    ORDER BY pb.id
                """)
                
                bookings_data = cursor.fetchall()
                
                # Define complete columns for all project booking data (with Select checkbox)
                complete_columns = (
                    "Select", "ID", "Cost Center", "GHRS ID", "Employee Name", "Department",
                    "Hub", "Work Location", "Business Unit", "Tipo", "Tipo Description", "SAP Tipo",
                    "SAABU Rate (EUR)", "SAABU Rate (USD)", "Local Agency Rate (USD)", "Unit Rate (USD)",
                    "Monthly Hours", "Annual Hours", "Workload 2025_Planned", "Workload 2025_Actual",
                    "Remark", "Project", "Item", "Technical Unit", "Activities", 
                    "Booking Period From", "Booking Period To",
                    "Actual Hours", "Hourly Rate", "Total Cost", "Status",
                    "Booking Date", "Start Date", "End Date"
                )
                
                # Update treeview columns if needed
                if hasattr(self, 'employee_tree'):
                    # Clear existing columns and set new ones
                    current_columns = self.employee_tree['columns']
                    if current_columns != complete_columns:
                        self.employee_tree.configure(columns=complete_columns)
                        
                        # Set column widths for complete view (with Select checkbox)
                        column_widths = {
                            "Select": 60, "ID": 50, "Cost Center": 100, "GHRS ID": 80, "Employee Name": 150,
                            "Department": 150, "Hub": 150, "Work Location": 120, "Business Unit": 120, "Tipo": 80,
                            "Tipo Description": 150, "SAP Tipo": 80, "SAABU Rate (EUR)": 100, "SAABU Rate (USD)": 100,
                            "Local Agency Rate (USD)": 120, "Unit Rate (USD)": 100, "Monthly Hours": 100, "Annual Hours": 100,
                            "Workload 2025_Planned": 120, "Workload 2025_Actual": 120, "Remark": 150, "Project": 150,
                            "Item": 120, "Technical Unit": 120, "Activities": 150, 
                            "Booking Period From": 120, "Booking Period To": 120,
                            "Actual Hours": 80, "Hourly Rate": 80, "Total Cost": 100, "Status": 80,
                            "Booking Date": 100, "Start Date": 100, "End Date": 100
                        }
                        
                        for col in complete_columns:
                            # Set basic column properties first
                            width = column_widths.get(col, 100)
                            anchor = 'w' if col in ["Employee Name", "Project", "Technical Unit", "Activities", 
                                                  "Department", "Hub", "Work Location", "Business Unit", 
                                                  "Tipo Description", "Remark"] else 'center'
                            self.employee_tree.column(col, width=width, anchor=anchor, minwidth=70)
                            
                            # Set headers (filter arrows will be added later after DataFrame creation)
                            self.employee_tree.heading(col, text=col)
                
                # Populate the grid with complete booking data (including checkbox)
                formatted_data_list = []  # Store for DataFrame
                for booking_data in bookings_data:
                    # Format the data for display (handle None values)
                    formatted_data = ["☐"]  # Start with unchecked checkbox
                    formatted_row = {}  # For DataFrame
                    
                    for i, value in enumerate(booking_data):
                        col_name = complete_columns[i + 1]  # +1 because we added Select column
                        
                        if value is None:
                            display_value = "N/A"
                            formatted_row[col_name] = ""
                        elif i in [10, 11, 12, 13, 15, 16, 17, 18, 26, 27, 28, 29]:  # Decimal/money columns (rates, hours, costs)
                            try:
                                display_value = f"{float(value):.2f}" if value else "0.00"
                                formatted_row[col_name] = float(value) if value else 0.0
                            except (ValueError, TypeError):
                                display_value = str(value) if value else "N/A"
                                formatted_row[col_name] = str(value) if value else ""
                        elif i in [14, 15]:  # Integer hours columns
                            try:
                                display_value = str(int(value)) if value else "0"
                                formatted_row[col_name] = int(value) if value else 0
                            except (ValueError, TypeError):
                                display_value = str(value) if value else "N/A"
                                formatted_row[col_name] = str(value) if value else ""
                        elif i in [24, 25, 30, 31, 32]:  # Date columns (booking_period_from, booking_period_to, booking_date, start_date, end_date)
                            display_value = str(value) if value else "N/A"
                            formatted_row[col_name] = str(value) if value else ""
                        else:
                            display_value = str(value) if value else "N/A"
                            formatted_row[col_name] = str(value) if value else ""
                            
                        formatted_data.append(display_value)
                    
                    # Add Select column to DataFrame
                    formatted_row["Select"] = False
                    formatted_data_list.append(formatted_row)
                    
                    self.employee_tree.insert("", "end", values=formatted_data)
                
                # Create DataFrame for filtering
                if formatted_data_list:
                    self.df = pd.DataFrame(formatted_data_list)
                    # Store original data for filter reset
                    self.original_df = self.df.copy()
                    
                    # Update column headers with filter arrows (like Fabsi app)
                    from functools import partial
                    for col in complete_columns:
                        if col not in ["Select", "ID"]:
                            header_text = f"{col} ▼"
                            self.employee_tree.heading(col, text=header_text, 
                                                     command=partial(self.show_filter_menu, col))
                else:
                    self.df = pd.DataFrame()
                    self.original_df = pd.DataFrame()
                
                conn.close()
                
                print(f"Loaded {len(bookings_data)} project booking records with full details")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load project booking data: {e}")
            logging.error(f"Project booking data grid loading error: {e}")
            import traceback
            traceback.print_exc()
    
    def delete_employee_record(self):
        """Delete selected project booking record"""
        if not hasattr(self, 'employee_tree'):
            return
            
        selected_item = self.employee_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a project booking to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this project booking record?"):
            try:
                item_values = self.employee_tree.item(selected_item[0])['values']
                booking_id = item_values[1]  # ID is now in position 1 due to checkbox
                
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Delete the project booking record
                cursor.execute("DELETE FROM project_bookings WHERE id = ?", (booking_id,))
                
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Success", "Project booking deleted successfully")
                # Auto refresh removed - user can manually refresh if needed
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete project booking: {e}")
                logging.error(f"Project booking deletion error: {e}")
    
    def load_employee_data_grid_for_filter(self):
        """Load employee data grid - same as load_employee_data_grid but for filter system"""
        self.load_employee_data_grid()
    
    def position_filter_buttons(self):
        """Position filter buttons above their respective columns"""
        try:
            if not hasattr(self, 'employee_tree') or not self.filter_menus:
                return
            
            # Clear existing buttons from header frame
            for widget in self.header_frame.winfo_children():
                widget.destroy()
            
            # Get column positions and create buttons
            columns = list(self.employee_tree['columns'])
            
            for i, col in enumerate(columns):
                if col in self.filter_menus and col != "Select":
                    # Calculate position based on column width
                    col_width = self.employee_tree.column(col, 'width')
                    x_pos = sum(self.employee_tree.column(columns[j], 'width') for j in range(i))
                    
                    # Create filter button
                    filter_btn = ctk.CTkButton(
                        self.header_frame,
                        text="🔽",
                        width=20,
                        height=20,
                        command=lambda c=col: self.show_column_filter(c),
                        fg_color="#003d52",
                        hover_color="#255c7b",
                        font=ctk.CTkFont(size=8)
                    )
                    filter_btn.place(x=x_pos + col_width - 25, y=2)
                    self.filter_menus[col] = filter_btn
                    
        except Exception as e:
            logging.error(f"Position filter buttons error: {e}")
    
    def show_column_filter(self, column):
        """Show Excel-like filter popup for the column - exact copy from Fabsi app"""
        try:
            # Close existing popup if any
            if hasattr(self, 'filter_popup') and self.filter_popup:
                self.filter_popup.destroy()
            
            # Get button widget and position
            if column not in self.filter_menus:
                return
                
            filter_btn = self.filter_menus[column]
            x = filter_btn.winfo_rootx()
            y = filter_btn.winfo_rooty() + filter_btn.winfo_height()
            
            # Create popup window
            self.filter_popup = ctk.CTkToplevel(self.root)
            self.filter_popup.wm_overrideredirect(True)
            self.filter_popup.geometry(f"220x380+{x}+{y}")
            
            # Main container frame
            main_frame = ctk.CTkFrame(self.filter_popup, corner_radius=8)
            main_frame.pack(fill='both', expand=True, padx=3, pady=3)
            
            # Store current filter values for search
            self.current_filter_values = []
            if column in self.df.columns:
                self.current_filter_values = sorted(self.df[column].unique().tolist())
                self.current_filter_values = [str(val) for val in self.current_filter_values if pd.notnull(val)]
            
            # Frame for sort options at the top
            sort_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            sort_frame.pack(fill='x', pady=(5, 5))
            
            # Sort buttons
            ctk.CTkButton(sort_frame, text="Sort A → Z", 
                         command=lambda: self.sort_column(column, ascending=True),
                         height=30, font=ctk.CTkFont(size=10)).pack(fill='x', pady=2)
            ctk.CTkButton(sort_frame, text="Sort Z → A",
                         command=lambda: self.sort_column(column, ascending=False),
                         height=30, font=ctk.CTkFont(size=10)).pack(fill='x', pady=2)
            
            # Separator
            separator_frame = ctk.CTkFrame(main_frame, height=2, fg_color="#22505f")
            separator_frame.pack(fill='x', pady=5)
            
            # Filter options frame
            filter_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            filter_frame.pack(fill='both', expand=True, padx=5)
            
            # Search entry
            search_var = tk.StringVar()
            search_entry = ctk.CTkEntry(filter_frame, textvariable=search_var, 
                                       placeholder_text="Search...",
                                       font=ctk.CTkFont(family='Arial', size=10))
            search_entry.pack(fill='x', pady=(0, 5))
            
            # Checkboxes container with scrollbar
            checkbox_container = ctk.CTkScrollableFrame(filter_frame, height=200)
            checkbox_container.pack(fill='both', expand=True)
            
            # Dictionary to store checkbox variables
            self.current_column_filter_vars = {}
            
            # Add checkboxes
            for val in self.current_filter_values:
                var = tk.BooleanVar()
                var.set(True)  # Start with all selected
                self.current_column_filter_vars[val] = var
                cb = ctk.CTkCheckBox(checkbox_container, text=str(val), variable=var,
                                   font=ctk.CTkFont(family='Arial', size=9))
                cb.pack(anchor='w', padx=5, pady=2)
            
            # Update filter options based on search
            def update_search(*args):
                search_text = search_var.get().lower()
                for widget in checkbox_container.winfo_children():
                    if hasattr(widget, 'cget'):
                        try:
                            if search_text in widget.cget('text').lower():
                                widget.pack(anchor='w', padx=5, pady=2)
                            else:
                                widget.pack_forget()
                        except:
                            pass
            
            search_var.trace('w', update_search)
            
            # Buttons frame at the bottom
            btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            btn_frame.pack(side='bottom', fill='x', padx=5, pady=(5, 10))
            
            # Pack buttons with equal spacing
            ctk.CTkButton(btn_frame, text="Apply", 
                         command=lambda: self.apply_column_filter(column),
                         fg_color='#003d52', hover_color='#255c7b',
                         width=60, height=30, font=ctk.CTkFont(size=9)).pack(side='left', expand=True, padx=2)
            ctk.CTkButton(btn_frame, text="Clear",
                         command=lambda: self.clear_column_filter(column),
                         fg_color='#ef8827', hover_color='#22505f',
                         width=60, height=30, font=ctk.CTkFont(size=9)).pack(side='left', expand=True, padx=2)
            ctk.CTkButton(btn_frame, text="Cancel",
                         command=self.filter_popup.destroy,
                         fg_color='#255c7b', hover_color='#22505f',
                         width=60, height=30, font=ctk.CTkFont(size=9)).pack(side='left', expand=True, padx=2)
            
            # Set focus to search entry
            search_entry.focus()
            
        except Exception as e:
            logging.error(f"Show column filter error: {e}")
            import traceback
            traceback.print_exc()
    
    def sort_column(self, column, ascending=True):
        """Sort the data by column - like Fabsi app"""
        try:
            # Close filter popup if open
            if hasattr(self, 'filter_popup') and self.filter_popup:
                self.filter_popup.destroy()
            
            if column not in self.df.columns:
                return
            
            # Sort the DataFrame
            self.df = self.df.sort_values(by=column, ascending=ascending, na_position='last')
            self.refresh_display()
            
        except Exception as e:
            logging.error(f"Sort column error: {e}")
    
    def apply_column_filter(self, column):
        """Apply filter to the column based on selected values - like Fabsi app"""
        try:
            if not hasattr(self, 'current_column_filter_vars'):
                return
            
            # Get selected values
            selected_values = []
            for value, var in self.current_column_filter_vars.items():
                if var.get():
                    selected_values.append(value)
            
            # Store the filter
            if selected_values:
                self.active_column_filters[column] = selected_values
            else:
                # If nothing selected, remove filter
                if column in self.active_column_filters:
                    del self.active_column_filters[column]
            
            # Apply all filters and refresh display
            self.apply_all_filters()
            
            # Close popup
            if hasattr(self, 'filter_popup') and self.filter_popup:
                self.filter_popup.destroy()
                
        except Exception as e:
            logging.error(f"Apply column filter error: {e}")
            import traceback
            traceback.print_exc()
    
    def clear_column_filter(self, column):
        """Clear filter for a specific column"""
        try:
            if column in self.active_column_filters:
                del self.active_column_filters[column]
            
            self.apply_all_filters()
            
            # Close popup
            if hasattr(self, 'filter_popup') and self.filter_popup:
                self.filter_popup.destroy()
                
        except Exception as e:
            logging.error(f"Clear column filter error: {e}")
    
    def apply_all_filters(self):
        """Apply all active filters to the DataFrame - PRESERVE SELECTIONS"""
        try:
            # Store current selections BEFORE reloading data
            current_selections = {}
            for item in self.employee_tree.get_children():
                values = self.employee_tree.item(item, 'values')
                if values and len(values) > 1:
                    row_id = values[1]  # ID is in second column
                    is_selected = values[0] == "☑"
                    if is_selected:
                        current_selections[str(row_id)] = True
            
            # Start with full dataset - reload from database
            self.load_employee_data_grid_for_filter()
            
            # Apply each active filter
            for column, values in self.active_column_filters.items():
                if column in self.df.columns:
                    # Convert values to string for comparison
                    values_str = [str(v) for v in values]
                    self.df = self.df[self.df[column].astype(str).isin(values_str)]
            
            # Refresh the display AND restore selections
            self.refresh_display_with_selections(current_selections)
            
        except Exception as e:
            logging.error(f"Apply all filters error: {e}")
            import traceback
            traceback.print_exc()
    
    def refresh_display(self):
        """Refresh the tree display with current DataFrame data"""
        try:
            # Clear existing items
            for item in self.employee_tree.get_children():
                self.employee_tree.delete(item)
            
            # Populate with filtered data
            for _, row in self.df.iterrows():
                values = []
                columns = list(self.employee_tree['columns'])
                
                for col in columns:
                    if col in row:
                        value = row[col]
                        if col == "Select":
                            values.append("☐")
                        else:
                            values.append(str(value) if pd.notnull(value) else "N/A")
                    else:
                        values.append("N/A")
                
                self.employee_tree.insert("", "end", values=values)
                
        except Exception as e:
            logging.error(f"Refresh display error: {e}")
            import traceback
            traceback.print_exc()
    
    def refresh_display_with_selections(self, preserved_selections):
        """Refresh the tree display with current DataFrame data and restore selections"""
        try:
            # Clear existing items
            for item in self.employee_tree.get_children():
                self.employee_tree.delete(item)
            
            # Clear and rebuild selected_rows set
            self.selected_rows.clear()
            
            # Populate with filtered data
            for _, row in self.df.iterrows():
                values = []
                columns = list(self.employee_tree['columns'])
                
                for col in columns:
                    if col in row:
                        value = row[col]
                        if col == "Select":
                            # Check if this row was previously selected
                            row_id = str(row.get('id', ''))
                            if row_id in preserved_selections:
                                values.append("☑")
                            else:
                                values.append("☐")
                        else:
                            # Handle None values
                            if pd.isna(value):
                                values.append("")
                            else:
                                values.append(str(value))
                    else:
                        values.append("")
                
                # Insert the row
                item = self.employee_tree.insert("", "end", values=values)
                
                # Add to selected_rows if it was selected
                row_id = str(row.get('id', ''))
                if row_id in preserved_selections:
                    self.selected_rows.add(item)
                    
        except Exception as e:
            logging.error(f"Refresh display with selections error: {e}")
            import traceback
            traceback.print_exc()
    
    def load_employee_data_grid_for_filter(self):
        """Reload the full dataset for filtering"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Get all project bookings with all columns and foreign key lookups
            cursor.execute("""
                SELECT 
                    pb.id,
                    pb.cost_center,
                    COALESCE(e.ghrs_id, pb.ghrs_id, 'N/A') as ghrs_id,
                    COALESCE(e.name, pb.employee_name, 'N/A') as employee_name,
                    pb.dept_description,
                    pb.work_location,
                    pb.business_unit,
                    pb.tipo,
                    pb.tipo_description,
                    pb.sap_tipo,
                    pb.saabu_rate_eur,
                    pb.saabu_rate_usd,
                    pb.local_agency_rate_usd,
                    pb.unit_rate_usd,
                    pb.monthly_hours,
                    pb.annual_hours,
                    pb.workload_2025_planned,
                    pb.workload_2025_actual,
                    pb.remark,
                    COALESCE(pb.project_name, p.name, 'N/A') as project,
                    pb.item,
                    COALESCE(pb.technical_unit_name, tu.name, 'N/A') as technical_unit,
                    COALESCE(pb.activities_name, a.name, 'N/A') as activities,
                    pb.booking_hours,
                    pb.booking_cost_forecast,
                    pb.booking_period,
                    pb.booking_hours_accepted,
                    pb.booking_period_accepted,
                    pb.booking_hours_extra,
                    pb.pb.actual_hours,
                    pb.hourly_rate,
                    pb.total_cost,
                    pb.booking_status,
                    pb.booking_date,
                    pb.start_date,
                    pb.end_date FROM project_bookings pb
                LEFT JOIN employee e ON pb.employee_id = e.id
                LEFT JOIN technical_unit tu ON pb.technical_unit_id = tu.id
                LEFT JOIN project p ON pb.project_id = p.id
                LEFT JOIN service s ON pb.service_id = s.id
                LEFT JOIN title t ON s.title_id = t.id
                LEFT JOIN activities a ON s.activities_id = a.id
                ORDER BY pb.id
            """)
            
            bookings_data = cursor.fetchall()
            
            # Define complete columns for all project booking data (with Select checkbox)
            complete_columns = (
                "Select", "ID", "Cost Center", "GHRS ID", "Employee Name", "Dept. Description",
                "Work Location", "Business Unit", "Tipo", "Tipo Description", "SAP Tipo",
                "SAABU Rate (EUR)", "SAABU Rate (USD)", "Local Agency Rate (USD)", "Unit Rate (USD)",
                "Monthly Hours", "Annual Hours", "Workload 2025_Planned", "Workload 2025_Actual",
                "Remark", "Project", "Item", "Technical Unit", "Activities", "Booking Hours",
                "Booking Cost (Forecast)", "Booking Period", "Booking hours (Accepted by Project)",
                "Booking Period (Accepted by Project)", "Booking hours (Extra)",
                "Actual Hours", "Hourly Rate", "Total Cost", "Status",
                "Booking Date", "Start Date", "End Date"
            )
            
            # Create DataFrame for filtering
            formatted_data_list = []
            for booking_data in bookings_data:
                formatted_row = {"Select": False}  # Start with checkbox unchecked
                
                for i, value in enumerate(booking_data):
                    col_name = complete_columns[i + 1]  # +1 because we added Select column
                    
                    if value is None:
                        formatted_row[col_name] = ""
                    elif i in [10, 11, 12, 13, 16, 17, 23, 24, 26, 28, 29, 30, 31, 32]:  # Decimal/money columns
                        formatted_row[col_name] = float(value) if value else 0.0
                    elif i in [14, 15]:  # Integer hours columns
                        formatted_row[col_name] = int(value) if value else 0
                    elif i in [34, 35, 36]:  # Date columns
                        formatted_row[col_name] = str(value) if value else ""
                    else:
                        formatted_row[col_name] = str(value) if value else ""
                        
                formatted_data_list.append(formatted_row)
            
            self.df = pd.DataFrame(formatted_data_list)
            conn.close()
            
        except Exception as e:
            logging.error(f"Load data for filter error: {e}")
            import traceback
            traceback.print_exc()
    
    def reset_filters(self):
        """Reset all filters - like Fabsi app"""
        try:
            self.active_column_filters.clear()
            self.apply_all_filters()
            
        except Exception as e:
            logging.error(f"Reset filters error: {e}")
    
    def clear_all_filters(self):
        """Clear all active filters"""
        try:
            self.active_column_filters.clear()
            self.apply_all_filters()
            messagebox.showinfo("Filters Cleared", "All filters have been cleared")
            
        except Exception as e:
            logging.error(f"Clear filters error: {e}")
    
    def on_simplified_cell_edit(self, event):
        """Excel-like cell editing: Click to edit any cell directly"""
        try:
            # Get the clicked item and column
            item = self.employee_tree.selection()[0]
            column = self.employee_tree.identify_column(event.x)
            
            # Get column index (remove '#' prefix)
            col_idx = int(column.replace('#', '')) - 1
            
            # Get current values
            current_values = list(self.employee_tree.item(item, 'values'))
            if col_idx >= len(current_values):
                return
                
            # Get the actual column names from the data query to map to database fields
            complete_columns = (
                "Select", "ID", "Cost Center", "GHRS ID", "Employee Name", "Dept. Description",
                "Work Location", "Business Unit", "Tipo", "Tipo Description", "SAP Tipo",
                "SAABU Rate (EUR)", "SAABU Rate (USD)", "Local Agency Rate (USD)", "Unit Rate (USD)",
                "Monthly Hours", "Annual Hours", "Workload 2025_Planned", "Workload 2025_Actual",
                "Remark", "Project", "Item", "Technical Unit", "Activities", "Booking Hours",
                "Booking Cost (Forecast)", "Booking Period", "Booking hours (Accepted by Project)",
                "Booking Period (Accepted by Project)", "Booking hours (Extra)",
                "Actual Hours", "Hourly Rate", "Total Cost", "Status",
                "Booking Date", "Start Date", "End Date"
            )
            
            # Map display columns to database column names (adjusted for checkbox)
            db_column_map = {
                0: None,  # Select checkbox - not editable
                1: None,  # ID - not editable
                2: "cost_center",
                3: "ghrs_id", 
                4: "employee_name",
                5: "dept_description",
                6: "work_location",
                7: "business_unit",
                8: "tipo",
                9: "tipo_description", 
                10: "sap_tipo",
                11: "saabu_rate_eur",
                12: "saabu_rate_usd",
                13: "local_agency_rate_usd",
                14: "unit_rate_usd",
                15: "monthly_hours",
                16: "annual_hours",
                17: "workload_2025_planned",
                18: "workload_2025_actual",
                19: "remark",
                20: "project_name",
                21: "item",
                22: "technical_unit_name",
                23: "activities_name",
                24: "booking_hours",
                25: "booking_cost_forecast",
                26: "booking_period",
                27: "booking_hours_accepted",
                28: "booking_period_accepted", 
                29: "booking_hours_extra",
                30: "actual_hours",
                31: "hourly_rate",
                32: "total_cost",
                33: "booking_status",
                34: "booking_date",
                35: "start_date",
                36: "end_date"
            }
            
            # Check if column is editable
            if col_idx in [0, 1]:  # Select checkbox and ID column
                if col_idx == 0:
                    messagebox.showinfo("Info", "Use single click to toggle row selection")
                else:
                    messagebox.showinfo("Info", "Booking ID cannot be edited")
                return
                
            db_column = db_column_map.get(col_idx)
            if not db_column:
                messagebox.showinfo("Info", "This column cannot be edited")
                return
            
            column_name = complete_columns[col_idx]
            current_value = current_values[col_idx] if col_idx < len(current_values) else ""
            
            # Remove "N/A" or format current value for editing
            if current_value == "N/A":
                current_value = ""
            
            # Create inline editing dialog
            self.create_inline_edit_dialog(item, col_idx, column_name, current_value, db_column)
            
        except (IndexError, ValueError, KeyError):
            pass  # No item selected or invalid column

    def create_inline_edit_dialog(self, tree_item, col_idx, column_name, current_value, db_column):
        """Create a small dialog for editing cell value like Excel"""
        # Get the booking ID for database update (now in column 1 due to checkbox)
        item_values = self.employee_tree.item(tree_item)['values']
        booking_id = item_values[1]  # ID is now in position 1
        
        # Store current row values for updating the display
        current_row_values = list(item_values)
        
        # Create compact edit dialog
        edit_dialog = ctk.CTkToplevel(self.root)
        edit_dialog.title(f"Edit {column_name}")
        edit_dialog.geometry("400x200")
        edit_dialog.transient(self.root)
        edit_dialog.grab_set()
        
        # Center the dialog
        edit_dialog.update_idletasks()
        x = (edit_dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (edit_dialog.winfo_screenheight() // 2) - (200 // 2)
        edit_dialog.geometry(f"400x200+{x}+{y}")
        
        # Content frame
        content_frame = ctk.CTkFrame(edit_dialog)
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Label
        ctk.CTkLabel(content_frame, text=f"Edit {column_name}:", 
                    font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(0, 10))
        
        # Determine input type based on column
        numeric_columns = ["saabu_rate_eur", "saabu_rate_usd", "local_agency_rate_usd", "unit_rate_usd",
                          "workload_2025_planned", "workload_2025_actual", "booking_hours", 
                          "booking_cost_forecast", "booking_hours_accepted", "booking_hours_extra",
                          "actual_hours", "hourly_rate", "total_cost"]
        
        integer_columns = ["monthly_hours", "annual_hours"]
        
        date_columns = ["booking_date", "start_date", "end_date"]
        
        text_area_columns = ["remark", "dept_description", "tipo_description"]
        
        if db_column in text_area_columns:
            # Text area for longer text
            edit_widget = ctk.CTkTextbox(content_frame, height=80, width=350)
            edit_widget.pack(pady=5, fill="x")
            if current_value and current_value != "N/A":
                edit_widget.insert("1.0", current_value)
        else:
            # Single line entry
            edit_widget = ctk.CTkEntry(content_frame, width=350, placeholder_text=f"Enter {column_name.lower()}")
            edit_widget.pack(pady=5, fill="x")
            if current_value and current_value != "N/A":
                edit_widget.insert(0, current_value)
        
        # Validation info
        if db_column in numeric_columns:
            info_text = "Enter a decimal number (e.g., 123.45)"
        elif db_column in integer_columns:
            info_text = "Enter a whole number (e.g., 40)"
        elif db_column in date_columns:
            info_text = "Enter date in YYYY-MM-DD format (e.g., 2025-01-15)"
        else:
            info_text = "Enter text value"
            
        ctk.CTkLabel(content_frame, text=info_text, 
                    font=ctk.CTkFont(size=10), text_color="gray").pack(pady=(0, 10))
        
        # Buttons
        button_frame = ctk.CTkFrame(content_frame)
        button_frame.pack(fill="x", pady=(10, 0))
        
        def save_change():
            try:
                # Get the new value
                if db_column in text_area_columns:
                    new_value = edit_widget.get("1.0", "end-1c").strip()
                else:
                    new_value = edit_widget.get().strip()
                
                # Validate numeric inputs
                if db_column in numeric_columns and new_value:
                    try:
                        float(new_value)
                    except ValueError:
                        messagebox.showerror("Error", "Please enter a valid decimal number")
                        return
                        
                if db_column in integer_columns and new_value:
                    try:
                        int(new_value)
                    except ValueError:
                        messagebox.showerror("Error", "Please enter a valid whole number")
                        return
                
                # Date validation
                if db_column in date_columns and new_value:
                    try:
                        from datetime import datetime
                        datetime.strptime(new_value, '%Y-%m-%d')
                    except ValueError:
                        messagebox.showerror("Error", "Please enter date in YYYY-MM-DD format")
                        return
                
                # Update database
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Handle empty values
                if new_value == "":
                    new_value = None
                
                # Update the specific column
                cursor.execute(f"UPDATE project_bookings SET {db_column} = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", 
                              (new_value, booking_id))
                
                # Special handling for calculated fields
                if db_column in ["hourly_rate"]:
                    # Recalculate total_cost if hours or rate changed
                    cursor.execute("SELECT hourly_rate FROM project_bookings WHERE id = ?", (booking_id,))
                    hours_rate = cursor.fetchone()
                    if hours_rate and hours_rate[0] and hours_rate[1]:
                        total_cost = float(hours_rate[0]) * float(hours_rate[1])
                        cursor.execute("UPDATE project_bookings SET total_cost = ? WHERE id = ?", (total_cost, booking_id))
                
                conn.commit()
                conn.close()
                
                # Update the stored row values
                current_row_values[col_idx] = new_value if new_value is not None else "N/A"
                
                # Format display value
                if db_column in numeric_columns and new_value:
                    current_row_values[col_idx] = f"{float(new_value):.2f}"
                elif db_column in integer_columns and new_value:
                    current_row_values[col_idx] = str(int(new_value))
                
                # Find the item by booking_id and update it (ID is now in position 1)
                for item in self.employee_tree.get_children():
                    values = self.employee_tree.item(item)['values']
                    if values and len(values) > 1 and values[1] == booking_id:
                        self.employee_tree.item(item, values=current_row_values)
                        break
                
                edit_dialog.destroy()
                messagebox.showinfo("Success", f"{column_name} updated successfully!")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update {column_name}: {e}")
                import traceback
                traceback.print_exc()
        
        def cancel_edit():
            edit_dialog.destroy()
        
        # Buttons
        ctk.CTkButton(button_frame, text="Save", command=save_change, width=100).pack(side="left", padx=(0, 10))
        ctk.CTkButton(button_frame, text="Cancel", command=cancel_edit, width=100).pack(side="left")
        
        # Focus on the input widget and select all text
        edit_widget.focus()
        if not db_column in text_area_columns and current_value and current_value != "N/A":
            edit_widget.select_range(0, tk.END)
        
        # Bind Enter key to save
        edit_dialog.bind('<Return>', lambda e: save_change())
        edit_dialog.bind('<Escape>', lambda e: cancel_edit())

    def on_keyboard_edit(self, event):
        """Handle F2 key press for editing like Excel"""
        try:
            if self.employee_tree.selection():
                # Simulate a double-click at the center of the selected cell
                item = self.employee_tree.selection()[0]
                bbox = self.employee_tree.bbox(item)
                if bbox:
                    # Create a fake event for the center of the first column
                    fake_event = type('Event', (), {})()
                    fake_event.x = bbox[0] + 50  # Approximate center of first editable column
                    fake_event.y = bbox[1] + bbox[3] // 2
                    self.on_simplified_cell_edit(fake_event)
        except Exception:
            pass
    
    def on_cell_select(self, event):
        """Handle cell selection for visual feedback"""
        try:
            # This provides visual feedback when clicking on cells
            # The actual editing happens on double-click
            pass
        except Exception:
            pass
            
    def toggle_edit_mode(self):
        """Toggle between single-click and double-click editing modes"""
        if self.edit_mode == "double":
            self.edit_mode = "single"
            self.employee_tree.bind('<Button-1>', self.on_simplified_cell_edit)
            self.edit_mode_btn.configure(text="Edit Mode: Single-Click")
            messagebox.showinfo("Edit Mode", "Single-click editing enabled!\nClick any cell to edit directly.")
        else:
            self.edit_mode = "double"
            self.employee_tree.unbind('<Button-1>')
            self.employee_tree.bind('<Button-1>', self.on_cell_select)
            self.edit_mode_btn.configure(text="Edit Mode: Double-Click")
            messagebox.showinfo("Edit Mode", "Double-click editing enabled!\nDouble-click any cell to edit, or press F2.")
    
    def open_employee_dialog(self, employee_id=None):
        """Open dialog for adding/editing employee with all 29 fields"""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Employee Record" if not employee_id else "Edit Employee Record")
        dialog.geometry("800x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Create scrollable frame
        scrollable_frame = ctk.CTkScrollableFrame(dialog, width=760, height=550)
        scrollable_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_text = "Add New Employee" if not employee_id else "Edit Employee"
        ctk.CTkLabel(scrollable_frame, text=title_text, font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(0, 20))
        
        # Dictionary to store entry widgets
        entry_widgets = {}
        
        # Define all 29 fields with their labels
        fields = [
            ("cost_center", "Cost Center"),
            ("ghrs_id", "GHRS ID"),
            ("last_name", "Last Name"),
            ("first_name", "First Name"),
            ("dept_description", "Dept. Description"),
            ("work_location", "Work Location"),
            ("business_unit", "Business Unit"),
            ("tipo", "Tipo"),
            ("tipo_description", "Tipo Description"),
            ("sap_tipo", "SAP Tipo"),
            ("saabu_rate_eur", "SAABU Rate (EUR)"),
            ("saabu_rate_usd", "SAABU Rate (USD)"),
            ("local_agency_rate_usd", "Local Agency Rate (USD)"),
            ("unit_rate_usd", "Unit Rate (USD)"),
            ("monthly_hours", "Monthly Hours"),
            ("annual_hours", "Annual Hours"),
            ("workload_2025_planned", "Workload 2025_Planned"),
            ("workload_2025_actual", "Workload 2025_Actual"),
            ("remark", "Remark"),
            ("project_assigned", "Project"),
            ("item", "Item"),
            ("technical_unit_assigned", "Technical Unit"),
            ("activities", "Activities"),
            ("booking_hours", "Booking Hours"),
            ("booking_cost_forecast", "Booking Cost (Forecast)"),
            ("booking_period", "Booking Period"),
            ("booking_hours_accepted", "Booking hours (Accepted by Project)"),
            ("booking_period_accepted", "Booking Period (Accepted by Project)"),
            ("booking_hours_extra", "Booking hours (Extra)")
        ]
        
        # Create input fields in a grid layout
        for i, (field_name, field_label) in enumerate(fields):
            row = i // 2
            col = i % 2
            
            field_frame = ctk.CTkFrame(scrollable_frame)
            field_frame.grid(row=row, column=col, padx=10, pady=5, sticky="ew")
            
            ctk.CTkLabel(field_frame, text=field_label + ":").pack(anchor="w", padx=5, pady=(5, 0))
            
            if field_name in ["remark", "activities"]:
                # Text area for longer fields
                entry = ctk.CTkTextbox(field_frame, height=60)
                entry.pack(fill="x", padx=5, pady=(0, 5))
            else:
                # Regular entry
                entry = ctk.CTkEntry(field_frame)
                entry.pack(fill="x", padx=5, pady=(0, 5))
            
            entry_widgets[field_name] = entry
        
        # Configure grid weights
        for i in range(2):
            scrollable_frame.grid_columnconfigure(i, weight=1)
        
        # Load existing data if editing
        if employee_id:
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT ghrs_id, last_name, first_name, cost_center, dept_description
                    FROM employee WHERE id = ?
                """, (employee_id,))
                
                emp_data = cursor.fetchone()
                conn.close()
                
                if emp_data:
                    # Simple display of employee information
                    info_text = f"""
Employee Details:
GHRS ID: {emp_data[0] or 'N/A'}
Name: {emp_data[2] or ''} {emp_data[1] or ''}
Cost Center: {emp_data[3] or 'N/A'}
Department: {emp_data[4] or 'N/A'}
                    """
                    
                    # Create simple info display
                    info_label = ctk.CTkLabel(
                        self.employee_details_frame,
                        text=info_text,
                        justify="left",
                        font=ctk.CTkFont(size=12)
                    )
                    info_label.pack(pady=10, padx=10, anchor="w")
                            
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load employee data: {e}")
        
        # Buttons frame  
        button_frame = ctk.CTkFrame(dialog)
        button_frame.pack(fill="x", padx=20, pady=10)
        
        close_btn = ctk.CTkButton(button_frame, text="Close", command=dialog.destroy, width=100)
        close_btn.pack(side="left", padx=10)
    
    def add_service_assignment(self):
        """Add new service assignment"""
        if not self.selected_employee.get():
            messagebox.showwarning("Warning", "Please select an employee first")
            return
        
        # Create a new window for service assignment
        self.open_service_assignment_dialog()
    
    def open_service_assignment_dialog(self):
        """Open dialog for adding/editing service assignment"""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Service Assignment")
        dialog.geometry("600x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Service selection
        ctk.CTkLabel(dialog, text="Select Service:", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(20, 5))
        
        # Load available services
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT s.id, p.name, a.name, t.name 
            FROM service s
            LEFT JOIN project p ON s.project_id = p.id
            LEFT JOIN activities a ON s.activities_id = a.id
            LEFT JOIN title t ON s.title_id = t.id
            ORDER BY p.name, a.name
        """)
        services = cursor.fetchall()
        conn.close()
        
        service_options = [f"{s[1]} - {s[2]} ({s[3]})" for s in services]
        service_var = tk.StringVar()
        service_dropdown = ctk.CTkComboBox(dialog, variable=service_var, values=service_options, width=400)
        service_dropdown.pack(pady=5)
        
        # Hours input
        ctk.CTkLabel(dialog, text="Estimated Hours:").pack(pady=(20, 5))
        hours_entry = ctk.CTkEntry(dialog, placeholder_text="Enter hours")
        hours_entry.pack(pady=5)
        
        # Rate input
        ctk.CTkLabel(dialog, text="Hourly Rate (USD):").pack(pady=(10, 5))
        rate_entry = ctk.CTkEntry(dialog, placeholder_text="Enter rate")
        rate_entry.pack(pady=5)
        
        # Date inputs
        ctk.CTkLabel(dialog, text="Start Date (YYYY-MM-DD):").pack(pady=(10, 5))
        start_date_entry = ctk.CTkEntry(dialog, placeholder_text="2025-01-01")
        start_date_entry.pack(pady=5)
        
        ctk.CTkLabel(dialog, text="End Date (YYYY-MM-DD):").pack(pady=(10, 5))
        end_date_entry = ctk.CTkEntry(dialog, placeholder_text="2025-12-31")
        end_date_entry.pack(pady=5)
        
        # Notes
        ctk.CTkLabel(dialog, text="Notes:").pack(pady=(10, 5))
        notes_entry = ctk.CTkTextbox(dialog, height=80)
        notes_entry.pack(pady=5, padx=20, fill="x")
        
        # Buttons
        button_frame = ctk.CTkFrame(dialog)
        button_frame.pack(pady=20)
        
        def save_assignment():
            try:
                # Get selected service ID
                selected_service = service_var.get()
                if not selected_service:
                    messagebox.showwarning("Warning", "Please select a service")
                    return
                
                service_id = services[[s[1] + " - " + s[2] + " (" + s[3] + ")" for s in services].index(selected_service)][0]
                
                # Get employee info from mapping
                selected_emp = self.selected_employee.get()
                if selected_emp not in self.employee_map:
                    messagebox.showerror("Error", "Employee not found")
                    return
                
                employee_info = self.employee_map[selected_emp]
                employee_id = employee_info["id"]
                employee_type = employee_info["type"]
                
                hours = float(hours_entry.get() or 0)
                rate = float(rate_entry.get() or 0)
                total_cost = hours * rate
                
                # Get selected project and technical unit IDs from mappings
                project_id = None
                tech_unit_id = None
                
                if self.selected_project.get() and self.selected_project.get() in self.project_map:
                    project_id = self.project_map[self.selected_project.get()]
                
                if self.selected_technical_unit.get() and self.selected_technical_unit.get() in self.technical_unit_map:
                    tech_unit_id = self.technical_unit_map[self.selected_technical_unit.get()]
                
                # Save to database
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                if employee_type == "extended":
                    # Get comprehensive employee data for the new booking
                    cursor.execute("""
                        SELECT cost_center, ghrs_id, COALESCE(first_name || ' ' || last_name, last_name, first_name, 'N/A') as employee_name, dept_description,
                               work_location, business_unit, tipo, tipo_description, sap_tipo,
                               saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                               monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                               remark
                        FROM employee_extended WHERE id = ?
                    """, (employee_id,))
                    emp_data = cursor.fetchone()
                    
                    # Fallback: get employee name from main employee table if extended doesn't have it
                    if not emp_data or not emp_data[2]:
                        cursor.execute("SELECT name FROM employee WHERE id = ?", (employee_id,))
                        emp_name_fallback = cursor.fetchone()
                        emp_name = emp_name_fallback[0] if emp_name_fallback else "N/A"
                    else:
                        emp_name = emp_data[2]
                    
                    # Get project, technical unit, and activity names
                    cursor.execute("SELECT name FROM project WHERE id = ?", (project_id,))
                    project_data = cursor.fetchone()
                    project_name_val = project_data[0] if project_data else "N/A"
                    
                    cursor.execute("SELECT name FROM technical_unit WHERE id = ?", (tech_unit_id,))
                    tu_data = cursor.fetchone()
                    tu_name_val = tu_data[0] if tu_data else "N/A"
                    
                    cursor.execute("""
                        SELECT a.name FROM service s 
                        LEFT JOIN activities a ON s.activities_id = a.id 
                        WHERE s.id = ?
                    """, (service_id,))
                    activity_data = cursor.fetchone()
                    activity_name_val = activity_data[0] if activity_data else "N/A"
                    
                    # Prepare values with defaults
                    if emp_data:
                        (cost_center, ghrs_id, employee_name_from_query, dept_description,
                         work_location, business_unit, tipo, tipo_description, sap_tipo,
                         saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                         monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                         remark) = emp_data
                    else:
                        # Default values when employee_extended data not available
                        (cost_center, ghrs_id, dept_description,
                         work_location, business_unit, tipo, tipo_description, sap_tipo,
                         saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                         monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                         remark) = (None, None, None, None, None, None, None, None,
                                   0.00, 0.00, 0.00, 0.00, 0, 0, 0.00, 0.00, None)
                    
                    cursor.execute("""
                        INSERT INTO project_bookings 
                        (employee_id, technical_unit_id, project_id, service_id, hourly_rate, total_cost, start_date, end_date,
                         booking_status, booking_date, created_at, updated_at,
                         cost_center, ghrs_id, employee_name, dept_description,
                         work_location, business_unit, tipo, tipo_description, sap_tipo,
                         saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                         monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                         remark, project_name, item, technical_unit_name, activities_name,
                         booking_hours, booking_cost_forecast, booking_period,
                         booking_hours_accepted, booking_period_accepted, booking_hours_extra)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Pending', CURRENT_DATE, 
                                CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                                ?, ?, ?, ?, ?, 0.00, 0.00, NULL, 0.00, NULL, 0.00)
                    """, (employee_id, tech_unit_id, project_id, service_id, hours, rate, 
                          total_cost, start_date_entry.get(), end_date_entry.get(),
                          cost_center, ghrs_id, emp_name, dept_description,
                          work_location, business_unit, tipo, tipo_description, sap_tipo,
                          saabu_rate_eur, saabu_rate_usd, local_agency_rate_usd, unit_rate_usd,
                          monthly_hours, annual_hours, workload_2025_planned, workload_2025_actual,
                          remark, project_name_val, tu_name_val, activity_name_val))
                
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Success", "Service assignment saved successfully")
                dialog.destroy()
                self.load_employee_services()
                # Auto refresh removed - user can manually refresh if needed
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save assignment: {e}")
                logging.error(f"Assignment save error: {e}")
        
        save_btn = ctk.CTkButton(button_frame, text="Save", command=save_assignment)
        save_btn.pack(side="left", padx=10)
        
        cancel_btn = ctk.CTkButton(button_frame, text="Cancel", command=dialog.destroy)
        cancel_btn.pack(side="left", padx=10)
    
    def edit_service_assignment(self):
        """Edit selected service assignment"""
        selected_item = self.service_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a service assignment to edit")
            return
        
        messagebox.showinfo("Info", "Edit functionality will be implemented")
    
    def delete_service_assignment(self):
        """Delete selected service assignment"""
        selected_item = self.service_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a service assignment to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this assignment?"):
            # Implementation for deletion
            messagebox.showinfo("Info", "Delete functionality will be implemented")
    
    def save_booking(self):
        """Save current booking"""
        if not self.selected_employee.get():
            messagebox.showwarning("Warning", "Please select an employee first")
            return
        
        messagebox.showinfo("Success", "Booking saved successfully")
    
    def approve_booking(self):
        """Approve current booking"""
        selected_item = self.service_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a booking to approve")
            return
        
        messagebox.showinfo("Success", "Booking approved successfully")
    
    def reject_booking(self):
        """Reject current booking"""
        selected_item = self.service_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a booking to reject")
            return
        
        messagebox.showinfo("Success", "Booking rejected")
    
    def import_excel_data(self):
        """Import employee data from Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        try:
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Map columns to database fields (based on 29-column structure)
            required_columns = [
                'Cost Center', 'GHRS ID', 'Last Name', 'First Name', 'Dept. Description',
                'Work Location', 'Business Unit', 'Tipo', 'Tipo Description', 'SAP Tipo',
                'SAABU Rate (EUR)', 'SAABU Rate (USD)', 'Local Agency Rate (USD)', 'Unit Rate (USD)',
                'Monthly Hours', 'Annual Hours', 'Workload 2025_Planned', 'Workload 2025_Actual',
                'Remark', 'Project', 'Item', 'Technical Unit', 'Activities', 'Booking Hours',
                'Booking Cost (Forecast)', 'Booking Period', 'Booking hours (Accepted by Project)',
                'Booking Period (Accepted by Project)', 'Booking hours (Extra)'
            ]
            
            # Check if columns exist
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messagebox.showwarning("Warning", f"Missing columns: {missing_columns}")
            
            # Import data
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            imported_count = 0
            for _, row in df.iterrows():
                try:
                    cursor.execute("""
                        INSERT INTO employee (
                            cost_center, ghrs_id, last_name, first_name, dept_description
                        ) VALUES (?, ?, ?, ?, ?)
                    """, (
                        row.get('cost_center', None),
                        row.get('ghrs_id', None), 
                        row.get('last_name', None),
                        row.get('first_name', None),
                        row.get('dept_description', None)
                    ))
                    imported_count += 1
                except Exception as e:
                    print(f"Error importing row: {e}")
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Success", f"Successfully imported {imported_count} employee records")
            self.load_data()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to import Excel data: {e}")
            logging.error(f"Excel import error: {e}")
    
    def export_report(self):
        """Export booking report to Excel with proper formatting like Fabsi app"""
        try:
            # Check if there's data to export
            if not hasattr(self, 'employee_tree') or not self.employee_tree.get_children():
                messagebox.showwarning("Warning", "No data to export")
                return
            
            file_path = filedialog.asksaveasfilename(
                title="Save report as",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not file_path:
                return
            
            # Get data from the current table view
            all_data = []
            columns = []
            
            # Get column headers
            if hasattr(self, 'employee_tree'):
                columns = list(self.employee_tree['columns'])
                
                # Get all rows from the tree
                for item in self.employee_tree.get_children():
                    values = self.employee_tree.item(item)['values']
                    all_data.append(values)
            
            if not all_data:
                messagebox.showwarning("Warning", "No data to export")
                return
            
            # Create DataFrame
            df = pd.DataFrame(all_data, columns=columns)
            
            # Remove the Select checkbox column from export
            if 'Select' in df.columns:
                df = df.drop('Select', axis=1)
            
            # Create Excel file with formatting
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Project Bookings')
                ws = writer.sheets['Project Bookings']
                
                # Apply formatting similar to Fabsi app
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                
                # Format header row
                for cell in ws[1]:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = header_fill
                    cell.border = border
                    cell.font = Font(bold=True)
                
                # Format data cells and auto-size columns
                for col in ws.columns:
                    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 25)
                    
                    for cell in col:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border
                
                # Add auto-filter
                ws.auto_filter.ref = ws.dimensions
                ws.row_dimensions[1].height = 48
            
            # Open the file automatically
            import subprocess
            subprocess.Popen(['start', '', file_path], shell=True)
            messagebox.showinfo("Success", f"Report exported to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report: {e}")
            logging.error(f"Export error: {e}")
            import traceback
            traceback.print_exc()
    
    def schedule_auto_refresh(self):
        """Schedule automatic refresh of employee data grid"""
        if self.auto_refresh_enabled:
            try:
                # FIXED: Use smart refresh that preserves filters and selections
                self.smart_refresh_with_preservation()
                # Schedule next refresh in 30 seconds (30000 milliseconds)
                self.root.after(30000, self.schedule_auto_refresh)
            except Exception as e:
                logging.error(f"Auto-refresh error: {e}")
                # Continue scheduling even if refresh fails
                self.root.after(30000, self.schedule_auto_refresh)
    
    def refresh_employee_data(self):
        """Refresh employee data grid and related displays - PRESERVES FILTERS"""
        try:
            # FIXED: Use smart refresh instead of full reload
            self.smart_refresh_with_preservation()
            # Also refresh service data if an employee is selected
            if self.selected_employee.get():
                self.load_employee_services()
        except Exception as e:
            logging.error(f"Data refresh error: {e}")
    
    def smart_refresh_with_preservation(self):
        """Smart refresh that preserves current filters and selections"""
        try:
            # Store current filter state
            current_filters = self.active_column_filters.copy()
            
            # Store current selections
            current_selections = {}
            for item in self.employee_tree.get_children():
                values = self.employee_tree.item(item, 'values')
                if values and len(values) > 1:
                    row_id = values[1]  # ID is in second column
                    is_selected = values[0] == "☑"
                    if is_selected:
                        current_selections[str(row_id)] = True
            
            # Reload data from database
            self.load_employee_data_grid_for_filter()
            
            # Reapply filters
            self.active_column_filters = current_filters
            for column, values in self.active_column_filters.items():
                if column in self.df.columns:
                    values_str = [str(v) for v in values]
                    self.df = self.df[self.df[column].astype(str).isin(values_str)]
            
            # Refresh display with preserved selections
            self.refresh_display_with_selections(current_selections)
            
            logging.info(f"Smart refresh completed - preserved {len(current_filters)} filters and {len(current_selections)} selections")
            
        except Exception as e:
            logging.error(f"Smart refresh error: {e}")
            # Fallback to regular refresh
            self.load_employee_data_grid()
    
    def manual_refresh(self):
        """Manual refresh triggered by user action"""
        try:
            self.refresh_employee_data()
            messagebox.showinfo("Refresh", "Employee data refreshed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh data: {e}")
    
    def toggle_auto_refresh(self):
        """Toggle automatic refresh on/off"""
        self.auto_refresh_enabled = not self.auto_refresh_enabled
        button_text = "Auto-Refresh: ON" if self.auto_refresh_enabled else "Auto-Refresh: OFF"
        self.auto_refresh_btn.configure(text=button_text)
        
        if self.auto_refresh_enabled:
            # Restart auto-refresh
            self.schedule_auto_refresh()
            messagebox.showinfo("Auto-Refresh", "Automatic refresh enabled (every 30 seconds)")
        else:
            messagebox.showinfo("Auto-Refresh", "Automatic refresh disabled")
    
    def smart_refresh(self):
        """Smart refresh that deletes rows with all zero values in specified fields"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Define the fields to check - all must be 0 for deletion
            fields_to_check = [
                'monthly_hours', 
                'annual_hours', 
                'workload_2025_planned', 
                'workload_2025_actual',
                'booking_hours', 
                'booking_period',  # This might be text, so we'll treat it differently
                'booking_hours_accepted', 
                'booking_period_accepted',  # This might be text, so we'll treat it differently
                'booking_hours_extra'
            ]
            
            # Build the WHERE clause to find rows where all numeric fields are 0 or NULL
            # For period fields, we check if they are NULL, empty, or contain only '0'
            numeric_conditions = []
            for field in fields_to_check:
                if 'period' in field:
                    # For period fields, check if NULL, empty, or just '0'
                    numeric_conditions.append(f"({field} IS NULL OR {field} = '' OR {field} = '0')")
                else:
                    # For numeric fields, check if 0 or NULL
                    numeric_conditions.append(f"({field} IS NULL OR {field} = 0)")
            
            where_clause = " AND ".join(numeric_conditions)
            
            # Find rows that match the deletion criteria
            query = f"""
                SELECT id, employee_name, project_name, technical_unit_name
                FROM project_bookings 
                WHERE {where_clause}
            """
            
            cursor.execute(query)
            rows_to_delete = cursor.fetchall()
            
            if rows_to_delete:
                # Ask for confirmation
                message = f"Found {len(rows_to_delete)} rows where all specified fields are zero.\n"
                message += "These rows will be deleted:\n\n"
                
                for i, row in enumerate(rows_to_delete[:5]):  # Show first 5 rows
                    message += f"- ID: {row[0]}, Employee: {row[1]}, Project: {row[2]}\n"
                
                if len(rows_to_delete) > 5:
                    message += f"... and {len(rows_to_delete) - 5} more rows\n"
                
                message += "\nDo you want to proceed with deletion?"
                
                if messagebox.askyesno("Confirm Deletion", message):
                    # Delete the identified rows
                    ids_to_delete = [str(row[0]) for row in rows_to_delete]
                    cursor.execute(f"DELETE FROM project_bookings WHERE id IN ({','.join(ids_to_delete)})")
                    
                    conn.commit()
                    
                    # Refresh the data display
                    self.load_employee_data_grid()
                    
                    messagebox.showinfo("Success", f"Deleted {len(rows_to_delete)} rows with all zero values.")
                else:
                    messagebox.showinfo("Cancelled", "Deletion cancelled by user.")
            else:
                # Just do regular refresh
                self.load_employee_data_grid()
                messagebox.showinfo("Refresh Complete", "No rows found with all zero values. Data refreshed.")
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to perform smart refresh: {e}")
            logging.error(f"Smart refresh error: {e}")
            import traceback
            traceback.print_exc()
    
    def toggle_row_selection(self, event):
        """Toggle row selection when clicking on the Select column"""
        try:
            # Get the clicked item and column
            region = self.employee_tree.identify_region(event.x, event.y)
            if region != "cell":
                return
                
            item = self.employee_tree.identify_row(event.y)
            column = self.employee_tree.identify_column(event.x)
            
            # Check if clicked on the Select column (first column)
            if column == "#1":  # First column is the Select column
                if item:
                    current_values = list(self.employee_tree.item(item, 'values'))
                    if current_values:
                        # Toggle the selection
                        if current_values[0] == "☑":
                            current_values[0] = "☐"
                            self.selected_rows.discard(item)
                        else:
                            current_values[0] = "☑"
                            self.selected_rows.add(item)
                        
                        self.employee_tree.item(item, values=current_values)
                        
        except Exception as e:
            logging.error(f"Row selection toggle error: {e}")
    
    def select_all_rows(self):
        """Select all rows in the CURRENT FILTERED VIEW (not all data)"""
        try:
            # Clear current selection set
            self.selected_rows.clear()
            
            # Iterate through ONLY THE VISIBLE ITEMS in the tree (filtered data)
            for item in self.employee_tree.get_children():
                current_values = list(self.employee_tree.item(item, 'values'))
                if current_values:
                    # Set checkbox to selected
                    current_values[0] = "☑"
                    self.employee_tree.item(item, values=current_values)
                    # Add to selected rows set
                    self.selected_rows.add(item)
            
            # DO NOT call apply_all_filters() here - this was causing the issue!
            print(f"Selected {len(self.selected_rows)} rows from current filtered view")
                    
        except Exception as e:
            logging.error(f"Select all rows error: {e}")
    
    def deselect_all_rows(self):
        """Deselect all rows in the employee tree"""
        try:
            # Iterate through all items in the tree
            for item in self.employee_tree.get_children():
                current_values = list(self.employee_tree.item(item, 'values'))
                if current_values:
                    # Set checkbox to unselected
                    current_values[0] = "☐"
                    self.employee_tree.item(item, values=current_values)
            
            # Clear the selected rows set
            self.selected_rows.clear()
                    
        except Exception as e:
            logging.error(f"Deselect all rows error: {e}")
    
    def delete_selected_rows(self):
        """Delete all selected rows"""
        try:
            if not self.selected_rows:
                messagebox.showwarning("Warning", "No rows selected for deletion")
                return
            
            # Get the IDs of selected rows
            ids_to_delete = []
            invalid_items = []
            
            for item in self.selected_rows:
                values = self.employee_tree.item(item, 'values')
                if values and len(values) > 1:  # Make sure we have values and ID is in position 1
                    item_id = values[1]  # ID is in the second column now
                    if item_id and str(item_id).strip():  # Check if ID is not empty
                        ids_to_delete.append(item_id)
                    else:
                        invalid_items.append(item)
                else:
                    invalid_items.append(item)
            
            if not ids_to_delete:
                messagebox.showwarning("Warning", "No valid rows selected for deletion")
                return
            
            # Check which IDs actually exist in the database
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Verify IDs exist in database
            existing_ids = []
            missing_ids = []
            
            for item_id in ids_to_delete:
                cursor.execute("SELECT id FROM project_bookings WHERE id = ?", (item_id,))
                if cursor.fetchone():
                    existing_ids.append(item_id)
                else:
                    missing_ids.append(item_id)
            
            if missing_ids:
                missing_str = ", ".join(missing_ids)
                logging.warning(f"Attempted to delete non-existent IDs: {missing_str}")
                
                if not existing_ids:
                    conn.close()
                    messagebox.showerror("Error", f"None of the selected items exist in the database. Missing IDs: {missing_str}")
                    # Refresh the data to sync with database
                    self.selected_rows.clear()
                    self.load_employee_data_grid()
                    return
                else:
                    # Some exist, some don't - ask user what to do
                    message = f"Some selected items don't exist in the database:\nMissing: {missing_str}\n\nDo you want to delete the {len(existing_ids)} existing items?"
                    if not messagebox.askyesno("Partial Deletion", message):
                        conn.close()
                        return
            
            if existing_ids:
                message = f"Are you sure you want to delete {len(existing_ids)} row(s)?"
                if messagebox.askyesno("Confirm Deletion", message):
                    # Delete only the existing IDs
                    placeholders = ','.join(['?' for _ in existing_ids])
                    cursor.execute(f"DELETE FROM project_bookings WHERE id IN ({placeholders})", existing_ids)
                    deleted_count = cursor.rowcount
                    
                    conn.commit()
                    conn.close()
                    
                    # Clear selection and refresh
                    self.selected_rows.clear()
                    self.load_employee_data_grid()
                    
                    if missing_ids:
                        messagebox.showinfo("Partial Success", f"Deleted {deleted_count} row(s) successfully.\n{len(missing_ids)} items were not found in the database.")
                    else:
                        messagebox.showinfo("Success", f"Deleted {deleted_count} row(s) successfully")
                else:
                    conn.close()
            else:
                conn.close()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete selected rows: {e}")
            logging.error(f"Selected rows deletion error: {e}")
            import traceback
            traceback.print_exc()
    
    def on_column_header_click(self, column):
        """Handle column header click for sorting and filtering"""
        try:
            # Right-click for filter, left-click for sort
            def show_context_menu(event):
                context_menu = tk.Menu(self.root, tearoff=0)
                context_menu.add_command(label=f"Sort {column} ↑", 
                                       command=lambda: self.sort_column(column, True))
                context_menu.add_command(label=f"Sort {column} ↓", 
                                       command=lambda: self.sort_column(column, False))
                context_menu.add_separator()
                context_menu.add_command(label=f"Filter {column}...", 
                                       command=lambda: self.show_column_filter(column))
                context_menu.post(event.x_root, event.y_root)
            
            # For now, just do simple sort on click
            if self.current_sort_column == column:
                # Toggle sort direction
                self.current_sort_ascending = not self.current_sort_ascending
            else:
                # New column, start with ascending
                self.current_sort_column = column
                self.current_sort_ascending = True
                
            self.sort_column(column, self.current_sort_ascending)
            
        except Exception as e:
            logging.error(f"Column header click error: {e}")
    
    def sort_column(self, column, ascending=True):
        """Sort the table by column"""
        try:
            if not hasattr(self, 'employee_tree') or not self.employee_tree.get_children():
                return
            
            # Get all data from the tree
            data = []
            for item in self.employee_tree.get_children():
                values = self.employee_tree.item(item)['values']
                data.append(values)
            
            # Find column index
            columns = list(self.employee_tree['columns'])
            if column not in columns:
                return
                
            col_idx = columns.index(column)
            
            # Sort data by the specified column
            def sort_key(row):
                value = row[col_idx] if col_idx < len(row) else ""
                # Try to convert to number if possible
                try:
                    return float(str(value).replace(',', '').replace('N/A', '0'))
                except:
                    return str(value).lower()
            
            data.sort(key=sort_key, reverse=not ascending)
            
            # Clear and repopulate the tree
            for item in self.employee_tree.get_children():
                self.employee_tree.delete(item)
                
            for row_data in data:
                self.employee_tree.insert("", "end", values=row_data)
            
            # Update column header to show sort direction
            for col in columns:
                if col == column:
                    symbol = " ↑" if ascending else " ↓"
                    self.employee_tree.heading(col, text=col + symbol)
                else:
                    self.employee_tree.heading(col, text=col,
                                             command=lambda c=col: self.on_column_header_click(c))
            
        except Exception as e:
            logging.error(f"Sort column error: {e}")
            import traceback
            traceback.print_exc()
    
    def show_filter_menu(self, column):
        """Show Excel-like filter menu for the selected column (based on Fabsi app)"""
        print(f"Opening filter for column: {column}")  # Debug
        
        # Store current filter column for use in other functions
        self.current_filter_column = column
        
        # Close any existing filter window
        if hasattr(self, 'filter_window') and self.filter_window:
            try:
                self.filter_window.destroy()
            except:
                pass
        
        # Create new filter window with proper layout
        self.filter_window = tk.Toplevel(self.root)
        self.filter_window.title(f"Filter - {column}")
        self.filter_window.geometry("350x550")  # Increased size for better visibility
        self.filter_window.resizable(False, False)
        self.filter_window.transient(self.root)
        self.filter_window.grab_set()
        
        # Position the window near the mouse cursor
        try:
            x = self.root.winfo_pointerx() - 175
            y = self.root.winfo_pointery() - 50
            self.filter_window.geometry(f"350x550+{x}+{y}")
        except:
            pass
        
        # Main container with fixed structure
        main_frame = tk.Frame(self.filter_window, bg='white')
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Title
        title_label = tk.Label(main_frame, text=f"Filter: {column}", 
                              font=('Arial', 12, 'bold'), bg='white')
        title_label.pack(pady=(0, 10))
        
        # Sort section
        sort_frame = tk.LabelFrame(main_frame, text="Sort Options", bg='white', font=('Arial', 9, 'bold'))
        sort_frame.pack(fill='x', pady=(0, 10))
        
        # Sort buttons
        sort_btn_frame = tk.Frame(sort_frame, bg='white')
        sort_btn_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Button(sort_btn_frame, text="↑ Sort A to Z", 
                 command=lambda: self.apply_sort(column, True),
                 bg='#003d52', fg='white', relief='flat', font=('Arial', 9),
                 width=15, pady=5).pack(side='left', padx=(0, 5))
        
        tk.Button(sort_btn_frame, text="↓ Sort Z to A",
                 command=lambda: self.apply_sort(column, False), 
                 bg='#003d52', fg='white', relief='flat', font=('Arial', 9),
                 width=15, pady=5).pack(side='left')
        
        # Filter section with FIXED height to ensure buttons show
        filter_frame = tk.LabelFrame(main_frame, text="Filter Values", bg='white', 
                                   font=('Arial', 9, 'bold'), height=320)
        filter_frame.pack(fill='x', pady=(0, 10))
        filter_frame.pack_propagate(False)  # Prevent expansion
        
        # Search box
        search_frame = tk.Frame(filter_frame, bg='white')
        search_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(search_frame, text="Search:", bg='white', font=('Arial', 9)).pack(anchor='w')
        
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=('Arial', 9))
        search_entry.pack(fill='x', pady=2)
        
        # Select All checkbox
        select_all_frame = tk.Frame(filter_frame, bg='white')
        select_all_frame.pack(fill='x', padx=5, pady=2)
        
        self.select_all_var = tk.BooleanVar(value=True)
        select_all_cb = tk.Checkbutton(select_all_frame, text="Select All", 
                                      variable=self.select_all_var,
                                      command=self.toggle_select_all,
                                      bg='white', font=('Arial', 9))
        select_all_cb.pack(anchor='w')
        
        # Separator line
        tk.Frame(filter_frame, height=1, bg='gray').pack(fill='x', padx=5, pady=2)
        
        # Values frame with checkboxes (Excel-like)
        values_frame = tk.Frame(filter_frame, bg='white')
        values_frame.pack(fill='both', expand=True, padx=5, pady=2)
        
        # Create scrollable frame for checkboxes
        canvas = tk.Canvas(values_frame, bg='white', highlightthickness=0)
        scrollbar = tk.Scrollbar(values_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Store checkbox variables
        self.checkbox_vars = {}
        
        # Populate checkboxes from current data
        self.populate_filter_checkboxes(scrollable_frame)
        
        # Bind search functionality
        self.search_var.trace('w', lambda *args: self.update_filter_search_checkboxes(scrollable_frame))
        
        # Bind mousewheel to canvas
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Action buttons - FIXED position at bottom
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(side='bottom', fill='x', pady=(10, 0))
        
        tk.Button(button_frame, text="OK", command=lambda: self.apply_filter(column),
                 bg='#003d52', fg='white', font=('Arial', 10, 'bold'),
                 width=8, pady=5).pack(side='left', padx=(0, 5))
        
        tk.Button(button_frame, text="Cancel", command=self.filter_window.destroy,
                 bg='#003d52', fg='white', font=('Arial', 10, 'bold'),
                 width=8, pady=5).pack(side='left', padx=(0, 5))
        
        tk.Button(button_frame, text="Clear Filter", command=lambda: self.clear_filter(column),
                 bg='#003d52', fg='white', font=('Arial', 10, 'bold'),
                 width=10, pady=5).pack(side='right')
    
    def populate_filter_checkboxes(self, parent_frame):
        """Populate the filter checkboxes (Excel-like)"""
        try:
            # Get unique values from the current data
            if hasattr(self, 'df') and hasattr(self, 'current_filter_column') and self.current_filter_column in self.df.columns:
                unique_values = sorted(self.df[self.current_filter_column].dropna().unique().tolist())
                self.filter_values = [str(val) for val in unique_values]
            else:
                # Fallback to getting values from the tree
                values = set()
                if hasattr(self, 'current_filter_column') and hasattr(self, 'employee_tree'):
                    try:
                        col_index = list(self.employee_tree['columns']).index(self.current_filter_column)
                        for item in self.employee_tree.get_children():
                            try:
                                val = self.employee_tree.item(item)['values'][col_index]
                                if val and str(val).strip():
                                    values.add(str(val))
                            except (IndexError, KeyError):
                                continue
                    except ValueError:
                        pass
                self.filter_values = sorted(list(values))
            
            # Clear existing checkboxes
            for widget in parent_frame.winfo_children():
                widget.destroy()
            self.checkbox_vars.clear()
            
            # Create checkboxes for each unique value
            for value in self.filter_values:
                var = tk.BooleanVar(value=True)  # All selected by default
                self.checkbox_vars[value] = var
                
                checkbox = tk.Checkbutton(
                    parent_frame, 
                    text=str(value), 
                    variable=var,
                    bg='white',
                    font=('Arial', 9),
                    anchor='w',
                    justify='left'
                )
                checkbox.pack(anchor='w', padx=5, pady=1, fill='x')
            
        except Exception as e:
            print(f"Error populating filter checkboxes: {e}")
            self.filter_values = []
    
    def update_filter_search_checkboxes(self, parent_frame):
        """Update filter checkboxes based on search text"""
        try:
            search_text = self.search_var.get().lower()
            
            # Clear existing checkboxes
            for widget in parent_frame.winfo_children():
                widget.destroy()
            
            # Filter values based on search
            filtered_values = [val for val in self.filter_values 
                             if search_text in str(val).lower()]
            
            # Recreate checkboxes for filtered values
            for value in filtered_values:
                if value in self.checkbox_vars:
                    var = self.checkbox_vars[value]
                else:
                    var = tk.BooleanVar(value=True)
                    self.checkbox_vars[value] = var
                
                checkbox = tk.Checkbutton(
                    parent_frame, 
                    text=str(value), 
                    variable=var,
                    bg='white',
                    font=('Arial', 9),
                    anchor='w',
                    justify='left'
                )
                checkbox.pack(anchor='w', padx=5, pady=1, fill='x')
                
        except Exception as e:
            print(f"Error updating filter search checkboxes: {e}")
    
    def toggle_select_all(self):
        """Toggle select all checkbox for filter checkboxes"""
        try:
            select_all_state = self.select_all_var.get()
            
            # Update all checkbox variables
            for value, var in self.checkbox_vars.items():
                var.set(select_all_state)
                
        except Exception as e:
            print(f"Error toggling select all: {e}")
    
    def apply_sort(self, column, ascending=True):
        """Apply sorting to the column"""
        try:
            if hasattr(self, 'df') and column in self.df.columns:
                self.df = self.df.sort_values(by=column, ascending=ascending, na_position='last')
                self.render_employee_table()
            else:
                # Fallback sorting using treeview
                self.sort_employee_data_column(column, ascending)
            
            # Close filter window
            if hasattr(self, 'filter_window') and self.filter_window:
                self.filter_window.destroy()
                
        except Exception as e:
            print(f"Error applying sort: {e}")
    
    def apply_filter(self, column):
        """Apply filter based on selected checkbox values"""
        try:
            # Get selected values from checkboxes
            selected_values = [value for value, var in self.checkbox_vars.items() if var.get()]
            
            if not selected_values:
                # If nothing selected, show all
                self.clear_filter(column)
                return
            
            # Apply filter to DataFrame if available
            if hasattr(self, 'df') and column in self.df.columns:
                self.df = self.df[self.df[column].astype(str).isin(selected_values)]
                self.render_employee_table()
            else:
                # Fallback: filter the treeview directly
                self.filter_treeview_by_column(column, selected_values)
            
            # Store active filter
            self.active_column_filters[column] = selected_values
            
            # Close filter window
            if hasattr(self, 'filter_window') and self.filter_window:
                self.filter_window.destroy()
                
        except Exception as e:
            print(f"Error applying filter: {e}")
    
    def clear_filter(self, column):
        """Clear filter for the specified column"""
        try:
            # Remove from active filters
            if column in self.active_column_filters:
                del self.active_column_filters[column]
            
            # Reset data to original
            if hasattr(self, 'original_df') and not self.original_df.empty:
                self.df = self.original_df.copy()
                # Re-apply any remaining filters
                for col, values in self.active_column_filters.items():
                    if col in self.df.columns:
                        self.df = self.df[self.df[col].astype(str).isin(values)]
                self.render_employee_table()
            else:
                # Fallback: reload all data
                self.load_employee_data_grid()
            
            # Close filter window
            if hasattr(self, 'filter_window') and self.filter_window:
                self.filter_window.destroy()
                
        except Exception as e:
            print(f"Error clearing filter: {e}")
    
    def filter_treeview_by_column(self, column, selected_values):
        """Filter treeview directly by hiding/showing items"""
        try:
            col_index = list(self.employee_tree['columns']).index(column)
            
            # Get all items
            all_items = self.employee_tree.get_children()
            
            # Hide items that don't match filter
            for item in all_items:
                try:
                    item_values = self.employee_tree.item(item)['values']
                    if len(item_values) > col_index:
                        cell_value = str(item_values[col_index])
                        if cell_value not in selected_values:
                            self.employee_tree.delete(item)
                except (IndexError, KeyError):
                    continue
                    
        except Exception as e:
            print(f"Error filtering treeview: {e}")
    
    def render_employee_table(self):
        """Render the employee table with current data"""
        try:
            # Clear existing items
            for item in self.employee_tree.get_children():
                self.employee_tree.delete(item)
            
            # Populate with filtered data
            if hasattr(self, 'df') and not self.df.empty:
                for i, row in self.df.iterrows():
                    # Prepare row values
                    row_values = []
                    for col in self.employee_tree['columns']:
                        if col in self.df.columns:
                            value = row[col]
                            # Handle special formatting
                            if col == 'Select':
                                value = '☑' if i in getattr(self, 'selected_rows', set()) else '☐'
                            elif pd.isna(value):
                                value = ""
                            else:
                                value = str(value)
                            row_values.append(value)
                        else:
                            row_values.append("")
                    
                    self.employee_tree.insert('', 'end', iid=str(i), values=row_values)
            
        except Exception as e:
            print(f"Error rendering employee table: {e}")
    
    def reset_filters(self):
        """Reset all filters and show original data"""
        try:
            from functools import partial
            
            # Clear all active filters
            self.active_column_filters.clear()
            
            # Reset to original data
            if hasattr(self, 'original_df') and not self.original_df.empty:
                self.df = self.original_df.copy()
                self.render_employee_table()
            else:
                self.load_employee_data_grid()
            
            # Update column headers to remove filter indicators
            for col in self.employee_tree['columns']:
                if col not in ["Select", "ID"]:
                    header_text = f"{col} ▼"
                    self.employee_tree.heading(col, text=header_text, 
                                             command=partial(self.show_filter_menu, col))
                    
        except Exception as e:
            print(f"Error resetting filters: {e}")

    def toggle_edit_mode(self):
        """Toggle edit mode between single and double click"""
        try:
            if self.edit_mode == "double":
                self.edit_mode = "single"
                self.edit_mode_btn.configure(text="Edit Mode: Single-Click")
                # Remove double-click binding and add single-click
                self.employee_tree.unbind('<Double-1>')
                self.employee_tree.bind('<Button-1>', self.on_simplified_cell_edit)
            else:
                self.edit_mode = "double"
                self.edit_mode_btn.configure(text="Edit Mode: Double-Click")
                # Remove single-click binding and add double-click
                self.employee_tree.unbind('<Button-1>')
                self.employee_tree.bind('<Double-1>', self.on_simplified_cell_edit)
                # Re-add checkbox functionality
                self.employee_tree.bind('<Button-1>', self.toggle_row_selection, add='+')
                
        except Exception as e:
            logging.error(f"Toggle edit mode error: {e}")
    
    def run(self):
        """Run the application"""
        self.root.mainloop()

def main():
    """Main function to run the Project Booking Application"""
    try:
        app = ProjectBookingApp()
        app.run()
    except Exception as e:
        print(f"Application error: {e}")
        logging.error(f"Application error: {e}")

if __name__ == "__main__":
    main()
