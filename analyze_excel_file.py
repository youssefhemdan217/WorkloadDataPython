import pandas as pd
import sqlite3
import openpyxl
from openpyxl import load_workbook
import os

def analyze_excel_file():
    """
    Analyze the test.xlsm file to:
    1. Review the BookingManHours tab columns
    2. Extract data for comparison with project booking table
    3. Check VBA code if accessible
    """
    
    excel_file_path = "test.xlsm"
    
    if not os.path.exists(excel_file_path):
        print("❌ Excel file not found!")
        return
    
    print("🔍 ANALYZING EXCEL FILE: test.xlsm")
    print("=" * 60)
    
    try:
        # Load the workbook to check sheets
        wb = load_workbook(excel_file_path, keep_vba=True)
        print(f"📊 Available worksheets: {wb.sheetnames}")
        
        # Check if BookingManHours sheet exists
        if 'BookingManHours' not in wb.sheetnames:
            print("❌ 'BookingManHours' sheet not found!")
            print(f"Available sheets: {wb.sheetnames}")
            return
        
        # Read the BookingManHours sheet
        print("\n📋 Reading BookingManHours sheet...")
        df = pd.read_excel(excel_file_path, sheet_name='BookingManHours')
        
        print(f"📏 Sheet dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
        
        # Display column information
        print("\n📑 EXCEL COLUMNS:")
        print("-" * 40)
        for i, col in enumerate(df.columns, 1):
            print(f"{i:2}. {col}")
        
        # Show sample data
        print("\n📊 SAMPLE DATA (first 5 rows):")
        print("-" * 80)
        print(df.head().to_string())
        
        # Check for VBA code
        print("\n🔧 VBA CODE ANALYSIS:")
        print("-" * 40)
        
        if hasattr(wb, 'vba_archive') and wb.vba_archive:
            print("✅ VBA code detected in the workbook")
            
            # Try to extract VBA modules
            try:
                # VBA code exists but extraction requires specialized tools
                print("📝 VBA modules detected but full extraction requires additional libraries")
                print("� Recommendation: Use Visual Studio Code with VBA extension or Excel VBA editor to view code")
            except Exception as e:
                print(f"VBA extraction error: {e}")
        else:
            print("❌ No VBA code detected or VBA not preserved")
        
        # Get current database table structure for comparison
        print("\n🗃️ COMPARING WITH DATABASE TABLE:")
        print("-" * 40)
        
        conn = sqlite3.connect('workload.db')
        cursor = conn.cursor()
        
        cursor.execute('PRAGMA table_info(project_bookings)')
        db_columns = cursor.fetchall()
        
        print("Current project_bookings table columns:")
        db_column_names = [col[1] for col in db_columns]
        for i, col_name in enumerate(db_column_names, 1):
            print(f"{i:2}. {col_name}")
        
        conn.close()
        
        # Compare columns
        excel_columns = df.columns.tolist()
        
        print(f"\n📊 COLUMN COMPARISON:")
        print("-" * 40)
        print(f"Excel columns: {len(excel_columns)}")
        print(f"Database columns: {len(db_column_names)}")
        
        # Find columns in Excel but not in DB
        excel_only = [col for col in excel_columns if col not in db_column_names]
        if excel_only:
            print("\n🔵 Columns in Excel but NOT in Database:")
            for col in excel_only:
                print(f"  • {col}")
        
        # Find columns in DB but not in Excel
        db_only = [col for col in db_column_names if col not in excel_columns]
        if db_only:
            print("\n🔴 Columns in Database but NOT in Excel:")
            for col in db_only:
                print(f"  • {col}")
        
        # Find matching columns
        matching = [col for col in excel_columns if col in db_column_names]
        if matching:
            print("\n🟢 Matching columns:")
            for col in matching:
                print(f"  • {col}")
        
        # Save analysis results
        analysis_results = {
            'excel_file': excel_file_path,
            'sheet_name': 'BookingManHours',
            'excel_columns': excel_columns,
            'db_columns': db_column_names,
            'excel_only_columns': excel_only,
            'db_only_columns': db_only,
            'matching_columns': matching,
            'data_shape': df.shape,
            'sample_data': df.head(3).to_dict('records') if not df.empty else []
        }
        
        return df, analysis_results
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def save_excel_data_for_import(df, analysis_results):
    """Save Excel data in a format ready for database import"""
    
    if df is None or df.empty:
        print("❌ No data to save")
        return
    
    # Save to CSV for easy review
    output_file = "excel_booking_data_for_import.csv"
    df.to_csv(output_file, index=False)
    print(f"\n💾 Excel data saved to: {output_file}")
    
    # Create SQL insert script
    sql_file = "import_excel_booking_data.sql"
    with open(sql_file, 'w') as f:
        f.write("-- SQL script to import Excel data into project_bookings table\n")
        f.write("-- Generated from test.xlsm BookingManHours sheet\n\n")
        
        f.write("-- First, clear existing data (optional)\n")
        f.write("-- DELETE FROM project_bookings;\n\n")
        
        f.write("-- Insert Excel data\n")
        
        for _, row in df.iterrows():
            values = []
            for col in df.columns:
                value = row[col]
                if pd.isna(value):
                    values.append("NULL")
                elif isinstance(value, str):
                    values.append(f"'{value.replace('\'', '\'\'')}'")
                else:
                    values.append(str(value))
            
            columns_str = ", ".join([f"[{col}]" for col in df.columns])
            values_str = ", ".join(values)
            f.write(f"INSERT INTO project_bookings ({columns_str}) VALUES ({values_str});\n")
    
    print(f"💾 SQL import script saved to: {sql_file}")

if __name__ == "__main__":
    df, analysis = analyze_excel_file()
    
    if df is not None and analysis is not None:
        save_excel_data_for_import(df, analysis)
        
        print("\n✅ ANALYSIS COMPLETE")
        print("Next steps:")
        print("1. Review the column comparison above")
        print("2. Check excel_booking_data_for_import.csv for data preview")
        print("3. Use import_excel_booking_data.sql to import data if needed")
    else:
        print("❌ Analysis failed")
