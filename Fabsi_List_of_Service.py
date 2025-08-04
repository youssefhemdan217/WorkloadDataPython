import os
import re
import logging
import sqlalchemy
logging.basicConfig(
    filename=os.path.join(os.path.dirname(__file__), 'app.log'),
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s %(message)s'
)
import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
import pandas as pd
import subprocess
import traceback
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, date
from PIL import Image, ImageTk

try:
    from tkcalendar import Calendar, DateEntry
    CALENDAR_AVAILABLE = True
except ImportError:
    CALENDAR_AVAILABLE = False
    print("tkcalendar not available. Installing...")
    try:
        subprocess.check_call(["pip", "install", "tkcalendar"])
        from tkcalendar import Calendar, DateEntry
        CALENDAR_AVAILABLE = True
        print("tkcalendar installed successfully!")
    except Exception as e:
        print(f"Failed to install tkcalendar: {e}")
        CALENDAR_AVAILABLE = False

# Set customtkinter appearance
ctk.set_appearance_mode("light")  # "light" or "dark"
ctk.set_default_color_theme("blue")  # "blue", "green", "dark-blue"

class CustomDatePicker(ctk.CTkFrame):
    """Custom date picker widget that integrates with CustomTkinter"""
    
    def __init__(self, parent, width=180, **kwargs):
        super().__init__(parent, **kwargs)
        
        self.width = width
        self.date_var = tk.StringVar()
        self.date_var.set("")
        
        # Create frame for the date picker
        self.configure(width=width, height=32)
        
        # Entry to display selected date
        self.date_entry = ctk.CTkEntry(
            self, 
            width=width-40, 
            placeholder_text="YYYY-MM-DD",
            font=ctk.CTkFont(family="Arial", size=11)
        )
        self.date_entry.pack(side='left', fill='x', expand=True)
        
        # Bind validation and keyboard shortcuts
        self.date_entry.bind('<FocusOut>', self.validate_date)
        self.date_entry.bind('<KeyRelease>', self.format_date_input)
        self.date_entry.bind('<Double-1>', lambda e: self.open_calendar())  # Double-click to open calendar
        self.date_entry.bind('<F4>', lambda e: self.open_calendar())  # F4 to open calendar (like Excel)
        
        # Calendar button
        self.calendar_btn = ctk.CTkButton(
            self,
            text="📅",
            width=30,
            height=32,
            command=self.open_calendar,
            font=ctk.CTkFont(family="Arial", size=12)
        )
        self.calendar_btn.pack(side='right', padx=(5, 0))
        
        # Make the frame non-expandable
        self.pack_propagate(False)
        
    def open_calendar(self):
        """Open calendar popup for date selection"""
        if not CALENDAR_AVAILABLE:
            messagebox.showwarning("Calendar Not Available", 
                                 "Calendar widget is not available. Please enter date manually in YYYY-MM-DD format.")
            return
            
        # Create popup window
        popup = tk.Toplevel(self)
        popup.title("Select Date")
        popup.geometry("350x400")
        popup.resizable(False, False)
        popup.transient(self.winfo_toplevel())
        popup.grab_set()
        
        # Center the popup
        popup.update_idletasks()
        x = (popup.winfo_screenwidth() // 2) - (350 // 2)
        y = (popup.winfo_screenheight() // 2) - (400 // 2)
        popup.geometry(f"350x400+{x}+{y}")
        
        # Create main frame for calendar
        cal_frame = tk.Frame(popup, bg='white')
        cal_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Create calendar widget - this will show the full calendar with years, months, and days
        cal = Calendar(
            cal_frame,
            font=('Arial', 12),
            selectmode='day',
            cursor="hand1",
            background='white',
            foreground='black',
            bordercolor='#22505f',
            headersbackground='#003d52',
            headersforeground='white',
            selectbackground='#ef8827',
            selectforeground='white',
            weekendbackground='#5b93a4',
            weekendforeground='white',
            othermonthforeground='#255c7b',
            othermonthbackground='white',
            showweeknumbers=False,
            firstweekday='monday',
            mindate=date(1990, 1, 1),  # Allow dates from 1990
            maxdate=date(2050, 12, 31),  # Allow dates until 2050
            date_pattern='yyyy-mm-dd'
        )
        cal.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Add double-click functionality to calendar
        def on_date_double_click(event):
            select_date()
        
        cal.bind('<Double-1>', on_date_double_click)
        
        # Set current date if entry has a value
        current_value = self.date_entry.get()
        if current_value:
            try:
                current_date = datetime.strptime(current_value, '%Y-%m-%d').date()
                cal.selection_set(current_date)
            except ValueError:
                # If invalid date, set to today
                cal.selection_set(date.today())
        else:
            # Set to today by default
            cal.selection_set(date.today())
        
        # Buttons frame
        btn_frame = tk.Frame(popup, bg='white')
        btn_frame.pack(side='bottom', fill='x', padx=15, pady=15)
        
        def select_date():
            selected_date = cal.selection_get()
            self.date_entry.delete(0, 'end')
            self.date_entry.insert(0, selected_date.strftime('%Y-%m-%d'))
            popup.destroy()
            
        def select_today():
            today = date.today()
            self.date_entry.delete(0, 'end')
            self.date_entry.insert(0, today.strftime('%Y-%m-%d'))
            popup.destroy()
            
        def clear_date():
            self.date_entry.delete(0, 'end')
            popup.destroy()
            
        def cancel():
            popup.destroy()
        
        # Add buttons with improved layout
        tk.Button(btn_frame, text="Select", command=select_date, 
                 bg='#003d52', fg='white', font=('Arial', 10, 'bold'),
                 width=8, pady=5, relief='raised', bd=2).pack(side='left', padx=2)
        tk.Button(btn_frame, text="Today", command=select_today,
                 bg='#5b93a4', fg='white', font=('Arial', 10, 'bold'),
                 width=8, pady=5, relief='raised', bd=2).pack(side='left', padx=2)
        tk.Button(btn_frame, text="Clear", command=clear_date,
                 bg='#ef8827', fg='white', font=('Arial', 10, 'bold'),
                 width=8, pady=5, relief='raised', bd=2).pack(side='left', padx=2)
        tk.Button(btn_frame, text="Cancel", command=cancel,
                 bg='#255c7b', fg='white', font=('Arial', 10, 'bold'),
                 width=8, pady=5, relief='raised', bd=2).pack(side='right', padx=2)
    
    def validate_date(self, event=None):
        """Validate the date format when user types manually"""
        date_text = self.date_entry.get().strip()
        if not date_text:
            return
            
        try:
            # Try to parse the date
            parsed_date = datetime.strptime(date_text, '%Y-%m-%d')
            # If successful, ensure it's properly formatted
            self.date_entry.delete(0, 'end')
            self.date_entry.insert(0, parsed_date.strftime('%Y-%m-%d'))
            self.date_entry.configure(text_color="black")
        except ValueError:
            # Invalid date format - highlight in red
            self.date_entry.configure(text_color="red")
    
    def format_date_input(self, event=None):
        """Auto-format date input as user types"""
        current_text = self.date_entry.get()
        
        # Remove any non-digit characters except hyphens
        cleaned = ''.join(c for c in current_text if c.isdigit() or c == '-')
        
        # Auto-add hyphens in the right places
        if len(cleaned) >= 4 and cleaned[4] != '-':
            cleaned = cleaned[:4] + '-' + cleaned[4:]
        if len(cleaned) >= 7 and cleaned[7] != '-':
            cleaned = cleaned[:7] + '-' + cleaned[7:]
        
        # Limit to 10 characters (YYYY-MM-DD)
        if len(cleaned) > 10:
            cleaned = cleaned[:10]
        
        # Update the entry if it changed
        if cleaned != current_text:
            cursor_pos = self.date_entry.index(tk.INSERT)
            self.date_entry.delete(0, 'end')
            self.date_entry.insert(0, cleaned)
            # Try to maintain cursor position
            try:
                self.date_entry.icursor(min(cursor_pos, len(cleaned)))
            except:
                pass
    
    def get(self):
        """Get the current date value"""
        return self.date_entry.get()
    
    def set(self, value):
        """Set the date value"""
        self.date_entry.delete(0, 'end')
        if value:
            self.date_entry.insert(0, str(value))
        # Validate after setting
        self.validate_date()
    
    def delete(self, start, end=None):
        """Delete content from the entry"""
        if end is None:
            end = 'end'
        self.date_entry.delete(start, end)
    
    def insert(self, index, value):
        """Insert value at index"""
        self.date_entry.insert(index, value)
        # Validate after inserting
        self.validate_date()

class ExcelActivityApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FABSI - List of Service")
        self.root.resizable(True, True)
        
        # Set window size and center it
        window_width = 1400
        window_height = 900
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")

        self.display_columns = [
            "Select", "ID", "Stick-Built", "Module", "Document Number", "Activities", "Title", "Department",
            "Technical Unit", "Assigned to", "Progress", "Estimated internal",
            "Estimated external", "Start date", "Due date", "Notes", "Professional Role"
        ]
        self.df = pd.DataFrame(columns=self.display_columns)
        self.original_df = self.df.copy()
        self.selected_rows = set()  # Track selected rows
        
        # Filter state management
        self.active_column_filters = {}  # Store active column filters
        self.refreshing_data = False  # Flag to control table rendering during data refresh

        self.file_path = ""
        self.tree = None
        self.entries = {}
        self.foreign_key_fields = [
            ("Stick-Built", "stickbuilts"),
            ("Module", "modules"),
            ("Activities", "activitiess"),
            ("Title", "titles"),
            ("Technical Unit", "technicalunits"),
            ("Assigned to", "employees"),
            ("Progress", "progresss"),
            ("Professional Role", "professionalunits")
        ]
        self.foreign_key_options = {}
        self.load_foreign_key_options_from_db()
        self.setup_ui()
    def load_foreign_key_options_from_db(self):
        # Print all available table names for debugging
        import sqlalchemy
        import re
        db_path = os.path.join(os.path.dirname(__file__), 'Workload.db')
        engine = sqlalchemy.create_engine(f'sqlite:///{db_path}')
        insp = sqlalchemy.inspect(engine)
        available_tables = insp.get_table_names()
        print("Available tables in database:", available_tables)
        # Try to map expected table names to actual table names
        def normalize(name):
            return re.sub(r'[^a-z0-9]', '', name.lower())
        # List of possible table name variants for each field
        table_name_variants = {
            "stickbuilts": ["stickbuilt", "stickbuilts"],
            "modules": ["module", "modules"],
            "activitiess": ["activities", "activity", "activitiess"],
            "titles": ["title", "titles"],
            "technicalunits": ["technicalunit", "technicalunits"],
            "employees": ["employee", "employees"],
            "progresss": ["progress", "progresss"],
            "professionalunits": ["professionalunit", "professionalunits"],
            "projects": ["project", "projects"]
        }
        # Build a mapping from endpoint to actual table name in DB
        endpoint_to_table = {}
        for endpoint, variants in table_name_variants.items():
            found = None
            for t in available_tables:
                t_norm = normalize(t)
                for v in variants:
                    if t_norm == normalize(v):
                        found = t
                        break
                if found:
                    break
            if found:
                endpoint_to_table[endpoint] = found
            else:
                endpoint_to_table[endpoint] = None
        print("Endpoint to DB table mapping:", endpoint_to_table)
        # Fetch dropdown options directly from SQLite using SQLAlchemy models
        from sqlalchemy.orm import sessionmaker
        from sqlalchemy.ext.automap import automap_base
        import traceback
        if not os.path.exists(db_path):
            from tkinter import messagebox
            messagebox.showerror("DB Error", f"Database file not found: {db_path}")
            print(f"Database file not found: {db_path}")
            return
        engine = sqlalchemy.create_engine(f'sqlite:///{db_path}')
        Base = automap_base()
        try:
            Base.prepare(autoload_with=engine)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("DB Error", f"Could not reflect database tables.\n{e}")
            print("Could not reflect database tables:", e)
            traceback.print_exc()
            return
        Session = sessionmaker(bind=engine)
        session = Session()
        for field, endpoint in self.foreign_key_fields:
            try:
                table_name = endpoint_to_table.get(endpoint)
                if not table_name:
                    self.foreign_key_options[field] = []
                    print(f"No table mapping for endpoint: {endpoint}")
                    continue
                if not hasattr(Base.classes, table_name):
                    self.foreign_key_options[field] = []
                    print(f"Table '{table_name}' not found in database.")
                    from tkinter import messagebox
                    messagebox.showwarning("DB Table Missing", f"Table '{table_name}' not found in database.")
                    continue
                Model = getattr(Base.classes, table_name)
                options = session.query(Model).all()
                if not options:
                    print(f"No data found in table '{table_name}'.")
                    from tkinter import messagebox
                    messagebox.showwarning("DB Table Empty", f"No data found in table '{table_name}'.")
                self.foreign_key_options[field] = [
                    {"id": getattr(opt, "id", None), "name": getattr(opt, "name", str(opt))} for opt in options
                ]
            except Exception as e:
                self.foreign_key_options[field] = []
                print(f"Error loading options for {field} ({endpoint}): {e}")
                traceback.print_exc()
        session.close()
        self.project_combobox = None
        self.current_project = None
        self.tree_edit_widgets = {}
        self.header_labels = []
        self.filter_vars = {}

        self.header_frame = None
        self.sum_frame = None
        self.total_label_internal = None
        self.total_label_external = None
        self.role_summary_frame = None
        self.role_summary_tree = None
        self.role_summary_data = pd.DataFrame()
        self.edit_popup = None
        self.dark_mode = False  # Track dark mode state

    def load_logo_image(self, image_path, width, height):
        """Load and resize logo image for display"""
        try:
            if os.path.exists(image_path):
                # Load and resize the image
                pil_image = Image.open(image_path)
                # Convert to RGBA if not already
                if pil_image.mode != 'RGBA':
                    pil_image = pil_image.convert('RGBA')
                # Resize maintaining aspect ratio
                pil_image.thumbnail((width, height), Image.Resampling.LANCZOS)
                # Convert to CTkImage
                return ctk.CTkImage(light_image=pil_image, dark_image=pil_image, size=(width, height))
            else:
                print(f"Logo image not found: {image_path}")
                return None
        except Exception as e:
            print(f"Error loading logo image {image_path}: {e}")
            return None

    def toggle_dark_mode(self):
        """Toggle between light and dark mode"""
        self.dark_mode = not self.dark_mode
        if self.dark_mode:
            ctk.set_appearance_mode("dark")
        else:
            ctk.set_appearance_mode("light")

    def setup_ui(self):
        # Create header frame - significantly reduced height
        header_frame = ctk.CTkFrame(self.root, height=40, corner_radius=10)
        header_frame.pack(fill='x', padx=10, pady=(2, 3))
        header_frame.pack_propagate(False)

        # Left logo (Saipem) - much smaller size
        left_logo_frame = ctk.CTkFrame(header_frame, width=60, height=35, corner_radius=5)
        left_logo_frame.pack(side='left', padx=5)
        left_logo_frame.pack_propagate(False)
        
        # Load Saipem logo
        saipem_logo_path = os.path.join(os.path.dirname(__file__), 'photos', 'saipem_logo.png')
        saipem_logo_image = self.load_logo_image(saipem_logo_path, 55, 30)
        
        if saipem_logo_image:
            left_logo_label = ctk.CTkLabel(left_logo_frame, image=saipem_logo_image, text="")
        else:
            left_logo_label = ctk.CTkLabel(left_logo_frame, text="SAIPEM", font=("Arial", 7), text_color="#003d52")
        left_logo_label.pack(expand=True)

        # Title in center - smaller font but still readable
        title_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_frame.pack(side='left', expand=True)
        title_label = ctk.CTkLabel(
            title_frame, 
            text="FABSI - List of Service",
            font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
            text_color="#003d52"
        )
        title_label.pack(expand=True)

        # Right logo (FABSI) - much smaller size
        right_logo_frame = ctk.CTkFrame(header_frame, width=60, height=35, corner_radius=5)
        right_logo_frame.pack(side='right', padx=5)
        right_logo_frame.pack_propagate(False)
        
        # Load FABSI logo
        fabsi_logo_path = os.path.join(os.path.dirname(__file__), 'photos', 'fabsi_logo.png')
        fabsi_logo_image = self.load_logo_image(fabsi_logo_path, 55, 30)
        
        if fabsi_logo_image:
            right_logo_label = ctk.CTkLabel(right_logo_frame, image=fabsi_logo_image, text="")
        else:
            right_logo_label = ctk.CTkLabel(right_logo_frame, text="FABSI", font=("Arial", 7), text_color="#ef8827")
        right_logo_label.pack(expand=True)

        # Add a separator - thinner
        separator_frame = ctk.CTkFrame(self.root, height=1, fg_color="#22505f")
        separator_frame.pack(fill='x', padx=10, pady=(0, 3))

        # Main content area
        main_top = ctk.CTkFrame(self.root, fg_color="transparent")
        main_top.pack(fill='x', padx=10, pady=1)

        # Professional, very compact form frame - use full available width for maximum space efficiency
        self.entry_frame = ctk.CTkFrame(main_top, height=80, width=1380, corner_radius=10)
        self.entry_frame.pack_propagate(False)
        self.entry_frame.pack(fill='x', padx=5, pady=3)

        # Button frame with all options - reduced spacing
        button_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        button_frame.pack(fill='x', padx=15, pady=(3, 3))
        
        # Left side buttons
        ctk.CTkButton(button_frame, text="Professional Role Summary", 
                     command=self.open_role_summary_modal, 
                     fg_color="#003d52", hover_color="#255c7b", width=160).pack(side='left', padx=(0, 10))
        ctk.CTkButton(button_frame, text="🔄 Clear All Filters", 
                     command=self.reset_filters, 
                     fg_color="#ef8827", hover_color="#22505f", width=130).pack(side='left', padx=(0, 10))
        ctk.CTkButton(button_frame, text="🗑️ Delete Selected", 
                     command=self.delete_selected, 
                     fg_color="#255c7b", hover_color="#22505f", width=130).pack(side='left', padx=(0, 10))
        ctk.CTkButton(button_frame, text="✅ Select All", 
                     command=self.select_all_rows, 
                     fg_color="#5b93a4", hover_color="#255c7b", width=100).pack(side='left', padx=(0, 10))
        ctk.CTkButton(button_frame, text="❌ Deselect All", 
                     command=self.deselect_all_rows, 
                     fg_color="#22505f", hover_color="#003d52", width=110).pack(side='left', padx=(0, 10))
        
        # Export buttons
        export_frame = ctk.CTkFrame(button_frame, corner_radius=8)
        export_frame.pack(side='left', padx=10)
        
        ctk.CTkLabel(export_frame, text="Export Data:", 
                    font=ctk.CTkFont(family='Arial', size=11, weight='bold')).pack(side='left', padx=5)
        
        ctk.CTkButton(export_frame, text="📊 Excel", 
                     command=self.save_to_excel,
                     fg_color="#5b93a4", hover_color="#255c7b",
                     width=80).pack(side='left', padx=5, pady=5)
        
        ctk.CTkButton(export_frame, text="📄 PDF",
                     command=self.save_to_pdf,
                     fg_color="#255c7b", hover_color="#22505f",
                     width=80).pack(side='left', padx=5, pady=5)
        
        # Right side buttons
        ctk.CTkButton(button_frame, text="🧹 Clear Form", 
                     command=self.clear_form, 
                     fg_color="#ef8827", hover_color="#22505f", width=100).pack(side='right', padx=(0, 10))
        ctk.CTkButton(button_frame, text="Agregar Actividad", 
                     command=self.add_row, 
                     fg_color="#5b93a4", hover_color="#255c7b", width=140).pack(side='right')
        ctk.CTkButton(button_frame, text="📁 Open Excel File", 
                     command=self.open_file, 
                     fg_color="#003d52", hover_color="#255c7b", width=130).pack(side='right', padx=(0, 10))

        # Remove the always-visible small table from the main UI
        # self.role_summary_frame = ctk.CTkFrame(summary_right)
        # self.role_summary_frame.pack(padx=5, pady=5, anchor='ne')
        # self.role_summary_tree = ...

        # Add a visual separator between form and table - thinner
        separator = ctk.CTkFrame(self.root, height=1, fg_color="#22505f")
        separator.pack(fill='x', padx=10, pady=2)

        # Add duplicate rows info label
        info_frame = ctk.CTkFrame(self.root, height=25, fg_color="transparent")
        info_frame.pack(fill='x', padx=15, pady=(2, 0))
        info_frame.pack_propagate(False)
        
        duplicate_info_label = ctk.CTkLabel(
            info_frame, 
            text="ℹ️ Yellow highlighted rows indicate duplicate entries (same values in all columns except Document Number)",
            font=ctk.CTkFont(family="Arial", size=11),
            text_color="#22505f"
        )
        duplicate_info_label.pack(side='left', padx=5)

        # Create the main table frame (always present) - reduced padding
        self.table_frame = ctk.CTkFrame(self.root, corner_radius=10)
        self.table_frame.pack(fill='both', expand=True, padx=15, pady=(2, 0))

        self.render_table()
        self.build_entry_fields()
        
        self.update_sum_labels()
        self.update_role_summary()

        # Start auto-updating totals every 400ms
        self.auto_update_totals()

        # Bottom frame with buttons (outside of scrollable area) - reduced spacing
        bottom_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        bottom_frame.pack(pady=5, fill='x')
        ctk.CTkButton(bottom_frame, text="Reset Filtros", command=self.reset_filters,
                     fg_color="#ef8827", hover_color="#22505f").pack(side='left', padx=10)
        ctk.CTkButton(bottom_frame, text="Save & Print in Excel", command=self.save_to_excel,
                     fg_color="#5b93a4", hover_color="#255c7b").pack(side='right', padx=10)
        ctk.CTkButton(bottom_frame, text="Save & Print in PDF", command=self.save_to_pdf,
                     fg_color="#255c7b", hover_color="#22505f").pack(side='right', padx=10)

    def on_checkbox_click(self, event):
        """Handle checkbox column clicks for row selection"""
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if item and column:
            col_idx = int(column.replace('#', '')) - 1
            if col_idx == 0 and 'Select' in self.df.columns:  # First column is Select
                row_idx = int(item)
                if row_idx in self.selected_rows:
                    self.selected_rows.remove(row_idx)
                else:
                    self.selected_rows.add(row_idx)
                self.render_table()

    def select_all_rows(self):
        """Select all visible rows"""
        self.selected_rows.clear()  # Clear first to ensure clean state
        # Get the actual number of rows in the current DataFrame
        total_rows = len(self.df)
        print(f"DataFrame has {total_rows} rows")  # Debug
        
        # Add all current row indices to selected_rows (from 0 to total_rows-1)
        for i in range(total_rows):
            self.selected_rows.add(i)
        
        print(f"Selected {len(self.selected_rows)} rows out of {total_rows}: {sorted(self.selected_rows)}")  # Debug
        self.render_table()  # Re-render to show checkmarks

    def deselect_all_rows(self):
        """Deselect all rows"""
        self.selected_rows.clear()
        self.render_table()
    def open_role_summary_modal(self):
        """Open a modal window showing the Professional Role summary table"""
        if hasattr(self, 'role_summary_modal') and self.role_summary_modal and tk.Toplevel.winfo_exists(self.role_summary_modal):
            self.role_summary_modal.lift()
            return

        # Create new modal window
        self.role_summary_modal = ctk.CTkToplevel(self.root)
        self.role_summary_modal.title("Professional Role & Hours Summary")
        self.role_summary_modal.geometry("450x500")
        self.role_summary_modal.transient(self.root)
        self.role_summary_modal.grab_set()
        self.role_summary_modal.resizable(False, False)

        # Add title label
        title_label = ctk.CTkLabel(
            self.role_summary_modal,
            text="Total Hours per Professional Role",
            font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
            text_color="#003d52"
        )
        title_label.pack(pady=(20, 15))

        # Create main frame
        frame = ctk.CTkFrame(self.role_summary_modal, corner_radius=10)
        frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # Create treeview with fixed column widths
        self.role_summary_tree = ttk.Treeview(
            frame,
            columns=["Professional Role", "Manhours"],
            show='headings',
            height=15,
            style="Summary.Treeview"
        )

        # Configure treeview style
        style = ttk.Style()
        style.configure("Summary.Treeview",
                       background="white",
                       foreground="black",
                       rowheight=25,
                       fieldbackground="white")
        style.configure("Summary.Treeview.Heading",
                       font=('Arial', 9, 'bold'),
                       background="#5b93a4")

        # Configure headers
        self.role_summary_tree.heading("Professional Role", text="Professional Role",
                                     command=lambda: self.sort_summary("Professional Role"))
        self.role_summary_tree.heading("Manhours", text="Internal Hours",
                                     command=lambda: self.sort_summary("Manhours"))

        # Configure columns
        self.role_summary_tree.column("Professional Role", width=250, anchor='w')
        self.role_summary_tree.column("Manhours", width=120, anchor='e')

        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.role_summary_tree.yview)
        self.role_summary_tree.configure(yscrollcommand=scrollbar.set)

        # Pack tree and scrollbar
        self.role_summary_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Update the summary data
        self.update_role_summary()

        # Add close button
        close_btn = ctk.CTkButton(
            self.role_summary_modal,
            text="Close",
            command=self.role_summary_modal.destroy,
            fg_color="#003d52",
            hover_color="#255c7b",
            width=120,
            height=35,
            font=ctk.CTkFont(family="Arial", size=12)
        )
        close_btn.pack(pady=15)

    def sort_column(self, column, ascending=True):
        """Sort the data by column"""
        if self.filter_popup:
            self.filter_popup.destroy()
        
        if column in self.df.columns:
            self.df = self.df.sort_values(by=column, ascending=ascending, na_position='last')
            self.render_table()
    
    def apply_column_filter(self, column, selected_indices):
        """Apply filter to the column based on selected values"""
        if not selected_indices:  # If nothing selected, show all
            self.clear_column_filter(column)
            return
            
        # Get selected values from listbox
        values = []
        for idx in selected_indices:
            values.append(self.current_filter_values[idx])
        
        # Apply filter to DataFrame
        self.df = self.df[self.df[column].astype(str).isin(values)]
        
        # Close popup and update display
        if self.filter_popup:
            self.filter_popup.destroy()
        self.render_table()
        self.update_sum_labels()
        self.update_role_summary()
    
    def clear_column_filter(self, column):
        """Clear filter for the column and show all values"""
        # Reset to original data
        display_cols_without_select_id = [col for col in self.display_columns if col not in ['Select', 'ID']]
        available_cols = [col for col in display_cols_without_select_id if col in self.original_df.columns]
        
        # Create a new DataFrame from original data
        self.df = self.original_df[available_cols].copy()
        
        # Add Select and ID columns
        self.df.insert(0, 'Select', False)
        self.df.insert(1, 'ID', range(1, len(self.df) + 1))
        
        # Close popup and update display
        if self.filter_popup:
            self.filter_popup.destroy()
        self.render_table()
        self.update_sum_labels()
        self.update_role_summary()

    def sort_summary(self, column):
        """Sort the role summary table by clicking on headers"""
        if not hasattr(self, 'role_summary_tree') or not self.role_summary_tree:
            return

        # Get all items except the total row
        items = [(self.role_summary_tree.set(item, column), item)
                for item in self.role_summary_tree.get_children('')]
        
        if not items:
            return

        # Remove the total row if it exists (last row)
        if items[-1][0] == "TOTAL":
            total_values = self.role_summary_tree.item(items[-1][1])['values']
            items.pop()
        else:
            total_values = None

        # Sort items
        items.sort(reverse=hasattr(self, '_summary_sort_reverse') and not self._summary_sort_reverse)
        self._summary_sort_reverse = not getattr(self, '_summary_sort_reverse', False)

        # Move items in the tree
        for idx, (_, item) in enumerate(items):
            self.role_summary_tree.move(item, '', idx)

        # Re-add total row at the bottom if it existed
        if total_values:
            self.role_summary_tree.insert("", "end", values=total_values, tags=('total',))

        # (moved to setup_ui)

    def auto_update_totals(self):
        self.update_sum_labels()
        self.update_role_summary()
        self.root.after(400, self.auto_update_totals)

    def update_sum_labels(self):
        # No longer need the separate labels since we're using table row
        # But keep the calculation for debugging and potential future use
        
        try:
            df = self.df
            col_internal = "Estimated internal"
            col_external = "Estimated external"
            
            # Calculate internal total - only numeric values
            total_internal = 0
            if col_internal in df.columns and len(df) > 0:
                internal_series = pd.to_numeric(df[col_internal], errors='coerce')
                total_internal = internal_series.dropna().sum()
            
            # Calculate external total - only numeric values
            total_external = 0
            if col_external in df.columns and len(df) > 0:
                external_series = pd.to_numeric(df[col_external], errors='coerce')
                total_external = external_series.dropna().sum()
                
        except Exception as e:
            print(f"Error calculating sums: {e}")
            total_internal = 0
            total_external = 0
        
        # Update the summation row in the table if it exists
        self.update_summation_row_in_table(total_internal, total_external)
        
        print(f"Updated totals: Internal={total_internal:.2f}, External={total_external:.2f}")  # Debug

    def update_summation_row_in_table(self, total_internal, total_external):
        """Update the summation row values in the table"""
        if not hasattr(self, 'tree') or not self.tree:
            return
            
        try:
            # Check if the summation row exists
            if self.tree.exists('TOTALS_ROW'):
                # Update the existing summation row
                row_values = []
                for col in self.df.columns:
                    if col == "Select":
                        row_values.append("")
                    # elif col == "ID":
                    #     row_values.append("TOTAL")
                    # elif col == "Activities":
                    #     row_values.append("📈 TOTAL PROJECT HOURS")
                    elif col == "Estimated internal":
                        row_values.append(f"{total_internal:,.2f}")
                    elif col == "Estimated external":
                        row_values.append(f"{total_external:,.2f}")
                    # elif col in ["Department"]:
                    #     row_values.append("ALL")
                    else:
                        row_values.append("")
                
                # Update the row values
                self.tree.item('TOTALS_ROW', values=row_values)
        except Exception as e:
            print(f"Error updating summation row: {e}")

    def update_role_summary(self):
        """Update the role summary based on the currently displayed/filtered data"""
        if not hasattr(self, 'role_summary_tree') or not self.role_summary_tree:
            return
        
        # Check if the tree widget still exists (hasn't been destroyed)
        try:
            self.role_summary_tree.get_children()
        except tk.TclError:
            # Tree has been destroyed, skip update
            return
        
        # Clear existing items
        for row in self.role_summary_tree.get_children():
            self.role_summary_tree.delete(row)
        
        # Calculate summary from current filtered data
        if "Professional Role" in self.df.columns and "Estimated internal" in self.df.columns:
            try:
                # Create summary from current filtered data
                summary = (
                    self.df.groupby("Professional Role")["Estimated internal"]
                    .apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
                    .reset_index()
                )
                
                # Clean up the summary data
                summary = summary[summary["Professional Role"].notnull() & (summary["Professional Role"] != "")]
                summary = summary[summary["Estimated internal"] > 0]  # Only show roles with hours
                summary = summary.sort_values("Estimated internal", ascending=False)  # Sort by hours descending
                self.role_summary_data = summary
                
                # Add rows to the tree
                total_hours = 0
                for _, row in summary.iterrows():
                    hours = row["Estimated internal"] if pd.notnull(row["Estimated internal"]) else 0
                    total_hours += hours
                    manhours = f"{hours:,.0f}" if hours > 0 else "0"
                    self.role_summary_tree.insert("", "end", values=(row["Professional Role"], manhours))
                
                # Add total row if there are entries
                if len(summary) > 0:
                    self.role_summary_tree.insert("", "end", values=("TOTAL", f"{total_hours:,.0f}"),
                                                tags=('total',))
                    self.role_summary_tree.tag_configure('total', background='#5b93a4',
                                                       font=('Arial', 9, 'bold'))
            except Exception as e:
                print(f"Error updating role summary: {e}")
                traceback.print_exc()

    def edit_role_summary_cell(self, event):
        # Permite editar la tabla resumen haciendo doble clic
        item = self.role_summary_tree.identify_row(event.y)
        column = self.role_summary_tree.identify_column(event.x)
        if not item or not column:
            return
        col_idx = int(column.replace('#', '')) - 1
        col_name = ["Professional Role", "Estimated internal"][col_idx]
        x, y, width, height = self.role_summary_tree.bbox(item, column)
        value = self.role_summary_tree.item(item, "values")[col_idx]
        # Evita más de una edición a la vez
        if self.edit_popup:
            self.edit_popup.destroy()
        self.edit_popup = tk.Entry(self.role_summary_tree, width=20)
        self.edit_popup.insert(0, value)
        self.edit_popup.place(x=x, y=y, width=width, height=height)
        self.edit_popup.focus()
        def on_focus_out(event):
            new_value = self.edit_popup.get()
            values = list(self.role_summary_tree.item(item, "values"))
            values[col_idx] = new_value
            self.role_summary_tree.item(item, values=values)
            self.edit_popup.destroy()
            self.edit_popup = None
        self.edit_popup.bind("<Return>", on_focus_out)
        self.edit_popup.bind("<FocusOut>", on_focus_out)

    def open_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.file_path = path
            if not self.current_project:
                messagebox.showinfo("Select Project", "Please select a project first.")
                return
            import sqlalchemy
            import pandas as pd
            db_path = os.path.join(os.path.dirname(__file__), 'Workload.db')
            engine = sqlalchemy.create_engine(f'sqlite:///{db_path}')
            # Get project_id
            with engine.connect() as conn:
                result = conn.execute(sqlalchemy.text('SELECT id FROM project WHERE name = :name'), {'name': self.current_project}).fetchone()
                project_id = result[0] if result else None
            if not project_id:
                messagebox.showerror("Error", "Project not found in database.")
                return
            try:
                # Try to load the sheet with the same name as the project, else fallback to first sheet
                try:
                    df = pd.read_excel(path, sheet_name=self.current_project)
                except Exception:
                    df = pd.read_excel(path)
                # Map Excel columns to DB columns
                field_to_db = {
                    "Stick-Built": "stick_built_id",
                    "Module": "module_id",
                    "Activities": "activities_id",
                    "Title": "title_id",
                    "Technical Unit": "technical_unit_id",
                    "Assigned to": "employee_id",
                    "Progress": "progress_id",
                    "Professional Role": "professional_unit_id",
                    "Department": "department",
                    "Estimated internal": "estimated_internal_hours",
                    "Estimated external": "estimated_external_hours",
                    "Start date": "start_date",
                    "Due date": "due_date",
                    "Notes": "notes"
                }
                # Build foreign key name->id maps for all dropdowns
                fk_maps = {}
                for col, options in self.foreign_key_options.items():
                    fk_maps[col] = {o['name']: o['id'] for o in options if o['id'] is not None}
                # Prepare rows for insert
                rows = []
                for _, row in df.iterrows():
                    data = {'project_id': project_id}
                    for col, db_col in field_to_db.items():
                        if db_col.endswith('_id'):
                            val = row.get(col, None)
                            data[db_col] = fk_maps.get(col, {}).get(val, None) if pd.notnull(val) else None
                        else:
                            data[db_col] = row.get(col, None)
                    rows.append(data)
                # Insert all rows in a transaction
                with engine.begin() as conn:
                    for data in rows:
                        columns = ', '.join(data.keys())
                        placeholders = ', '.join([f':{k}' for k in data.keys()])
                        insert_sql = f"INSERT INTO service ({columns}) VALUES ({placeholders})"
                        conn.execute(sqlalchemy.text(insert_sql), data)
                logging.info(f"Imported {len(rows)} rows from Excel to service table for project {self.current_project}")
                messagebox.showinfo("Import", f"Imported {len(rows)} rows successfully.")
                self.on_project_selected(None)
            except Exception as e:
                import traceback
                logging.error(f"Failed to import from Excel: {e}")
                logging.error(traceback.format_exc())
                messagebox.showerror("Import Error", f"Failed to import: {e}")

    def load_project_data(self, sheet_name):
        if not self.file_path:
            return
        try:
            df_loaded = pd.read_excel(self.file_path, sheet_name=sheet_name)
            cols_present = [c for c in self.display_columns if c in df_loaded.columns]
            extra_cols = [c for c in df_loaded.columns if c not in cols_present]
            df_loaded = df_loaded[cols_present + extra_cols]
            self.df = df_loaded
            self.original_df = self.df.copy()
            self.render_table()
            self.build_entry_fields()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar la hoja '{sheet_name}':\n{e}")

    def on_project_selected(self, choice=None):
        project_name = self.project_combobox.get()
        self.current_project = project_name
        # Fetch project ID directly from DB
        import sqlalchemy
        db_path = os.path.join(os.path.dirname(__file__), 'Workload.db')
        engine = sqlalchemy.create_engine(f'sqlite:///{db_path}')
        with engine.connect() as conn:
            result = conn.execute(sqlalchemy.text('SELECT id FROM project WHERE name = :name'), {'name': project_name}).fetchone()
            project_id = result[0] if result else None
        # Load services for this project directly from DB
        import pandas as pd
        try:
            if project_id:
                # Join foreign key tables to get display names
                query = '''
                SELECT
                    s.id,
                    sb.name AS "Stick-Built",
                    m.name AS "Module",
                    s.id AS "Document Number",
                    a.name AS "Activities",
                    t.name AS "Title",
                    s.department AS "Department",
                    tu.name AS "Technical Unit",
                    e.name AS "Assigned to",
                    p.name AS "Progress",
                    s.estimated_internal_hours AS "Estimated internal",
                    s.estimated_external_hours AS "Estimated external",
                    s.start_date AS "Start date",
                    s.due_date AS "Due date",
                    s.notes AS "Notes",
                    pu.name AS "Professional Role"
                FROM service s
                LEFT JOIN stick_built sb ON s.stick_built_id = sb.id
                LEFT JOIN module m ON s.module_id = m.id
                LEFT JOIN activities a ON s.activities_id = a.id
                LEFT JOIN title t ON s.title_id = t.id
                LEFT JOIN technical_unit tu ON s.technical_unit_id = tu.id
                LEFT JOIN employee e ON s.employee_id = e.id
                LEFT JOIN progress p ON s.progress_id = p.id
                LEFT JOIN professional_unit pu ON s.professional_unit_id = pu.id
                WHERE s.project_id = :project_id
                '''
                with engine.connect() as conn:
                    result = conn.execute(sqlalchemy.text(query), {'project_id': project_id})
                    rows = result.fetchall()
                    columns = result.keys()
                df = pd.DataFrame(rows, columns=columns)
                # Store original data WITHOUT Select and ID columns for filtering
                if not df.empty:
                    # Store clean original data
                    display_cols_without_select_id = [col for col in self.display_columns if col not in ['Select', 'ID']]
                    available_cols = [col for col in display_cols_without_select_id if col in df.columns]
                    self.original_df = df[available_cols].copy()
                    
                    # Add the Select column (checkbox) and ID counter for display
                    df.insert(0, 'Select', False)
                    df.insert(1, 'ID', range(1, len(df) + 1))
                    # Reorder to match display_columns
                    final_cols = ['Select', 'ID'] + available_cols
                    self.df = df[final_cols]
                else:
                    self.original_df = pd.DataFrame(columns=[col for col in self.display_columns if col not in ['Select', 'ID']])
                    self.df = pd.DataFrame(columns=self.display_columns)
                
                # Only render table if we're not in the middle of a data refresh
                if not self.refreshing_data:
                    self.render_table()
                self.update_sum_labels()  # Update totals after loading data
            else:
                self.df = pd.DataFrame(columns=self.display_columns)
                self.original_df = pd.DataFrame(columns=[col for col in self.display_columns if col not in ['Select', 'ID']])
                # Only render table if we're not in the middle of a data refresh
                if not self.refreshing_data:
                    self.render_table()
                self.update_sum_labels()  # Update totals even when empty
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load services for project: {e}")

    def build_entry_fields(self):
        for widget in self.entry_frame.winfo_children():
            widget.destroy()
        self.entries.clear()

        # Get project names from database
        db_path = os.path.join(os.path.dirname(__file__), 'Workload.db')
        engine = sqlalchemy.create_engine(f'sqlite:///{db_path}')
        with engine.connect() as conn:
            result = conn.execute(sqlalchemy.text('SELECT name FROM project')).fetchall()
            project_names = [row[0] for row in result]

        # Keep the existing project selection if it exists
        existing_project = None
        if hasattr(self, 'project_combobox') and self.project_combobox is not None:
            try:
                existing_project = self.project_combobox.get()
            except:
                existing_project = None

        # Configure grid columns for maximum space utilization - 7 columns to use all available width
        for i in range(7):
            self.entry_frame.grid_columnconfigure(i, weight=1)
        
        # Row 1: Select Project, Stick-Built, Module, Document Number, Title, Department, Estimated Internal
        row1_fields = ["Select Project", "Stick-Built", "Module", "Document Number", "Title", "Department", "Estimated internal"]
        # Row 2: Technical Unit, Assigned to, Progress, Professional Role, Start Date, Due Date, Estimated External
        row2_fields = ["Technical Unit", "Assigned to", "Progress", "Professional Role", "Start date", "Due date", "Estimated external"]
        # Row 3: Activities (spanning 4 columns), Notes (spanning 3 columns)
        
        # Define field widths optimized for the 7-column layout - very compact but readable
        field_widths = {
            "Select Project": 140,
            "Stick-Built": 90,
            "Module": 90,
            "Document Number": 120,
            "Title": 130,
            "Department": 90,
            "Estimated internal": 110,
            "Technical Unit": 130,
            "Assigned to": 130,
            "Progress": 90,
            "Professional Role": 130,
            "Start date": 100,
            "Due date": 100,
            "Estimated external": 110,
            "Activities": 500,  # Will span 4 columns
            "Notes": 380,  # Will span 3 columns - same size approach as Activities
        }
        
        # Create Row 1 fields
        for col_idx, field in enumerate(row1_fields):
            # Create label - optimized font size for blind people
            lbl = ctk.CTkLabel(self.entry_frame, text=field, 
                              font=ctk.CTkFont(family="Arial", size=11, weight="bold"),
                              anchor="w")
            lbl.grid(row=0, column=col_idx, sticky='w', padx=2, pady=(2,0))
            
            # Create field widget
            width = field_widths.get(field, 90)
            
            if field == "Select Project":
                # Project dropdown - first field as requested
                widget = ctk.CTkComboBox(self.entry_frame, width=width, state="readonly",
                                       font=ctk.CTkFont(family="Arial", size=11), height=26)
                widget.configure(values=project_names)
                if hasattr(self, 'current_project') and self.current_project:
                    widget.set(self.current_project)
                widget.configure(command=self.on_project_selected)
                self.project_combobox = widget
            elif field in self.foreign_key_options:
                # Use dropdown for foreign key fields
                options = self.foreign_key_options[field]
                widget = ctk.CTkComboBox(self.entry_frame, width=width, state="readonly",
                                       font=ctk.CTkFont(family="Arial", size=11), height=26)
                widget.configure(values=[o['name'] for o in options])
            elif field == "Department":
                # Use entry field with default value
                widget = ctk.CTkEntry(self.entry_frame, width=width,
                                    font=ctk.CTkFont(family="Arial", size=11), height=26)
                widget.insert(0, "FABSI")
            else:
                # Regular entry field
                widget = ctk.CTkEntry(self.entry_frame, width=width,
                                    font=ctk.CTkFont(family="Arial", size=11), height=26)
            
            self.entries[field] = widget
            widget.grid(row=1, column=col_idx, sticky='we', padx=2, pady=(0,1))
        
        # Create Row 2 fields
        for col_idx, field in enumerate(row2_fields):
            # Create label - optimized font size for blind people
            lbl = ctk.CTkLabel(self.entry_frame, text=field, 
                              font=ctk.CTkFont(family="Arial", size=11, weight="bold"),
                              anchor="w")
            lbl.grid(row=2, column=col_idx, sticky='w', padx=2, pady=(2,0))
            
            # Create field widget
            width = field_widths.get(field, 90)
            
            if field in self.foreign_key_options:
                # Use dropdown for foreign key fields
                options = self.foreign_key_options[field]
                widget = ctk.CTkComboBox(self.entry_frame, width=width, state="readonly",
                                       font=ctk.CTkFont(family="Arial", size=11), height=26)
                widget.configure(values=[o['name'] for o in options])
            elif field in ["Start date", "Due date"]:
                # Use regular entry field for date fields with placeholder
                widget = ctk.CTkEntry(self.entry_frame, width=width,
                                    font=ctk.CTkFont(family="Arial", size=11), height=26)
                widget.configure(placeholder_text="DD/MM/YYYY")
            else:
                # Regular entry field
                widget = ctk.CTkEntry(self.entry_frame, width=width,
                                    font=ctk.CTkFont(family="Arial", size=11), height=26)
            
            self.entries[field] = widget
            widget.grid(row=3, column=col_idx, sticky='we', padx=2, pady=(0,1))
        
        # Create Row 3: Activities (spanning columns 0-3) and Notes (spanning columns 4-6)
        # Activities label and field
        activities_lbl = ctk.CTkLabel(self.entry_frame, text="Activities", 
                                     font=ctk.CTkFont(family="Arial", size=11, weight="bold"),
                                     anchor="w")
        activities_lbl.grid(row=4, column=0, sticky='w', padx=2, pady=(2,0))
        
        # Special searchable combobox for Activities - spanning 4 columns
        activities_widget = ctk.CTkComboBox(self.entry_frame, width=500, state="normal",
                                          font=ctk.CTkFont(family="Arial", size=11), height=26)
        # Get existing activities from foreign key options if available
        if "Activities" in self.foreign_key_options:
            activity_options = [o['name'] for o in self.foreign_key_options["Activities"]]
            activities_widget.configure(values=activity_options)
        else:
            activities_widget.configure(values=[])
        
        # Set combo box to start empty
        activities_widget.set("")
        
        # Bind key release event for real-time search
        def on_activities_key_release(event, combo_widget=activities_widget):
            self.filter_activities_dropdown(combo_widget)
        
        activities_widget.bind('<KeyRelease>', on_activities_key_release)
        self.entries["Activities"] = activities_widget
        activities_widget.grid(row=5, column=0, columnspan=4, sticky='we', padx=2, pady=(0,1))
        
        # Notes label and field in columns 4-6 - same size approach as Activities
        notes_lbl = ctk.CTkLabel(self.entry_frame, text="Notes", 
                                font=ctk.CTkFont(family="Arial", size=11, weight="bold"),
                                anchor="w")
        notes_lbl.grid(row=4, column=4, sticky='w', padx=2, pady=(2,0))
        
        notes_widget = ctk.CTkEntry(self.entry_frame, width=380,
                                   font=ctk.CTkFont(family="Arial", size=11), height=26)
        self.entries["Notes"] = notes_widget
        notes_widget.grid(row=5, column=4, columnspan=3, sticky='we', padx=2, pady=(0,1))

    def update_activities_filter(self, event):
        text = self.entries["Activities"].get()
        all_activities = self.df["Activities"].dropna().unique().tolist() if "Activities" in self.df.columns else []
        if text:
            matches = [act for act in all_activities if text.lower() in act.lower()]
            self.entries["Activities"]['values'] = matches if matches else all_activities
        else:
            self.entries["Activities"]['values'] = all_activities

    def filter_activities_dropdown(self, combo_widget):
        """Filter the activities dropdown based on user input"""
        current_text = combo_widget.get().lower()
        
        # Get all available activities from foreign key options
        all_activities = []
        if "Activities" in self.foreign_key_options:
            all_activities = [o['name'] for o in self.foreign_key_options["Activities"]]
        
        # Also include unique activities from current data
        if "Activities" in self.df.columns:
            data_activities = self.df["Activities"].dropna().unique().tolist()
            # Combine and remove duplicates
            all_activities = list(set(all_activities + data_activities))
        
        if current_text:
            # Filter activities that contain the typed text
            filtered_activities = [
                activity for activity in all_activities 
                if current_text in activity.lower()
            ]
            combo_widget.configure(values=filtered_activities)
        else:
            # Show all activities when no text is entered
            combo_widget.configure(values=all_activities)

    def render_table(self):
        """Render the main table with data"""
        # Clear existing widgets
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        
        # Create main container
        main_container = ctk.CTkFrame(self.table_frame, fg_color="transparent")
        main_container.pack(fill='both', expand=True)
        
        # Create the container for our data table - minimal padding for more space
        self.scrollable_frame = ctk.CTkFrame(main_container, corner_radius=8)
        self.scrollable_frame.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Create and configure treeview with scrollbars
        tree_frame = ttk.Frame(self.scrollable_frame)
        tree_frame.pack(fill='both', expand=True)
        
        # Create Treeview - increased height to show more rows
        self.tree = ttk.Treeview(tree_frame, columns=list(self.df.columns), show='headings', height=25)
        
        # Create scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        # Configure grid weights
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        
        # Configure headers and columns with optimized widths for better fit
        column_widths = {
            "Select": 40,               # Compact checkbox visibility
            "ID": 45,                  # Compact ID numbers
            "Stick-Built": 85,         # Compact but readable
            "Module": 85,              # Compact but readable
            "Activities": 280,         # Reduced from 400 but still readable
            "Title": 120,              # Reduced but adequate for titles
            "Department": 80,          # Compact department names
            "Technical Unit": 120,     # Compact but readable
            "Assigned to": 120,        # Compact but readable names
            "Progress": 80,            # Compact progress status
            "Estimated internal": 85,  # Compact numbers with decimals
            "Estimated external": 85,  # Compact numbers with decimals
            "Start date": 85,          # Compact date format
            "Due date": 85,            # Compact date format
            "Notes": 200,              # Reduced from 300 but still useful
            "Professional Role": 150   # Reduced but readable role names
        }
        
        header_map = {
            "Select": "☑", "ID": "ID", "Stick-Built": "Stick-Built",
            "Module": "Module", "Activities": "Activities",
            "Title": "Title", "Department": "Department",
            "Technical Unit": "Tech. Unit", "Assigned to": "Assigned To",
            "Progress": "Progress", "Notes": "Notes",
            "Professional Role": "Professional Role",
            "Estimated internal": "Est. Internal",
            "Estimated external": "Est. External",
            "Start date": "Start Date", "Due date": "Due Date"
        }
        
        # Configure columns with improved readability and Excel-like filtering
        for col in self.df.columns:
            # Add filter arrow only for non-Select and non-ID columns
            if col not in ["Select", "ID"]:
                header_text = f"{header_map.get(col, col)} ▼"
                # Use functools.partial to avoid lambda closure issues
                from functools import partial
                self.tree.heading(col, text=header_text, command=partial(self.show_filter_menu, col))
            else:
                header_text = header_map.get(col, col)
                self.tree.heading(col, text=header_text)
            
            width = column_widths.get(col, 100)
            anchor = 'w' if col in ["Activities", "Title", "Notes", "Technical Unit", 
                                  "Assigned to", "Professional Role", "Department"] else 'center'
            self.tree.column(col, width=width, minwidth=40, stretch=True, anchor=anchor)
        
        # Style the treeview with reduced row height to show more rows
        style = ttk.Style()
        style.configure("Treeview",
                       background="white",
                       foreground="black",
                       rowheight=28,  # Reduced row height to show more rows
                       fieldbackground="white",
                       font=('Arial', 10))  # Keep font size readable for blind client
        
        style.configure("Treeview.Heading",
                       background="#5b93a4",
                       font=('Arial', 9, 'bold'))
        
        # Find duplicate rows before inserting data
        duplicate_indices = self.find_duplicate_rows()
        
        # Insert data
        for i, row in self.df.iterrows():
            # Determine row tag based on duplicate status and position
            if i in duplicate_indices:
                tag = 'duplicate_row'  # Yellow highlighting for duplicates
            else:
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            
            row_values = list(row)
            if len(row_values) > 0 and 'Select' in self.df.columns:
                checkbox_idx = self.df.columns.get_loc('Select')
                row_values[checkbox_idx] = '☑' if i in self.selected_rows else '☐'
            self.tree.insert('', 'end', iid=str(i), values=row_values, tags=(tag,))
        
        # ADD SUMMATION ROW as the last row of the table
        self.add_summation_row_to_table()
        
        # Configure row colors
        self.tree.tag_configure('oddrow', background='white')
        self.tree.tag_configure('evenrow', background='#5b93a4')
        # Configure duplicate row highlighting in yellow
        self.tree.tag_configure('duplicate_row', background='yellow', foreground='black', font=('Arial', 10, 'bold'))
        # Configure summation row with bright colors
        self.tree.tag_configure('total_row', background='#ef8827', foreground='black', font=('Arial', 12, 'bold'))
        
        # Bind events
        self.tree.bind("<Double-1>", self.edit_cell)
        self.tree.bind("<Button-1>", self.on_checkbox_click)
        
        # Update the totals after rendering the table
        self.update_sum_labels()

    def find_duplicate_rows(self):
        """
        Find duplicate rows based on all column values (excluding Select, ID, and Document Number columns)
        Returns a set of row indices that are duplicates
        """
        duplicate_indices = set()
        
        # Create a copy of the dataframe without Select, ID, and Document Number columns for comparison
        comparison_df = self.df.copy()
        
        # Remove columns that should not be considered for duplicate detection
        columns_to_exclude = ['Select', 'ID', 'Document Number']
        for col in columns_to_exclude:
            if col in comparison_df.columns:
                comparison_df = comparison_df.drop(col, axis=1)
        
        # Find duplicates based on remaining columns
        # Keep='False' marks all duplicates as True (both original and duplicates)
        duplicate_mask = comparison_df.duplicated(keep=False)
        
        # Get indices of duplicate rows
        duplicate_indices = set(self.df.index[duplicate_mask].tolist())
        
        return duplicate_indices

    def add_summation_row_to_table(self):
        """Add a summation row as the last row of the table"""
        # Calculate totals
        total_internal = 0
        total_external = 0
        
        try:
            if "Estimated internal" in self.df.columns and len(self.df) > 0:
                internal_series = pd.to_numeric(self.df["Estimated internal"], errors='coerce')
                total_internal = internal_series.dropna().sum()
            
            if "Estimated external" in self.df.columns and len(self.df) > 0:
                external_series = pd.to_numeric(self.df["Estimated external"], errors='coerce')
                total_external = external_series.dropna().sum()
        except Exception as e:
            print(f"Error calculating totals for summation row: {e}")
        
        # Create summation row values
        row_values = []
        for col in self.df.columns:
            if col == "Select":
                row_values.append("")  # Special icon for totals row
            # elif col == "ID":
            #     row_values.append("TOTAL")
            # elif col == "Activities":
            #     row_values.append("📈 TOTAL PROJECT HOURS")
            elif col == "Estimated internal":
                row_values.append(f"{total_internal:,.2f}")
            elif col == "Estimated external":
                row_values.append(f"{total_external:,.2f}")
            # elif col in ["Department"]:
            #     row_values.append("ALL")
            else:
                row_values.append("")  # Empty for other columns
        
        # Insert the summation row with special styling
        self.tree.insert('', 'end', iid='TOTALS_ROW', values=row_values, tags=('total_row',))

    def get_visible_columns(self):
        """Get list of columns to display (excluding Document Number)"""
        columns = [col for col in self.df.columns if col != "Document Number"]
        return columns

    def create_unified_header(self):
        """Create header with filters that scrolls with data"""
        # Reset any existing filter menus
        self.filter_menus = {}
        
        header_map = {
            "Select": "☑",
            "ID": "ID",
            "Estimated internal": "Est. Internal",
            "Estimated external": "Est. External", 
            "Start date": "Start Date",
            "Due date": "Due Date",
            "Stick-Built": "Stick-Built",
            "Module": "Module",
            "Activities": "Activities",
            "Title": "Title",
            "Department": "Department",
            "Technical Unit": "Tech. Unit",
            "Assigned to": "Assigned To",
            "Progress": "Progress",
            "Notes": "Notes",
            "Professional Role": "Professional Role"
        }
        
        column_widths = {
            "Select": 40,
            "ID": 45,
            "Stick-Built": 85,
            "Module": 85,
            "Activities": 280,    # Reduced but still readable
            "Title": 120,
            "Department": 80,
            "Technical Unit": 120,
            "Assigned to": 120,
            "Progress": 80,
            "Estimated internal": 85,
            "Estimated external": 85,
            "Start date": 85,
            "Due date": 85,
            "Notes": 200,
            "Professional Role": 150
        }
        
        header_height = 35  # Further reduced height
        header_frame = tk.Frame(self.scrollable_frame, height=header_height, bg="#5b93a4", relief='groove', bd=1)
        header_frame.pack(fill='x', pady=(0,2))
        header_frame.pack_propagate(False)
        
        self.header_labels = []
        self.filter_vars = {}
        self.filter_menus = {}  # Store filter menu buttons
        x_offset = 0
        
        for idx, col in enumerate(self.df.columns):
            txt = header_map.get(col, col)
            width = column_widths.get(col, 100)
            
            cell_frame = tk.Frame(header_frame, width=width, height=header_height, bg="#5b93a4", 
                                relief='groove', bd=1)
            cell_frame.place(x=x_offset, y=0, width=width, height=header_height)
            cell_frame.pack_propagate(False)
            
            # Header container with text and filter button
            header_container = tk.Frame(cell_frame, bg="#5b93a4")
            header_container.pack(fill='x', expand=True, pady=(2,0))
            
            # Header text label
            header_label = tk.Label(header_container, text=txt, bg="#5b93a4", 
                                  font=('Arial', 8, 'bold'), anchor='w')
            header_label.pack(side='left', padx=(2,0))
            
            # Filter button - only add for non-Select and non-ID columns
            if col not in ["Select", "ID"]:
                filter_btn = tk.Button(header_container, text="↓", font=('Arial', 7),
                                     width=2, height=1, relief='flat', bg="#5b93a4",
                                     command=lambda c=col: self.show_column_filter(c))
                filter_btn.pack(side='right', padx=(0,1))
                self.filter_menus[col] = filter_btn
            
            x_offset += width
            
    def show_column_filter(self, column):
        """Show Excel-like filter popup for the column"""
        # Close existing popup if any
        if hasattr(self, 'filter_popup') and self.filter_popup:
            self.filter_popup.destroy()
        
        # Get button widget and position
        filter_btn = self.filter_menus[column]
        x = filter_btn.winfo_rootx()
        y = filter_btn.winfo_rooty() + filter_btn.winfo_height()
        
        # Create popup window
        self.filter_popup = ctk.CTkToplevel(self.root)
        self.filter_popup.wm_overrideredirect(True)
        self.filter_popup.geometry(f"220x380+{x}+{y}")  # Made taller to fit all elements
        
        # Main container frame
        main_frame = ctk.CTkFrame(self.filter_popup, corner_radius=8)
        main_frame.pack(fill='both', expand=True, padx=3, pady=3)
        
        # Store current filter values for search
        self.current_filter_values = []
        if column in self.df.columns:
            self.current_filter_values = sorted(self.df[column].unique().tolist())
            self.current_filter_values = [str(val) for val in self.current_filter_values if pd.notnull(val)]
        
        # Frame for sort options at the top
        sort_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        sort_frame.pack(fill='x', pady=(5, 5))
        
        # Sort buttons
        ctk.CTkButton(sort_frame, text="Sort A → Z", 
                     command=lambda: self.sort_column(column, ascending=True),
                     height=30, font=ctk.CTkFont(size=10)).pack(fill='x', pady=2)
        ctk.CTkButton(sort_frame, text="Sort Z → A",
                     command=lambda: self.sort_column(column, ascending=False),
                     height=30, font=ctk.CTkFont(size=10)).pack(fill='x', pady=2)
        
        # Separator
        separator_frame = ctk.CTkFrame(main_frame, height=2, fg_color="#22505f")
        separator_frame.pack(fill='x', pady=5)
        
        # Filter options frame
        filter_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        filter_frame.pack(fill='both', expand=True, padx=5)
        
        # Search entry
        search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(filter_frame, textvariable=search_var, 
                                   placeholder_text="Search...",
                                   font=ctk.CTkFont(family='Arial', size=10))
        search_entry.pack(fill='x', pady=(0, 5))
        
        # Checkboxes container with scrollbar
        checkbox_container = ctk.CTkScrollableFrame(filter_frame, height=200)
        checkbox_container.pack(fill='both', expand=True)
        
        # Dictionary to store checkbox variables
        self.current_column_filter_vars = {}
        
        # Add checkboxes
        for val in self.current_filter_values:
            var = tk.BooleanVar()
            self.current_column_filter_vars[val] = var
            cb = ctk.CTkCheckBox(checkbox_container, text=str(val), variable=var,
                               font=ctk.CTkFont(family='Arial', size=9))
            cb.pack(anchor='w', padx=5, pady=2)
        
        # Update filter options based on search
        def update_search(*args):
            search_text = search_var.get().lower()
            for widget in checkbox_container.winfo_children():
                if hasattr(widget, 'cget'):
                    try:
                        if search_text in widget.cget('text').lower():
                            widget.pack(anchor='w', padx=5, pady=2)
                        else:
                            widget.pack_forget()
                    except:
                        pass
        
        search_var.trace('w', update_search)
        
        # Buttons frame at the bottom
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(side='bottom', fill='x', padx=5, pady=(5, 10))
        
        # Pack buttons with equal spacing
        ctk.CTkButton(btn_frame, text="Apply", 
                     command=lambda: self.apply_column_filter(column),
                     fg_color='#003d52', hover_color='#255c7b',
                     width=60, height=30, font=ctk.CTkFont(size=9)).pack(side='left', expand=True, padx=2)
        ctk.CTkButton(btn_frame, text="Clear",
                     command=lambda: self.clear_column_filter(column),
                     fg_color='#ef8827', hover_color='#22505f',
                     width=60, height=30, font=ctk.CTkFont(size=9)).pack(side='left', expand=True, padx=2)
        ctk.CTkButton(btn_frame, text="Cancel",
                     command=self.filter_popup.destroy,
                     fg_color='#255c7b', hover_color='#22505f',
                     width=60, height=30, font=ctk.CTkFont(size=9)).pack(side='left', expand=True, padx=2)
        
        # Handle click outside popup
        def on_click_outside(event):
            if event.widget == self.filter_popup:
                self.filter_popup.destroy()
        
        self.filter_popup.bind('<Button-1>', on_click_outside)
        
        # Set focus to search entry
        search_entry.focus()

    def update_filter_list(self, search_text, listbox):
        """Update the filter listbox based on search text"""
        listbox.delete(0, tk.END)
        items = [item for item in self.current_filter_values 
                if str(search_text).lower() in str(item).lower()]
        for item in sorted(items):
            listbox.insert(tk.END, str(item))

    def sort_column(self, column, ascending=True):
        """Sort the data by column"""
        if self.filter_popup:
            self.filter_popup.destroy()
        
        if column in self.df.columns:
            self.df = self.df.sort_values(by=column, ascending=ascending, na_position='last')
            self.render_table()

    def apply_column_filter(self, column, *args):
        """Apply filter to the column based on selected checkboxes"""
        # Get selected values from checkboxes
        selected_values = [val for val, var in self.current_column_filter_vars.items() if var.get()]
        
        if not selected_values:  # If nothing selected, show all
            # Remove this column's filter
            if column in self.active_column_filters:
                del self.active_column_filters[column]
            self.clear_column_filter(column)
            return
        
        # Store the filter state for this column
        self.active_column_filters[column] = selected_values
        print(f"Stored filter for {column}: {selected_values}")  # Debug
        
        # Apply all active filters starting from original data
        self.apply_all_active_filters()
        
        if self.filter_popup:
            self.filter_popup.destroy()
            
        self.update_sum_labels()
        self.update_role_summary()

    def clear_column_filter(self, column):
        """Clear filter for the column"""
        # Remove this column's filter from active filters
        if column in self.active_column_filters:
            del self.active_column_filters[column]
            print(f"Cleared filter for {column}")  # Debug
        
        # Apply all remaining active filters
        self.apply_all_active_filters()
        
        if self.filter_popup:
            self.filter_popup.destroy()
            
        self.update_sum_labels()
        self.update_role_summary()

    def apply_all_active_filters(self):
        """Apply all active filters (both dropdown and column filters) to the data"""
        # Start with original data without Select and ID columns
        df_original_clean = self.original_df.copy()
        
        # Remove Select and ID columns if they exist in original_df
        if 'Select' in df_original_clean.columns:
            df_original_clean = df_original_clean.drop('Select', axis=1)
        if 'ID' in df_original_clean.columns:
            df_original_clean = df_original_clean.drop('ID', axis=1)
        
        df_filtered = df_original_clean.copy()
        
        # Apply dropdown filters first (if any exist)
        if hasattr(self, 'filter_vars'):
            for col, var in self.filter_vars.items():
                val = var.get()
                if val != "Todos" and col not in ["Select", "ID"] and col in df_filtered.columns:
                    df_filtered = df_filtered[df_filtered[col].astype(str) == val]
                    print(f"Applied dropdown filter {col}: {val}")  # Debug
        
        # Apply column filters (checkbox filters)
        for column, selected_values in self.active_column_filters.items():
            if column in df_filtered.columns and selected_values:
                mask = df_filtered[column].astype(str).isin([str(v) for v in selected_values])
                df_filtered = df_filtered[mask]
                print(f"Applied column filter {column}: {selected_values}")  # Debug
        
        # Re-add Select and ID columns
        if not df_filtered.empty:
            df_filtered.insert(0, 'Select', False)
            df_filtered.insert(1, 'ID', range(1, len(df_filtered) + 1))
        else:
            df_filtered = pd.DataFrame(columns=self.display_columns)
            
        self.df = df_filtered
        self.render_table()
        print(f"Applied all filters, resulting data shape: {df_filtered.shape}")  # Debug

    def create_data_rows(self):
        """Create the data rows section"""
        # Enhanced treeview styling
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background="white",
            foreground="black",
            rowheight=25,
            fieldbackground="white",
            bordercolor="#BFBFBF",
            borderwidth=1,
            font=('Arial', 9)
        )
        
        # Create treeview for data
        self.tree = ttk.Treeview(
            self.scrollable_frame,
            columns=list(self.df.columns),
            show='',  # Hide native headers
            selectmode='extended',
            height=15  # Limit height so summation is visible
        )
        self.tree.pack(fill='x', pady=(0,2))
        
        self.tree.bind("<Double-1>", self.edit_cell)
        self.tree.bind("<Button-1>", self.on_checkbox_click)
        
        # Column widths (optimized for better table fit)
        column_widths = {
            "Select": 40, "ID": 45, "Stick-Built": 85, "Module": 85, "Document Number": 65,
            "Activities": 280, "Title": 120, "Department": 80, "Technical Unit": 120,
            "Assigned to": 120, "Progress": 80, "Estimated internal": 85,
            "Estimated external": 85, "Start date": 85, "Due date": 85,
            "Notes": 200, "Professional Role": 150
        }
        
        for col in self.df.columns:
            anchor = 'w' if col in ["Activities", "Technical Unit", "Assigned to", "Professional Role"] else 'c'
            self.tree.column(col, width=column_widths.get(col, 120), anchor=anchor, stretch=False)
        
        # Add data rows
        for i, row in self.df.iterrows():
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            row_values = list(row)
            if len(row_values) > 0 and 'Select' in self.df.columns:
                checkbox_idx = self.df.columns.get_loc('Select')
                row_values[checkbox_idx] = '☑' if i in self.selected_rows else '☐'
            self.tree.insert('', 'end', iid=str(i), values=row_values, tags=(tag,))
        
        # Row styling
        self.tree.tag_configure('oddrow', background='white')
        self.tree.tag_configure('evenrow', background='#5b93a4')

    def show_filter_menu(self, column):
        """Show Excel-like filter menu for the selected column"""
        print(f"Opening filter for column: {column}")  # Debug
        
        # Close any existing filter window
        if hasattr(self, 'filter_window') and self.filter_window:
            try:
                self.filter_window.destroy()
            except:
                pass
        
        # Create new filter window with proper layout
        self.filter_window = tk.Toplevel(self.root)
        self.filter_window.title(f"Filter - {column}")
        self.filter_window.geometry("350x550")  # Increased size for better visibility
        self.filter_window.resizable(False, False)
        self.filter_window.transient(self.root)
        self.filter_window.grab_set()
        
        # Position the window near the mouse cursor
        try:
            x = self.root.winfo_pointerx() - 175
            y = self.root.winfo_pointery() - 50
            self.filter_window.geometry(f"350x550+{x}+{y}")
        except:
            pass
        
        # Main container with fixed structure
        main_frame = tk.Frame(self.filter_window, bg='white')
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Title
        title_label = tk.Label(main_frame, text=f"Filter: {column}", 
                              font=('Arial', 12, 'bold'), bg='white')
        title_label.pack(pady=(0, 10))
        
        # Sort section
        sort_frame = tk.LabelFrame(main_frame, text="Sort Options", bg='white', font=('Arial', 9, 'bold'))
        sort_frame.pack(fill='x', pady=(0, 10))
        
        # Sort buttons
        sort_btn_frame = tk.Frame(sort_frame, bg='white')
        sort_btn_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Button(sort_btn_frame, text="↑ Sort A to Z", 
                 command=lambda: self.apply_sort(column, True),
                 bg='#5b93a4', relief='flat', font=('Arial', 9),
                 width=15, pady=5).pack(side='left', padx=(0, 5))
        
        tk.Button(sort_btn_frame, text="↓ Sort Z to A",
                 command=lambda: self.apply_sort(column, False), 
                 bg='#5b93a4', relief='flat', font=('Arial', 9),
                 width=15, pady=5).pack(side='left')
        
        # Filter section with FIXED height to ensure buttons show
        filter_frame = tk.LabelFrame(main_frame, text="Filter Values", bg='white', 
                                   font=('Arial', 9, 'bold'), height=320)
        filter_frame.pack(fill='x', pady=(0, 10))
        filter_frame.pack_propagate(False)  # Prevent expansion
        
        # Search box
        search_frame = tk.Frame(filter_frame, bg='white')
        search_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(search_frame, text="Search:", bg='white', font=('Arial', 9)).pack(anchor='w')
        
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=('Arial', 9))
        search_entry.pack(fill='x', pady=2)
        
        # Select All checkbox
        select_all_frame = tk.Frame(filter_frame, bg='white')
        select_all_frame.pack(fill='x', padx=5, pady=2)
        
        self.select_all_var = tk.BooleanVar(value=True)
        select_all_cb = tk.Checkbutton(select_all_frame, text="Select All", 
                                      variable=self.select_all_var, bg='white',
                                      font=('Arial', 9, 'bold'),
                                      command=self.toggle_all_checkboxes)
        select_all_cb.pack(anchor='w')
        
        # Values list with scrollbar - FIXED HEIGHT for proper scrolling
        list_frame = tk.Frame(filter_frame, bg='white', height=220)
        list_frame.pack(fill='x', padx=5, pady=5)
        list_frame.pack_propagate(False)  # Prevent expansion
        
        # Create scrollable frame for checkboxes
        canvas = tk.Canvas(list_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Get distinct values for the column
        if hasattr(self, 'original_df') and column in self.original_df.columns:
            # Use original data to get all possible values
            all_values = self.original_df[column].dropna().astype(str).unique().tolist()
        elif column in self.df.columns:
            # Fallback to current data
            all_values = self.df[column].dropna().astype(str).unique().tolist()
        else:
            all_values = []
        
        # Sort values naturally
        def natural_sort_key(val):
            try:
                return (0, float(val))  # Numbers first
            except (ValueError, TypeError):
                return (1, str(val).lower())  # Then text
        
        all_values = sorted(all_values, key=natural_sort_key)
        self.filter_values = all_values
        
        # Store checkbox variables
        self.filter_checkboxes = {}
        self.current_filter_column = column
        
        # Create checkboxes for each value
        for value in all_values:
            var = tk.BooleanVar(value=True)
            self.filter_checkboxes[value] = var
            
            cb = tk.Checkbutton(scrollable_frame, text=f" {value}", 
                              variable=var, bg='white', anchor='w',
                              font=('Arial', 9))
            cb.pack(fill='x', anchor='w', pady=1)
        
        # Pack canvas and scrollbar properly
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Search functionality
        def update_search(*args):
            search_text = self.search_var.get().lower()
            for widget in scrollable_frame.winfo_children():
                if isinstance(widget, tk.Checkbutton):
                    text = widget.cget('text').lower()
                    if search_text in text:
                        widget.pack(fill='x', anchor='w', pady=1)
                    else:
                        widget.pack_forget()
            # Update scrollregion after search
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        self.search_var.trace('w', update_search)
        
        # BUTTONS SECTION - Always visible at the bottom of main_frame
        separator = tk.Frame(main_frame, height=2, bg='#22505f')
        separator.pack(fill='x', pady=(10, 5))
        
        # Button frame - FIXED at bottom
        btn_frame = tk.Frame(main_frame, bg='#5b93a4', relief='raised', bd=2, height=80)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)  # Prevent resizing
        
        # Button container for centered layout
        btn_container = tk.Frame(btn_frame, bg='#5b93a4')
        btn_container.pack(expand=True, pady=15)
        
        # Buttons with improved styling and clear visibility
        apply_btn = tk.Button(btn_container, text="✅ Apply Filter", 
                             command=self.apply_filter,
                             bg='#003d52', fg='white', font=('Arial', 10, 'bold'),
                             width=15, height=2, relief='raised', bd=3,
                             activebackground='#255c7b')
        apply_btn.pack(side='left', padx=8)
        
        clear_btn = tk.Button(btn_container, text="🔄 Clear Filter",
                             command=lambda: self.clear_filter(column),
                             bg='#ef8827', fg='white', font=('Arial', 10, 'bold'),
                             width=15, height=2, relief='raised', bd=3,
                             activebackground='#22505f')
        clear_btn.pack(side='left', padx=8)
        
        # cancel_btn = tk.Button(btn_container, text="❌ Cancel",
        #                       command=self.filter_window.destroy,
        #                       bg='#F44336', fg='white', font=('Arial', 10, 'bold'),
        #                       width=12, height=2, relief='raised', bd=3,
        #                       activebackground='#D32F2F')
        # cancel_btn.pack(side='left', padx=8)
        
        # Add mouse wheel scrolling support
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        # Bind mousewheel to canvas and window
        canvas.bind("<MouseWheel>", on_mousewheel)
        self.filter_window.bind("<MouseWheel>", on_mousewheel)
        
        # Focus on search entry for immediate typing
        search_entry.focus()
    
    def toggle_all_checkboxes(self):
        """Toggle all filter checkboxes based on Select All state"""
        if hasattr(self, 'filter_checkboxes'):
            select_all = self.select_all_var.get()
            for var in self.filter_checkboxes.values():
                var.set(select_all)
    
    def apply_sort(self, column, ascending):
        """Apply sorting to the column"""
        try:
            if column in self.df.columns:
                # Convert to appropriate type for sorting
                if self.df[column].dtype == 'object':
                    # Try to convert to numeric if possible, otherwise sort as string
                    try:
                        sorted_df = self.df.iloc[self.df[column].astype(float).argsort()]
                        if not ascending:
                            sorted_df = sorted_df.iloc[::-1]
                    except:
                        sorted_df = self.df.sort_values(by=column, ascending=ascending, 
                                                       key=lambda x: x.astype(str).str.lower())
                else:
                    sorted_df = self.df.sort_values(by=column, ascending=ascending)
                
                self.df = sorted_df.reset_index(drop=True)
                # Update ID column after sorting
                if 'ID' in self.df.columns:
                    self.df['ID'] = range(1, len(self.df) + 1)
                
                self.render_table()
                self.update_sum_labels()
                self.update_role_summary()
                
                # Close filter window
                if hasattr(self, 'filter_window') and self.filter_window:
                    self.filter_window.destroy()
                    
                print(f"Sorted by {column}, ascending={ascending}")
        except Exception as e:
            print(f"Error sorting: {e}")
            import traceback
            traceback.print_exc()
    
    def apply_filter(self):
        """Apply the selected filter values"""
        try:
            if not hasattr(self, 'filter_checkboxes') or not hasattr(self, 'current_filter_column'):
                return
            
            column = self.current_filter_column
            
            # Get selected values
            selected_values = []
            for value, var in self.filter_checkboxes.items():
                if var.get():
                    selected_values.append(value)
            
            if not selected_values:
                # If nothing selected, show empty dataframe
                self.df = pd.DataFrame(columns=self.df.columns)
            else:
                # Filter the original data
                if hasattr(self, 'original_df') and column in self.original_df.columns:
                    # Start with original data (without Select and ID columns)
                    filtered_df = self.original_df[self.original_df[column].astype(str).isin(selected_values)].copy()
                    
                    # Add Select and ID columns
                    if not filtered_df.empty:
                        filtered_df.insert(0, 'Select', False)
                        filtered_df.insert(1, 'ID', range(1, len(filtered_df) + 1))
                    else:
                        # Create empty dataframe with all columns
                        filtered_df = pd.DataFrame(columns=self.df.columns)
                    
                    self.df = filtered_df
                else:
                    # Fallback to current dataframe
                    self.df = self.df[self.df[column].astype(str).isin(selected_values)].copy()
                    if 'ID' in self.df.columns and not self.df.empty:
                        self.df['ID'] = range(1, len(self.df) + 1)
            
            # Clear selections since row indices have changed
            self.selected_rows.clear()
            
            # Update display
            self.render_table()
            self.update_sum_labels()
            self.update_role_summary()
            
            # Close filter window
            if hasattr(self, 'filter_window') and self.filter_window:
                self.filter_window.destroy()
                
            print(f"Applied filter to {column}: {len(selected_values)} values selected")
            
        except Exception as e:
            print(f"Error applying filter: {e}")
            import traceback
            traceback.print_exc()
    
    def clear_filter(self, column):
        """Clear filter and restore original data"""
        try:
            # Restore original data
            if hasattr(self, 'original_df'):
                # Recreate the full dataframe with Select and ID columns
                display_cols_without_select_id = [col for col in self.df.columns if col not in ['Select', 'ID']]
                available_cols = [col for col in display_cols_without_select_id if col in self.original_df.columns]
                
                restored_df = self.original_df[available_cols].copy()
                
                # Add Select and ID columns
                if not restored_df.empty:
                    restored_df.insert(0, 'Select', False)
                    restored_df.insert(1, 'ID', range(1, len(restored_df) + 1))
                else:
                    restored_df = pd.DataFrame(columns=self.df.columns)
                
                self.df = restored_df
            
            # Clear selections
            self.selected_rows.clear()
            
            # Update display
            self.render_table()
            self.update_sum_labels()
            self.update_role_summary()
            
            # Close filter window
            if hasattr(self, 'filter_window') and self.filter_window:
                self.filter_window.destroy()
                
            print(f"Cleared filter for {column}")
            
        except Exception as e:
            print(f"Error clearing filter: {e}")
            import traceback
            traceback.print_exc()

    def create_summation_row(self):
        """Create a fixed summary row below the table in dedicated frame"""
        # Clear any existing widgets in the totals frame
        for widget in self.totals_display_frame.winfo_children():
            widget.destroy()
        
        # Create a HUGE IMPOSSIBLE TO MISS title
        title_label = tk.Label(self.totals_display_frame, 
                              text="� ATTENTION: TOTAL HOURS SUMMARY 🚨", 
                              font=('Arial', 20, 'bold'),
                              bg='#22505f',
                              fg='white')
        title_label.pack(pady=(15, 20))
        
        # Create styled containers for totals with MAXIMUM visibility
        totals_container = tk.Frame(self.totals_display_frame, bg='#22505f')
        totals_container.pack(expand=True, fill='both', padx=30)
        
        # Internal total box - GIGANTIC and impossible to miss
        internal_frame = tk.Frame(totals_container, bg='#5b93a4', relief='raised', bd=10, height=80)
        internal_frame.pack(side='left', padx=30, pady=15, expand=True, fill='both')
        
        internal_label = tk.Label(internal_frame, 
                                text="💼 TOTAL INTERNAL HOURS:", 
                                font=('Arial', 16, 'bold'),
                                bg='#5b93a4',
                                fg='white')
        internal_label.pack(side='left', padx=25, pady=20)
        
        self.total_label_internal = tk.Label(internal_frame,
                                           text="0.00",
                                           font=('Arial', 24, 'bold'),
                                           bg='#5b93a4',
                                           fg='white')
        self.total_label_internal.pack(side='left', padx=(15, 25), pady=20)
        
        # External total box - GIGANTIC and impossible to miss
        external_frame = tk.Frame(totals_container, bg='#003d52', relief='raised', bd=10, height=80)
        external_frame.pack(side='left', padx=30, pady=15, expand=True, fill='both')
        
        external_label = tk.Label(external_frame,
                                text="🏢 TOTAL EXTERNAL HOURS:",
                                font=('Arial', 16, 'bold'),
                                bg='#003d52',
                                fg='white')
        external_label.pack(side='left', padx=25, pady=20)
        
        self.total_label_external = tk.Label(external_frame,
                                           text="0.00", 
                                           font=('Arial', 24, 'bold'),
                                           bg='#003d52',
                                           fg='white')
        self.total_label_external.pack(side='left', padx=(15, 25), pady=20)

    def on_checkbox_click(self, event):
        """Handle checkbox column clicks for row selection"""
        item = self.tree.identify('item', event.x, event.y)
        column = self.tree.identify('column', event.x, event.y)
        
        if not item or not column.startswith('#'):
            return
            
        col_idx = int(column.replace('#', '')) - 1
        col_name = self.df.columns[col_idx]
        
        if col_name == 'Select':
            row_idx = int(item)
            values = list(self.tree.item(item)['values'])
            
            # Toggle selection
            if row_idx in self.selected_rows:
                self.selected_rows.remove(row_idx)
                values[0] = '☐'
            else:
                self.selected_rows.add(row_idx)
                values[0] = '☑'
                
            # Update tree display without full refresh
            self.tree.item(item, values=values)
            print(f"Selected rows: {len(self.selected_rows)}")

    def apply_dropdown_filters(self, event=None):
        """Apply dropdown filters and preserve any existing column filters"""
        self.apply_all_active_filters()

    def reset_filters(self):
        """Reset all filters and restore original project data"""
        # Clear all dropdown filter selections
        if hasattr(self, 'filter_vars'):
            for col, var in self.filter_vars.items():
                var.set("Todos")
        
        # Clear all column filters
        self.active_column_filters.clear()
        print("Cleared all filters")  # Debug
        
        # Clear row selections
        self.selected_rows.clear()
        
        # Reload the full project data from database
        if self.current_project:
            self.on_project_selected(None)  # This will reload all data for the current project
        else:
            # If no project selected, just clear everything
            self.df = pd.DataFrame(columns=self.display_columns)
            self.original_df = pd.DataFrame(columns=[col for col in self.display_columns if col not in ['Select', 'ID']])
            self.render_table()
            self.update_sum_labels()

    def update_filter_dropdowns(self):
        """Update filter dropdown values based on current data"""
        for idx, (col, var) in enumerate(self.filter_vars.items()):
            if col not in ["Select", "ID"] and idx < len(self.header_labels):
                for lbl, cmb in self.header_labels:
                    if hasattr(cmb, 'get'):
                        try:
                            # Update dropdown values with current data
                            if col in self.df.columns:
                                values = sorted(self.df[col].dropna().astype(str).unique())
                            else:
                                values = []
                            values = ["Todos"] + values
                            cmb['values'] = values
                            if var.get() not in values:
                                var.set("Todos")
                        except:
                            pass

    def sort_column(self, column, ascending=True):
        """Sort the dataframe by the specified column"""
        if column in self.df.columns and column not in ["Select", "ID"]:
            try:
                # Try to sort numerically first, fallback to string sort
                if column in ["Estimated internal", "Estimated external"]:
                    # For numeric columns, convert to numeric for proper sorting
                    self.df[column] = pd.to_numeric(self.df[column], errors='coerce')
                    self.df = self.df.sort_values(by=column, ascending=ascending, na_position='last')
                else:
                    # For text columns, sort as strings
                    self.df = self.df.sort_values(by=column, ascending=ascending, na_position='last')
                
                # Update ID counter after sorting
                if 'ID' in self.df.columns:
                    self.df['ID'] = range(1, len(self.df) + 1)
                
                self.render_table()
                self.update_sum_labels()
            except Exception as e:
                print(f"Error sorting column {column}: {e}")

    def edit_cell(self, event):
        """Handle cell editing on double-click"""
        item = self.tree.identify('item', event.x, event.y)
        column = self.tree.identify('column', event.x, event.y)
        
        if not item or not column or not column.startswith('#'):
            return
            
        # Get column name from index
        col_idx = int(column.replace('#', '')) - 1
        col_name = self.df.columns[col_idx]
        
        # Don't edit checkbox column
        if col_name == 'Select':
            return
            
        # Create and position the edit entry
        x, y, width, height = self.tree.bbox(item, column)
        entry = tk.Entry(self.tree, width=20)
        entry.insert(0, self.tree.item(item)['values'][col_idx])
        entry.select_range(0, tk.END)
        
        def on_edit_done(event=None):
            """Complete the edit and update data"""
            new_value = entry.get()
            row_idx = int(item)  # Get numerical index
            
            # Update DataFrame
            self.df.at[row_idx, col_name] = new_value
            
            # Update tree
            values = list(self.tree.item(item)['values'])
            values[col_idx] = new_value
            self.tree.item(item, values=values)
            
            entry.destroy()
            
            # Update summaries if needed
            if col_name in ["Estimated internal", "Estimated external"]:
                self.update_sum_labels()
                self.update_role_summary()
                
        entry.place(x=x, y=y, width=width, height=height)
        entry.focus_set()
        
        # Bind events
        entry.bind("<Return>", on_edit_done)
        entry.bind("<Escape>", lambda e: entry.destroy())
        entry.bind("<FocusOut>", on_edit_done)

    def clear_form(self):
        """Clear all form fields for easy data entry"""
        if messagebox.askyesno("Clear Form", "Are you sure you want to clear all form fields?"):
            for col, entry in self.entries.items():
                if hasattr(entry, 'set') and col != "Department":  # Keep Department as "FABSI"
                    entry.set("")
                elif hasattr(entry, 'delete') and col != "Department":
                    entry.delete(0, 'end')
                    
                # Reset Department to FABSI if it was cleared
                if col == "Department":
                    if hasattr(entry, 'insert'):
                        entry.insert(0, "FABSI")
                    elif hasattr(entry, 'set'):
                        entry.set("FABSI")

    def add_row(self):
        required_fields = ["Stick-Built", "Module", "Department", "Activities"]
        missing_fields = [col for col in required_fields if not self.entries[col].get().strip()]
        if not self.current_project:
            messagebox.showerror("Error", "Please select a project first.")
            return
        if missing_fields:
            messagebox.showerror("Error", f"Por favor llena los campos requeridos:\n{', '.join(missing_fields)}")
            return
        # Prepare data for DB insert
        import sqlalchemy
        db_path = os.path.join(os.path.dirname(__file__), 'Workload.db')
        logging.debug(f"Using DB file for insert: {os.path.abspath(db_path)}")
        engine = sqlalchemy.create_engine(f'sqlite:///{db_path}')
        # Get project_id
        with engine.connect() as conn:
            result = conn.execute(sqlalchemy.text('SELECT id FROM project WHERE name = :name'), {'name': self.current_project}).fetchone()
            project_id = result[0] if result else None
        if not project_id:
            messagebox.showerror("Error", "Project not found in database.")
            return
        # Build insert dict
        data = { 'project_id': project_id }
        # Map form fields to DB columns (assume DB columns match form fields, adjust if needed)
        # You may need to adjust these mappings if your DB uses different column names
        # Mapping from form field to DB column
        field_to_db = {
            "Stick-Built": "stick_built_id",
            "Module": "module_id",
            "Activities": "activities_id",
            "Title": "title_id",
            "Technical Unit": "technical_unit_id",
            "Assigned to": "employee_id",
            "Progress": "progress_id",
            "Professional Role": "professional_unit_id",
            "Department": "department",
            "Estimated internal": "estimated_internal_hours",
            "Estimated external": "estimated_external_hours",
            "Start date": "start_date",
            "Due date": "due_date",
            "Notes": "notes"
        }
        for col in self.entries:
            val = self.entries[col].get()
            db_col = field_to_db.get(col)
            if db_col is None:
                continue  # Skip fields not mapped
            # Foreign key fields (those ending with _id)
            if db_col.endswith('_id'):
                options = self.foreign_key_options.get(col, [])
                match = next((o for o in options if o['name'] == val), None)
                data[db_col] = match['id'] if match else None
            else:
                data[db_col] = val
        # Insert into DB
        try:
            columns = ', '.join(data.keys())
            placeholders = ', '.join([f':{k}' for k in data.keys()])
            insert_sql = f"INSERT INTO service ({columns}) VALUES ({placeholders})"
            with engine.begin() as conn:
                result = conn.execute(sqlalchemy.text(insert_sql), data)
                logging.info(f"Insert result rowcount: {result.rowcount}, data: {data}")
            messagebox.showinfo("Success", "Service added successfully.")
            
            # Store current filter state BEFORE reloading data
            current_dropdown_filters = {}
            current_column_filters = {}
            
            # Save dropdown filter state
            if hasattr(self, 'filter_vars'):
                current_dropdown_filters = {col: var.get() for col, var in self.filter_vars.items() 
                                          if hasattr(var, 'get') and var.get() != "Todos"}
            
            # Save column filter state (these are already stored in self.active_column_filters)
            current_column_filters = self.active_column_filters.copy()
            
            print(f"Saving current dropdown filters: {current_dropdown_filters}")  # Debug
            print(f"Saving current column filters: {current_column_filters}")  # Debug
            
            # DO NOT clear form fields - keep them for easier data entry
            # Commented out form clearing to allow faster consecutive entries
            # for col, entry in self.entries.items():
            #     if hasattr(entry, 'set') and col != "Department":  # Keep Department as "FABSI"
            #         entry.set("")
            #     elif hasattr(entry, 'delete') and col != "Department":
            #         entry.delete(0, 'end')
            #         if col == "Department":
            #             entry.insert(0, "FABSI")
            
            # Reload ALL project data first (this updates self.original_df)
            self.refreshing_data = True  # Set flag to prevent table rendering
            self.on_project_selected(None)
            self.refreshing_data = False  # Clear flag
            
            # Restore dropdown filters
            if hasattr(self, 'filter_vars'):
                for col, value in current_dropdown_filters.items():
                    if col in self.filter_vars:
                        self.filter_vars[col].set(value)
                        print(f"Restored dropdown filter {col}: {value}")  # Debug
            
            # Restore column filters
            self.active_column_filters = current_column_filters
            print(f"Restored column filters: {self.active_column_filters}")  # Debug
            
            # Apply all filters to show only matching records
            if current_dropdown_filters or current_column_filters:
                self.apply_all_active_filters()
                print("Applied all filters after adding record")  # Debug
            else:
                print("No active filters to restore")  # Debug
        except Exception as e:
            import traceback
            logging.error(f"Failed to add service: {e}")
            logging.error(traceback.format_exc())
            messagebox.showerror("Error", f"Failed to add service: {e}")

    def load_services(self):
        # Not used anymore, replaced by on_project_selected
        pass

    def delete_selected(self):
        if not self.selected_rows:
            messagebox.showwarning("Warning", "Please select at least one activity to delete.")
            return
        
        if messagebox.askyesno("Confirm Delete", f"Do you want to delete {len(self.selected_rows)} selected activities from the database?"):
            # Get the actual database IDs from the Document Number column
            ids_to_delete = []
            missing_ids = []
            
            print(f"Selected rows to delete: {sorted(self.selected_rows)}")  # Debug
            print(f"DataFrame columns: {list(self.df.columns)}")  # Debug
            
            for row_idx in self.selected_rows:
                if row_idx < len(self.df):
                    if 'Document Number' in self.df.columns:
                        db_id = self.df.iloc[row_idx]['Document Number']
                        if pd.notnull(db_id) and str(db_id).strip():  # Check for valid ID
                            try:
                                ids_to_delete.append(int(db_id))
                                print(f"Row {row_idx}: Found DB ID {db_id}")  # Debug
                            except (ValueError, TypeError):
                                print(f"Row {row_idx}: Invalid DB ID format: {db_id}")  # Debug
                                missing_ids.append(row_idx)
                        else:
                            print(f"Row {row_idx}: Empty/null DB ID")  # Debug
                            missing_ids.append(row_idx)
                    else:
                        print("Document Number column not found in DataFrame")  # Debug
                        missing_ids.append(row_idx)
                else:
                    print(f"Row index {row_idx} is out of bounds (DataFrame has {len(self.df)} rows)")  # Debug
                    missing_ids.append(row_idx)
            
            print(f"Valid IDs to delete: {ids_to_delete}")  # Debug
            print(f"Rows with missing/invalid IDs: {missing_ids}")  # Debug
            
            # Delete from database
            if ids_to_delete:
                try:
                    import sqlalchemy
                    db_path = os.path.join(os.path.dirname(__file__), 'Workload.db')
                    engine = sqlalchemy.create_engine(f'sqlite:///{db_path}')
                    
                    deleted_count = 0
                    failed_deletes = []
                    
                    with engine.begin() as conn:
                        for db_id in ids_to_delete:
                            try:
                                result = conn.execute(sqlalchemy.text("DELETE FROM service WHERE id = :id"), {'id': db_id})
                                if result.rowcount > 0:
                                    deleted_count += 1
                                    print(f"Successfully deleted service ID {db_id}")  # Debug
                                    logging.info(f"Deleted service ID {db_id}, rows affected: {result.rowcount}")
                                else:
                                    failed_deletes.append(db_id)
                                    print(f"No rows affected when deleting service ID {db_id}")  # Debug
                            except Exception as delete_error:
                                failed_deletes.append(db_id)
                                print(f"Failed to delete service ID {db_id}: {delete_error}")  # Debug
                                logging.error(f"Failed to delete service ID {db_id}: {delete_error}")
                    
                    # Clear selection after successful deletion
                    self.selected_rows.clear()
                    self.selected_rows.clear()
                    
                    # Store current filter state BEFORE reloading data
                    current_dropdown_filters = {}
                    current_column_filters = {}
                    
                    # Save dropdown filter state
                    if hasattr(self, 'filter_vars'):
                        current_dropdown_filters = {col: var.get() for col, var in self.filter_vars.items() 
                                                  if hasattr(var, 'get') and var.get() != "Todos"}
                    
                    # Save column filter state
                    current_column_filters = self.active_column_filters.copy()
                    
                    print(f"Saving current dropdown filters before delete: {current_dropdown_filters}")  # Debug
                    print(f"Saving current column filters before delete: {current_column_filters}")  # Debug
                    
                    # Reload data
                    self.refreshing_data = True  # Set flag to prevent table rendering
                    self.on_project_selected(None)
                    self.refreshing_data = False  # Clear flag
                    
                    # Restore dropdown filters
                    if hasattr(self, 'filter_vars'):
                        for col, value in current_dropdown_filters.items():
                            if col in self.filter_vars:
                                self.filter_vars[col].set(value)
                                print(f"Restored dropdown filter {col}: {value}")  # Debug
                    
                    # Restore column filters
                    self.active_column_filters = current_column_filters
                    print(f"Restored column filters: {self.active_column_filters}")  # Debug
                    
                    # Apply all filters
                    if current_dropdown_filters or current_column_filters:
                        self.apply_all_active_filters()
                        print("Applied all filters after delete")  # Debug
                    
                    # Show detailed success message
                    success_msg = f"Successfully deleted {deleted_count} activities from database."
                    if failed_deletes:
                        success_msg += f"\nFailed to delete {len(failed_deletes)} activities (IDs: {failed_deletes})."
                    if missing_ids:
                        success_msg += f"\n{len(missing_ids)} selected rows had no valid database IDs."
                    
                    messagebox.showinfo("Deletion Complete", success_msg)
                    logging.info(f"Successfully deleted {deleted_count} services from database. Failed: {len(failed_deletes)}")
                    
                except Exception as e:
                    logging.error(f"Failed to delete services: {e}")
                    import traceback
                    logging.error(traceback.format_exc())
                    messagebox.showerror("Error", f"Failed to delete activities from database: {e}")
            else:
                # Handle case where no valid IDs were found
                if missing_ids:
                    messagebox.showwarning("Warning", f"No valid database IDs found for the {len(missing_ids)} selected rows. Make sure the data was properly loaded from the database.")
                else:
                    messagebox.showwarning("Warning", "No valid activities found to delete.")

    def save_to_excel(self):
        if self.df.empty:
            messagebox.showerror("Error", "No selected data to save.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            self.df.to_excel(writer, index=False, sheet_name='List of Service')
            ws = writer.sheets['List of Service']
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = header_fill
                cell.border = border
                cell.font = Font(bold=True)
            for col in ws.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = min(max_length + 2, 25)
                for cell in col:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
            ws.auto_filter.ref = ws.dimensions
            ws.row_dimensions[1].height = 48
        subprocess.Popen(['start', '', path], shell=True)
        messagebox.showinfo("Guardado", f"Archivo guardado:\n{path}")

    def save_to_pdf(self):
        """Export table data to a well-formatted PDF file"""
        if self.df.empty:
            messagebox.showerror("Error", "No selected data to save.")
            return

        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import A3, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            # Ask for save location
            path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")]
            )
            if not path:
                return
                
            # Configure document
            doc = SimpleDocTemplate(
                path,
                pagesize=landscape(A3),  # Using A3 for more space
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Prepare data
            # Filter out unwanted columns and rename them
            export_columns = [col for col in self.df.columns if col not in ['Select', 'Document Number']]
            df_export = self.df[export_columns].copy()
            
            # Convert DataFrame to list of lists for the table
            header = [col for col in df_export.columns]  # Use column names directly
            data = [header]
            
            # Convert DataFrame values to strings and handle None/NaN
            for _, row in df_export.iterrows():
                row_data = []
                for value in row:
                    if pd.isna(value):
                        row_data.append("")
                    else:
                        row_data.append(str(value))
                data.append(row_data)
            
            # Calculate column widths based on content
            col_widths = []
            for idx in range(len(header)):
                max_width = max(
                    len(str(row[idx])) for row in data
                ) * 6  # 6 points per character
                col_widths.append(min(max_width, 150))  # Cap at 150 points for better fit
                
            # Create table
            table = Table(data, colWidths=col_widths, repeatRows=1)
            
            # Table style
            table_style = TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5b93a4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Cell style
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                
                # Text alignment
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            
            # Alternate row colors
            for row in range(1, len(data)):
                if row % 2 == 0:
                    table_style.add('BACKGROUND', (0, row), (-1, row), colors.HexColor('#255c7b'))
                else:
                    table_style.add('BACKGROUND', (0, row), (-1, row), colors.white)
            
            table.setStyle(table_style)
            
            # Build document
            elements = []
            
            # Add title
            title = Paragraph(f"FABSI - List of Service: {self.current_project}", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))  # Add some space after title
            
            # Add table
            elements.append(table)
            
            # Generate PDF
            doc.build(elements)
            
            # Open the generated PDF
            subprocess.Popen(['start', '', path], shell=True)
            messagebox.showinfo("Success", f"PDF file saved successfully:\n{path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create PDF: {str(e)}")
            logging.error(f"PDF creation error: {str(e)}")
            logging.error(traceback.format_exc())

if __name__ == "__main__":
    root = ctk.CTk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.geometry(f"{screen_width}x{screen_height}+0+0")
    root.resizable(True, True)
    app = ExcelActivityApp(root)
    root.mainloop()

